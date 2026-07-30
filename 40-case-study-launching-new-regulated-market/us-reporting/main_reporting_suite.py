# Companion code for "The Backend of Luck" - Chapter 40, Case Study.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

# =============================================================================
# Main Reporting Suite
# Source: Production casino platform (sanitized)
# Chapter 40 - Case Study
#
# Orchestrates the full US regulatory report generation pipeline:
#   1. Fetch data from each supplier (IGT, NetEnt, Evolution, Kambi)
#   2. Build partial CSV/Excel sheets per report type
#   3. Save raw CSVs for audit trail
#   4. Merge into final XLSX using the regulator's template
#   5. Run variance checks and alert via OpsGenie if discrepancies found
# =============================================================================

from __future__ import annotations

import csv
import io
import logging
import os
from datetime import datetime, time
from decimal import Decimal
from typing import Optional, Union
from zoneinfo import ZoneInfo

from models import FailedStep
from reporting_supplier import ReportingSupplier
from variance_check import Variance, check_variances
from wsr_reports import WsrReports

logger = logging.getLogger(__name__)


class MainReportingSuite:
    """
    Orchestrates the full US regulatory report generation pipeline.
    """

    def __init__(self, suppliers: list[ReportingSupplier]) -> None:
        self._suppliers = suppliers
        self._wsr = WsrReports()

    def run(
        self,
        db,
        reporting_date: datetime,
        casino_day_start: time,
        timezone: ZoneInfo,
        csv_output_dir: str,
        xlsx_output_path: str,
        xlsx_template_path: str,
        jurisdiction: str,
    ) -> Union[str, FailedStep]:
        """
        Execute the full pipeline. Returns "OK" on success or a FailedStep on error.
        """
        logger.info(
            "Generating report: date=%s jurisdiction=%s gaming_day_start=%s tz=%s",
            reporting_date.date(), jurisdiction, casino_day_start, timezone,
        )

        # Step 1 & 2: Fetch data and build sheets
        try:
            wsr_sheets = self._wsr.create(db, reporting_date, self._suppliers, xlsx_template_path)
        except Exception as exc:
            logger.error("Data fetching failed: %s", exc)
            return FailedStep("Data fetching", str(exc), exc)

        # Step 3: Save raw CSVs
        try:
            os.makedirs(csv_output_dir, exist_ok=True)
            for sheet_name, rows in wsr_sheets.items():
                csv_path = os.path.join(csv_output_dir, f"{sheet_name.replace(' ', '_')}.csv")
                with open(csv_path, "w", newline="", encoding="utf-8") as f:
                    writer = io.StringIO()
                    for row in rows:
                        f.write(",".join(str(v) for v in row.values()) + "\n")
            logger.info("Raw CSVs saved to %s", csv_output_dir)
        except Exception as exc:
            logger.error("Failed to save CSVs: %s", exc)
            return FailedStep("Saving CSV", str(exc), exc)

        # Step 4 & 5: Merge into XLSX and run variance checks
        try:
            self._save_xlsx(wsr_sheets, xlsx_output_path, xlsx_template_path, reporting_date)
        except Exception as exc:
            logger.error("Failed to save XLSX: %s", exc)
            return FailedStep("Saving XLSX", str(exc), exc)

        logger.info("Report generated successfully: %s", xlsx_output_path)
        return "OK"

    def _save_xlsx(
        self,
        sheets: dict,
        output_path: str,
        template_path: str,
        reporting_date: datetime,
    ) -> None:
        """
        Merge data sheets into the regulatory XLSX template, run variance
        checks, and write the final file.
        """
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(template_path)
        us_date = reporting_date.strftime("%m%d%Y")

        # Write data rows to their respective worksheets
        for sheet_name, rows in sheets.items():
            if sheet_name not in workbook.sheetnames:
                ws = workbook.create_sheet(sheet_name)
            else:
                ws = workbook[sheet_name]
            for row_dict in rows:
                ws.append(list(row_dict.values()))

        # Variance check
        variance_results = check_variances(workbook, us_date)
        if isinstance(variance_results, FailedStep):
            logger.error("Variance check failed: %s", variance_results.details)
            # Alert via OpsGenie in production — log for now
        elif isinstance(variance_results, list):
            for result in variance_results:
                if isinstance(result, Variance):
                    logger.warning(
                        "Variance detected: cell=%s date=%s value=%s",
                        result.cell, result.date, result.value,
                    )

        workbook.save(output_path)
        logger.info("XLSX saved: %s", output_path)
