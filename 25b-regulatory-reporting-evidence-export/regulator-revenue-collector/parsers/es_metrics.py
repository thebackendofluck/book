# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Spain DGOJ quarterly market data parser.

Source files (Alfresco CMIS, confirmed 1T-2026 / 2026-06-09):
  GGR CSV  — /cmis/document/alfresco/9dce6b70-db0b-4c8a-8377-eca50bc06fdb
  XLSX summary — /cmis/document/alfresco/76e38aff-7136-432c-a513-56f168b020e9

CSV schema (semicolon-delimited, Spanish locale numbers):
  Año | Trimestre | Mes | Juego | Cantidades jugadas | GGR

  - Values are in EUR (not millions) as integers with dot-thousands separators
    (e.g. "9.846.318" = 9 846 318 EUR).
  - ``Trimestre`` column contains labels like "2013.T1", "2025.T4".
  - ``Mes`` column contains full Spanish month names ("Enero", "Febrero", …).
  - One row per (month × game type); GGR can be negative (player-favoured month).

Vertical mapping (``Juego`` column → canonical vertical):
  Sports wagering
    - Ap. Dep. de Contrap. Convencionales  → sports-wagering
    - Ap. Dep. de Contrap. en Vivo         → sports-wagering
    - Ap. Cruzadas (exchange bets)         → sports-wagering
    - Carreras de Caballos                 → sports-wagering  (horse racing)
    - Otras Apuestas Deportivas            → sports-wagering

  iGaming (casino + bingo + concursos)
    - Black Jack                           → igaming
    - Ruleta                               → igaming
    - Bacará                               → igaming
    - Bingo                                → igaming
    - Concursos                            → igaming
    - Otros Juegos de Casino               → igaming
    - Máquinas de Azar (slots)             → igaming

  Online Poker (reported separately in the dashboard)
    - Póker (Torneos)                      → online-poker
    - Póker (Cash)                         → online-poker

  Any unrecognised ``Juego`` value is tagged as "igaming" (conservative default).

Period mapping:
  ``Año`` + ``Mes`` → "YYYY-MM" using the Spanish-to-number month table.
  ``Trimestre`` is ignored (redundant given Año + Mes).

FX conversion:
  EUR → USD cents:  int(round(eur_value * 1.08 * 100))
  No million scaling — CSV values are already in full EUR.

Expected fact count:
  ~11 game types × ~12 months × ~13 years (2013–2025) ≈ 1 700 rows in the
  CSV; after aggregating handle+GGR the parser emits ~1 700 MetricFact rows
  (handle + ggr separately → ~3 400 total).
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Optional

import openpyxl

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_EUR_TO_USD: float = 1.08

_MONTH_ES: dict[str, int] = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

# Mapping: lowercase substring of ``Juego`` column → canonical vertical.
# Order matters: most specific first.
_JUEGO_MAP: list[tuple[str, str]] = [
    # Sports wagering
    ("ap. dep.",           "sports-wagering"),
    ("ap. cruzada",        "sports-wagering"),
    ("apuesta",            "sports-wagering"),
    ("carrera",            "sports-wagering"),
    ("hípic",              "sports-wagering"),
    ("hipic",              "sports-wagering"),
    ("otras apuestas",     "sports-wagering"),
    # Online poker (before igaming to avoid mis-match on generic "juego")
    ("póker",              "online-poker"),
    ("poker",              "online-poker"),
    # iGaming / casino
    ("black jack",         "igaming"),
    ("blackjack",          "igaming"),
    ("ruleta",             "igaming"),
    ("bacará",             "igaming"),
    ("bacara",             "igaming"),
    ("bingo",              "igaming"),
    ("concurso",           "igaming"),
    ("máquina",            "igaming"),
    ("maquina",            "igaming"),
    ("casino",             "igaming"),
    ("slots",              "igaming"),
]

_DEFAULT_VERTICAL = "igaming"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _vertical(juego: str) -> str:
    """Map a Juego label to a canonical vertical."""
    lower = juego.lower()
    for fragment, vertical in _JUEGO_MAP:
        if fragment in lower:
            return vertical
    return _DEFAULT_VERTICAL


def _eur_to_usd_cents(eur: float) -> int:
    """Convert EUR to USD cents at fixed 1.08 rate."""
    return int(round(eur * _EUR_TO_USD * 100))


# Defensive ceiling: no single regulator-reported metric in any country has
# ever exceeded $100B for a single (state, vertical, period) point. Anything
# larger is a parser glitch (e.g. mis-identifying turnover as GGY, mis-reading
# thousand separators, or hitting a regional aggregate sheet that mixes units).
# Returning False here causes the calling code to skip the fact rather than
# pollute the dashboard.
_MAX_SINGLE_FACT_USD_CENTS = 100 * 1_000_000_000 * 100  # $100B in cents


def _value_is_sane(usd_cents: int) -> bool:
    return -_MAX_SINGLE_FACT_USD_CENTS <= usd_cents <= _MAX_SINGLE_FACT_USD_CENTS


def _parse_es_number(raw: str) -> Optional[float]:
    """Parse a Spanish-locale number to float.

    DGOJ CSV uses dot as thousands separator and no decimal separator for
    integer GGR values (e.g. "9.846.318").  Handles:
      - "9.846.318"  → 9846318.0
      - "-1.234.567" → -1234567.0
      - "0"          → 0.0
      - ""  or "-"   → None
    """
    s = raw.strip()
    if not s or s in ("-", "N/A", "n/a", "nd"):
        return None
    # Remove dot-thousands separators then swap comma decimal if present.
    # If both dot and comma present: European format (1.234,56).
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # comma as decimal separator (no thousands dot)
        s = s.replace(",", ".")
    else:
        # Only dots present → thousands separators; strip them.
        # Guard: if only one dot and <=2 digits after → decimal separator.
        dot_count = s.count(".")
        if dot_count == 1:
            integer_part, frac_part = s.rsplit(".", 1)
            if len(frac_part) <= 2 and len(integer_part) <= 4:
                # Looks like a decimal number (e.g. "62.6" or "9.5")
                pass  # leave as-is
            else:
                # Thousands separator (e.g. "9.846" means 9846)
                s = s.replace(".", "")
        else:
            # Multiple dots → all are thousands separators
            s = s.replace(".", "")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _month_num(mes: str) -> Optional[int]:
    """Convert Spanish month name to 1–12, or None if unrecognised."""
    return _MONTH_ES.get(mes.strip().lower())


def _period_from_row(ano: str, mes: str) -> Optional[str]:
    """Build "YYYY-MM" period string from Año and Mes columns."""
    try:
        year = int(ano.strip())
    except (ValueError, AttributeError):
        return None
    mo = _month_num(mes)
    if mo is None:
        return None
    return f"{year}-{mo:02d}"


# ---------------------------------------------------------------------------
# CSV parser (main data file)
# ---------------------------------------------------------------------------

def parse_es_csv(path: Path, source_url: str) -> list[MetricFact]:
    """Parse the DGOJ GGR/handle CSV and return MetricFact rows.

    The CSV is semicolon-delimited with Spanish locale numbers and contains
    monthly rows for all verticals from 2013 onward.

    Parameters
    ----------
    path:        Local path to the downloaded CSV file.
    source_url:  Original Alfresco URL recorded on each MetricFact.

    Returns
    -------
    list[MetricFact] with state="ES", operator="ES Statewide".
    Each row emits up to two facts: one for metric_name="handle" and one for
    metric_name="ggr".  Rows with zero values are skipped.
    """
    raw_bytes = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw_bytes.decode("latin-1", errors="replace")

    # Detect delimiter; DGOJ CSVs are semicolon-delimited.
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    # DGOJ publishes one row per (Juego, Mes, Año). Our 3-vertical mapping
    # collapses 17 Juego sub-products into {sports-wagering, igaming,
    # online-poker} — so multiple CSV rows land on the same dashboard key.
    # If we emitted them all individually the upsert would overwrite on
    # conflict of (state, op, vertical, period, metric) and only the last
    # sub-product would survive (under-counting ES by ~7×). Aggregate
    # handle / ggr per (period, vertical) BEFORE emitting.
    aggregated: dict[tuple[str, str], dict[str, float]] = {}
    for row in reader:
        ano = row.get("Año") or row.get("﻿Año") or row.get("Año") or ""
        mes = row.get("Mes") or ""
        juego = row.get("Juego") or ""
        raw_handle = row.get("Cantidades jugadas") or row.get("Cantidades Jugadas") or ""
        raw_ggr = row.get("GGR") or ""

        period = _period_from_row(ano, mes)
        if period is None:
            continue

        vertical = _vertical(juego)
        bucket = aggregated.setdefault((period, vertical), {"handle": 0.0, "ggr": 0.0})

        handle = _parse_es_number(raw_handle)
        ggr = _parse_es_number(raw_ggr)
        if handle is not None:
            bucket["handle"] += handle
        if ggr is not None:
            bucket["ggr"] += ggr

    facts: list[MetricFact] = []
    for (period, vertical), metrics in aggregated.items():
        for metric_name in ("handle", "ggr"):
            value = metrics[metric_name]
            if value == 0:
                continue
            v = _eur_to_usd_cents(value)
            if _value_is_sane(v):
                facts.append(MetricFact(
                    state="ES",
                    operator="ES Statewide",
                    vertical=vertical,
                    period=period,
                    metric_name=metric_name,
                    value_usd_cents=v,
                    source_url=source_url,
                ))

    return facts


# ---------------------------------------------------------------------------
# XLSX parser (summary / series workbook)
# ---------------------------------------------------------------------------

# Column header fragments for the XLSX summary workbook.
# The workbook has multiple sheets; we scan all of them.
_XLSX_HANDLE_HEADERS = re.compile(
    r"cantidades?\s+jugadas?|handle|mises?|wager", re.IGNORECASE
)
_XLSX_GGR_HEADERS = re.compile(
    r"\bggr\b|bruto\s+resultado|brutospel|gross\s+gam", re.IGNORECASE
)
_XLSX_JUEGO_HEADERS = re.compile(
    r"juego|game|tipo|type", re.IGNORECASE
)
_XLSX_PERIOD_HEADERS = re.compile(
    r"^(a[ñn]o|year|\d{4})", re.IGNORECASE
)


def parse_es_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse the DGOJ XLSX summary workbook and return MetricFact rows.

    The workbook ("Detalle datos del mercado español de juego") contains
    historical aggregates across sheets.  This parser uses a best-effort
    column-detection approach: locate period, game-type, and GGR columns
    by header name matching, then emit one fact per non-empty data cell.

    Falls back to an empty list gracefully if the sheet layout is unrecognised.

    Parameters
    ----------
    path:        Local path to the downloaded XLSX file.
    source_url:  Original Alfresco URL recorded on each MetricFact.

    Returns
    -------
    list[MetricFact] — may be empty if the workbook layout cannot be parsed.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return []

    facts: list[MetricFact] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the header row: the first row with at least 3 non-None cells
        # that contains recognisable column names.
        header_row_idx: Optional[int] = None
        header: list[str] = []
        for i, row in enumerate(rows):
            non_null = [str(c).strip() for c in row if c is not None]
            if len(non_null) >= 3:
                joined = " ".join(non_null).lower()
                if any(kw in joined for kw in ("ggr", "juego", "cantidades", "año", "ano")):
                    header_row_idx = i
                    header = [str(c).strip() if c is not None else "" for c in row]
                    break

        if header_row_idx is None:
            continue

        # Locate column indices.
        col_period: Optional[int] = None
        col_juego: Optional[int] = None
        col_ggr: Optional[int] = None
        col_handle: Optional[int] = None

        for idx, h in enumerate(header):
            if col_period is None and _XLSX_PERIOD_HEADERS.search(h):
                col_period = idx
            if col_juego is None and _XLSX_JUEGO_HEADERS.search(h):
                col_juego = idx
            if col_ggr is None and _XLSX_GGR_HEADERS.search(h):
                col_ggr = idx
            if col_handle is None and _XLSX_HANDLE_HEADERS.search(h):
                col_handle = idx

        # Need at least a period column and one value column to emit anything.
        if col_period is None or (col_ggr is None and col_handle is None):
            continue

        for row in rows[header_row_idx + 1:]:
            if all(c is None for c in row):
                continue

            # Period: expect an integer year or "YYYY-QN" style string.
            raw_period = row[col_period] if col_period < len(row) else None
            if raw_period is None:
                continue

            # Attempt period parsing: integer year → "YYYY-12" (annual).
            period: Optional[str] = None
            if isinstance(raw_period, (int, float)):
                year = int(raw_period)
                if 2000 <= year <= 2100:
                    period = f"{year}-12"
            elif isinstance(raw_period, str):
                # "2025.T4" → "2025-12", "2025-Q4" → "2025-12", "2025-01" kept
                m = re.match(r"(\d{4})[.\-_]?(?:[TQ](\d)|(\d{2}))?", raw_period.strip())
                if m:
                    year = int(m.group(1))
                    if m.group(2):
                        quarter = int(m.group(2))
                        # Map Q→end month: Q1→03, Q2→06, Q3→09, Q4→12
                        end_month = quarter * 3
                        period = f"{year}-{end_month:02d}"
                    elif m.group(3):
                        period = f"{year}-{m.group(3)}"
                    else:
                        period = f"{year}-12"

            if period is None:
                continue

            juego = str(row[col_juego]).strip() if col_juego is not None and col_juego < len(row) and row[col_juego] is not None else "Otros"
            vertical = _vertical(juego)

            if col_ggr is not None and col_ggr < len(row) and row[col_ggr] is not None:
                ggr = _parse_es_number(str(row[col_ggr]))
                if ggr is not None and ggr != 0:
                    v = _eur_to_usd_cents(ggr)
                    if _value_is_sane(v):
                        facts.append(MetricFact(
                            state="ES",
                            operator="ES Statewide",
                            vertical=vertical,
                            period=period,
                            metric_name="ggr",
                            value_usd_cents=v,
                            source_url=source_url,
                        ))

            if col_handle is not None and col_handle < len(row) and row[col_handle] is not None:
                handle = _parse_es_number(str(row[col_handle]))
                if handle is not None and handle != 0:
                    v = _eur_to_usd_cents(handle)
                    if _value_is_sane(v):
                        facts.append(MetricFact(
                            state="ES",
                            operator="ES Statewide",
                            vertical=vertical,
                            period=period,
                            metric_name="handle",
                            value_usd_cents=v,
                            source_url=source_url,
                        ))

    return facts
