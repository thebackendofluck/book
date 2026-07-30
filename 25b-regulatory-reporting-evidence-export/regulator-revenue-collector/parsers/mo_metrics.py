# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Missouri Gaming Commission monthly XLSX parsers.

Two file shapes are supported:

  Casino "WEB{MMYY}.xlsx" (Casino_Gaming/rb_financials)
    Tab "MONTHLY STATS" — per-boat blocks. Each block has 8 monthly rows
    (one per FYTD month) followed by a "TOTALS" row. Column layout:
      0  BOAT label (set on first month-row of each block; None on the rest)
      1  YEAR (datetime, e.g. 2025-07-01 → period 2025-07)
     10  TOTAL AGR (current year, USD)

    Tab "YTD TAXES" — wide matrix of Gaming Tax per property per month.
    Used to extract per-property monthly tax_paid.

  Sports wagering "SW Monthly Financials {MMYY}.xlsx"
    Tab "MONTHLY STATS RETAIL" and "MONTHLY STATS MOBILE":
      0  LICENSEE (set on first month-row; None on rest until next block)
      1  MO/YR (datetime)
      3  TOTAL HANDLE
      7  TOTAL TAXABLE AGR  (= ggr)
      8  SPORTS WAGERING TAX

Period for each emitted MetricFact is read from the `YEAR`/`MO/YR` cell of
its row, not from the URL — because each workbook spans the entire FYTD.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONEY_RE = re.compile(r"[\$,\s]")
_TOTAL_RE = re.compile(r"^\s*(totals?|fytd|grand\s+total|state\s+totals?)", re.I)


def _to_cents(raw) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        try:
            return int(round(float(raw) * 100))
        except (ValueError, OverflowError):
            return None
    s = _MONEY_RE.sub("", str(raw)).strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    if not s or s in ("-", "--", "N/A"):
        return None
    try:
        return int(round(float(s) * 100)) * (-1 if neg else 1)
    except (ValueError, InvalidOperation):
        return None


def _period(raw) -> str | None:
    """Extract YYYY-MM from a date cell or string."""
    if isinstance(raw, (datetime, date)):
        return f"{raw.year:04d}-{raw.month:02d}"
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.match(r"(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return None


def _is_pk(path: Path) -> bool:
    """Cheap guard against HTML/text masquerading as .xlsx."""
    try:
        with open(path, "rb") as fh:
            return fh.read(2) == b"PK"
    except OSError:
        return False


def _clean(label) -> str:
    return re.sub(r"\s+", " ", str(label or "")).strip()


# ---------------------------------------------------------------------------
# Casino: WEB workbook
# ---------------------------------------------------------------------------

# MONTHLY STATS column layout (0-indexed)
_CAS_COL_BOAT = 0
_CAS_COL_YEAR = 1
_CAS_COL_AGR = 10


def _parse_monthly_stats(ws, source_url: str) -> list[MetricFact]:
    out: list[MetricFact] = []
    current_boat: str | None = None
    in_data = False
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= _CAS_COL_AGR:
            continue
        col0 = _clean(row[_CAS_COL_BOAT])
        col1 = row[_CAS_COL_YEAR]

        # Detect header marker "BOAT" → enable data parsing on subsequent rows.
        if col0.upper() == "BOAT":
            in_data = True
            continue
        if not in_data:
            continue

        # New block starts with a non-empty boat name.
        if col0 and not _TOTAL_RE.match(col0):
            current_boat = col0
        elif _TOTAL_RE.match(col0):
            current_boat = None
            continue

        if not current_boat:
            continue

        period = _period(col1)
        if not period:
            continue

        agr_cents = _to_cents(row[_CAS_COL_AGR])
        if agr_cents is None:
            continue

        out.append(MetricFact(
            state="MO",
            operator=current_boat,
            vertical="commercial-casino",
            period=period,
            metric_name="ggr",
            value_usd_cents=agr_cents,
            source_url=source_url,
        ))
    return out


def _parse_ytd_taxes(ws, source_url: str) -> list[MetricFact]:
    """Pull the GAMING TAX matrix (per-property × per-month) for tax_paid."""
    out: list[MetricFact] = []
    rows = list(ws.iter_rows(values_only=True))

    # Locate the GAMING TAX section by header row.
    tax_header_idx: int | None = None
    for i, row in enumerate(rows):
        if not row:
            continue
        first = _clean(row[0]).upper()
        # GAMING TAX section header — followed by a MONTH row a few rows below.
        if first == "GAMING TAX":
            tax_header_idx = i
            break

    if tax_header_idx is None:
        return out

    # Within the GAMING TAX section, find the MONTH header row (column labels).
    month_row_idx: int | None = None
    for j in range(tax_header_idx + 1, min(tax_header_idx + 6, len(rows))):
        if rows[j] and _clean(rows[j][0]).upper() == "MONTH":
            month_row_idx = j
            break
    if month_row_idx is None:
        return out

    headers = [_clean(c) for c in rows[month_row_idx]]
    # Walk data rows below the MONTH header until a TOTALS row or empty gap.
    for k in range(month_row_idx + 1, len(rows)):
        r = rows[k]
        if not r:
            continue
        first = _clean(r[0])
        if not first:
            continue
        if _TOTAL_RE.match(first):
            break
        period = _period(r[0])
        if not period:
            # Not a date row; skip (could be a sub-header).
            continue
        for col_i in range(1, min(len(r), len(headers))):
            op = headers[col_i]
            if not op:
                continue
            # Skip aggregate/total columns (e.g. "TOTAL", "STATE TOTAL") —
            # they duplicate the per-operator rows and inflate tax_paid ~2x.
            if _TOTAL_RE.match(op):
                continue
            cents = _to_cents(r[col_i])
            if cents is None:
                continue
            out.append(MetricFact(
                state="MO",
                operator=op,
                vertical="commercial-casino",
                period=period,
                metric_name="tax_paid",
                value_usd_cents=cents,
                source_url=source_url,
            ))
    return out


def parse_mo_casino_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse a Missouri WEB{MMYY}.xlsx and emit per-boat ggr + tax_paid facts.

    Each WEB workbook covers the full fiscal-YTD (8 months for FY mid-year),
    so a single download backfills several months at once.
    """
    if not _is_pk(path):
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    facts: list[MetricFact] = []
    try:
        for name in wb.sheetnames:
            up = name.strip().upper()
            ws = wb[name]
            if up == "MONTHLY STATS":
                facts.extend(_parse_monthly_stats(ws, source_url))
            elif up == "YTD TAXES":
                facts.extend(_parse_ytd_taxes(ws, source_url))
    finally:
        wb.close()
    facts.sort(key=lambda f: (f.period, f.operator, f.metric_name))
    return facts


# ---------------------------------------------------------------------------
# Sports wagering: SW Monthly Financials workbook
# ---------------------------------------------------------------------------

# MONTHLY STATS RETAIL/MOBILE column layout
_SW_COL_LIC = 0
_SW_COL_DATE = 1
_SW_COL_HANDLE = 3
_SW_COL_AGR = 7
_SW_COL_TAX = 8


def _parse_sw_block(ws, channel: str, source_url: str) -> list[MetricFact]:
    """Walk a MONTHLY STATS RETAIL/MOBILE sheet → MetricFacts."""
    out: list[MetricFact] = []
    current_lic: str | None = None
    in_data = False
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= _SW_COL_TAX:
            continue
        col0 = _clean(row[_SW_COL_LIC])

        # Header detection: row whose first cell is "LICENSEE".
        if col0.upper() == "LICENSEE":
            in_data = True
            continue
        if not in_data:
            continue

        if col0 and not _TOTAL_RE.match(col0):
            current_lic = col0
        elif _TOTAL_RE.match(col0):
            current_lic = None
            continue

        if not current_lic:
            continue

        period = _period(row[_SW_COL_DATE])
        if not period:
            continue

        # Strip the trailing channel suffix sometimes baked into the licensee
        # label so that "ARGOSY CASINO - RETAIL" and an analogous mobile row
        # collapse onto a single operator name. Keep the raw label otherwise.
        op = re.sub(r"\s*[-–]\s*(retail|mobile)\s*$", "", current_lic, flags=re.I)

        for metric, col in (
            ("handle", _SW_COL_HANDLE),
            ("ggr", _SW_COL_AGR),
            ("tax_paid", _SW_COL_TAX),
        ):
            cents = _to_cents(row[col])
            if cents is None:
                continue
            out.append(MetricFact(
                state="MO",
                operator=f"{op} ({channel})",
                vertical="sports-wagering",
                period=period,
                metric_name=metric,
                value_usd_cents=cents,
                source_url=source_url,
            ))
    return out


def parse_mo_sw_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse a Missouri 'SW Monthly Financials {MMYY}.xlsx' file."""
    if not _is_pk(path):
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    facts: list[MetricFact] = []
    try:
        for name in wb.sheetnames:
            up = name.strip().upper()
            ws = wb[name]
            if "RETAIL" in up and "MONTHLY" in up:
                facts.extend(_parse_sw_block(ws, "retail", source_url))
            elif "MOBILE" in up and "MONTHLY" in up:
                facts.extend(_parse_sw_block(ws, "mobile", source_url))
    finally:
        wb.close()
    facts.sort(key=lambda f: (f.period, f.operator, f.metric_name))
    return facts


# ---------------------------------------------------------------------------
# Public dispatcher
# ---------------------------------------------------------------------------

def parse_mo_xlsx(path: Path, vertical: str, source_url: str) -> list[MetricFact]:
    """Entry point used by backfill — routes by vertical."""
    if vertical == "sports-wagering":
        return parse_mo_sw_xlsx(path, source_url)
    return parse_mo_casino_xlsx(path, source_url)
