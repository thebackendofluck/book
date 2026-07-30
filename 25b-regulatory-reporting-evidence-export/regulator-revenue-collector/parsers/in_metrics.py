# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Indiana Gaming Commission combined monthly revenue parser (XLSX and PDF).

One file per month covers both casino and sports-wagering.
Casino sheet/section: rows = property names; cols include Win, Wagering Tax,
  Supplemental Tax (sometimes merged as a single Tax column).
Sports sheet/section: rows = operator names; cols include Handle, Adjusted
  Gross Revenue, Tax (9.5 %).

IGC switched from XLSX to PDF in late 2024.  This module handles both:
  XLSX (openpyxl): pre-2024 files at  /igc/files/reports/{YYYY}/{YYYY}-{MM}-Revenue.xlsx
  PDF (pdfplumber): 2024+ files at     /igc/files/{YYYY}-{MM}-Revenue.pdf
                                  or   /igc/files/reports/{YYYY}/{YYYY}-{MM}-Revenue.pdf

The IGC monthly PDF (page 0) contains three distinct sections on one page:
  1. TOTAL TAX: columns = [Name] [Location] [Suppl Tax] [Sports Wag Tax]
                           [Wagering Tax] [Total Tax]
                → parse Wagering Tax column as tax_paid per operator
  2. Win section: columns = [Name] [Win] [Free Play] [Other*] [Taxable AGR]
                → not used directly; AGR is confirmed in section 3
  3. WAGERING TAX: columns = [Name] [No.Tables] [Table Win] [No.EGD]
                              [EGD/Slot Win] [AGR]
                → parse AGR (rightmost) column as ggr per operator

Page 1 is the YTD cumulative summary — it must be EXCLUDED to avoid
inflating monthly figures with year-to-date totals.

Sample URLs:
  https://www.in.gov/igc/files/reports/2026/2026-03-Revenue.pdf
  https://www.in.gov/igc/files/reports/2025/2025-01-Revenue.pdf
  https://www.in.gov/igc/files/reports/2024/2024-01-Revenue.pdf
"""
from __future__ import annotations

import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONEY_RE = re.compile(r"[\$,\s]")
_SKIP_RE = re.compile(
    r"^\s*(total|subtotal|grand\s+total|state\s+total|#|\bno\.\b|footnote|%|ytd|\*+)",
    re.I,
)

# Sheet-name patterns
_CASINO_SHEET_RE = re.compile(r"casino|riverboat|racino|gaming", re.I)
_SPORTS_SHEET_RE = re.compile(r"sport|wagering|mobile", re.I)

# Column-header patterns for casino
_WIN_RE = re.compile(r"\bwin\b", re.I)
_WAG_TAX_RE = re.compile(r"wagering\s+tax|wager\s+tax", re.I)
_SUPP_TAX_RE = re.compile(r"supplement", re.I)
_TAX_RE = re.compile(r"\btax\b", re.I)
_AGR_HDR_RE = re.compile(r"\bagr\b", re.I)            # standalone AGR column header
_TOTAL_TAX_HDR_RE = re.compile(r"total\s+tax", re.I)  # page-0 section heading

# Column-header patterns for sports
_HANDLE_RE = re.compile(r"\bhandle\b|\bwagers?\b", re.I)
_AGR_RE = re.compile(r"adjusted\s+gross|adj\.?\s*gross|\bagr\b", re.I)

# Page-level skip patterns — pages whose heading indicates YTD / cumulative data
_YTD_PAGE_RE = re.compile(r"ytd\s+summary|year\s+to\s+date", re.I)
# Sports "Detail" page has per-sportsbook breakdown; not a monthly summary table
_SPORTS_DETAIL_PAGE_RE = re.compile(r"detail\s+of\s+sports\s+wagering", re.I)

# "Monthly Sports Wagering Handle by Sport" / "State Wide Handle by Sport"
# sub-table sentinel — row text that marks the start of the sport-category
# breakdown section at the bottom of the sports summary page.  This section
# uses sport names as row labels and has a Month / YTD column layout; it must
# NEVER be parsed as operator rows.
#
# IGC used two different headings across the report history:
#   2023-07 – 2026-02:  "State Wide Handle by Sport"   (or "State Wide Handle by Sport Month YTD")
#   2026-03+:           "Monthly Sports Wagering Handle by Sport"
# Both are matched by this pattern.
_SPORT_SUBTABLE_RE = re.compile(
    r"monthly\s+sports\s+wagering|handle\s+by\s+sport|state\s+wide\s+handle",
    re.I,
)

# Explicit set of sport-category names that are NEVER valid operator labels.
# Acts as a belt-and-suspenders guard in every sports and casino row emitter.
# Names are stored in canonical (title-case) form; comparison is case-insensitive.
_SPORT_CATEGORY_NAMES: frozenset[str] = frozenset({
    "football", "basketball", "baseball", "parlay", "other", "hockey",
    "soccer", "tennis", "golf", "boxing", "mma", "auto racing",
    "table tennis", "cricket", "esports", "combined",
    "state wide handle by sport", "month", "month ytd",
})


def _is_sport_category(label: str) -> bool:
    """Return True if *label* is a sport-category name, not an operator."""
    return label.lower().strip() in _SPORT_CATEGORY_NAMES


# Known casino properties (lowercase) — used to validate operator rows
_CASINO_PROPERTIES = {
    "ameristar", "belterra", "blue chip", "caesars southern indiana",
    "french lick", "hard rock northern indiana", "harrah's hoosier park",
    "hollywood", "horseshoe hammond", "horseshoe indianapolis",
    "majestic star", "rising star", "tropicana",
}


def _period_from_url(url: str) -> str | None:
    m = re.search(r"(\d{4})-(\d{1,2})-Revenue", url, re.I)
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else None


def _to_cents(raw) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return int(round(float(raw) * 100))
    s = _MONEY_RE.sub("", str(raw)).strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    if not s or s in ("-", "--"):
        return None
    try:
        return int(round(float(s) * 100)) * (-1 if neg else 1)
    except ValueError:
        return None


def _header_str(cell) -> str:
    return re.sub(r"\s+", " ", str(cell or "").replace("\n", " ")).strip()


def _is_skip_row(label: str) -> bool:
    return bool(_SKIP_RE.match(label))


# ---------------------------------------------------------------------------
# Casino sheet parser (XLSX)
# ---------------------------------------------------------------------------
#
# IGC XLSX structure (Sheet1 / first sheet): one sheet covers the same three
# sections as PDF page 0.
#
# Section 1 — TOTAL TAX (header: "TOTAL TAX | Location | Supplemental Tax |
#   Sports Wagering Tax | Wagering Tax | Total Tax"):
#   → col index 4 = Wagering Tax  → emit as tax_paid
#
# Section 3 — WAGERING TAX (header: "WAGERING TAX | No. of Table Games |
#   Table Win | No. of EGD/Slots | EGD/Slot Win | AGR"):
#   → last non-None col = AGR  → emit as ggr
#
# Both sections appear on the same worksheet (no sheet-name filter needed).
# ---------------------------------------------------------------------------

_SECTION_TOTAL_TAX = "total_tax"
_SECTION_WAG_TAX = "wagering_tax"
_SECTION_WIN = "win_section"
_SECTION_OTHER = "other"


def _classify_header_row(cells: list[str]) -> str:
    """Determine which section a header row introduces."""
    merged = " ".join(cells).lower()
    # "TOTAL TAX" section has both a "wagering tax" sub-header AND "total tax"
    if _TOTAL_TAX_HDR_RE.search(merged) and _WAG_TAX_RE.search(merged):
        return _SECTION_TOTAL_TAX
    # "WAGERING TAX" breakdown section — has "EGD" and "AGR" (no handle)
    if re.search(r"\begd\b|\begd/slot", merged, re.I) and _AGR_HDR_RE.search(merged):
        return _SECTION_WAG_TAX
    # Win / Taxable AGR section — has "free play" and "taxable agr"
    if _WIN_RE.search(merged) and re.search(r"free\s+play|taxable", merged, re.I):
        return _SECTION_WIN
    return _SECTION_OTHER


def _parse_casino_sheet(ws, period: str, source_url: str) -> list[MetricFact]:
    """Extract GGR (AGR) and Wagering Tax for each IN casino property.

    Reads two sections from the combined monthly sheet:
    - TOTAL TAX: emits tax_paid from the Wagering Tax column.
    - WAGERING TAX: emits ggr from the AGR (last) column.
    Skips TOTAL/YTD rows and the Win/Free-Play intermediate section.
    """
    out: list[MetricFact] = []

    # Accumulators per section
    col_wag_tax_idx: int = -1     # column index for Wagering Tax (TOTAL TAX section)
    col_agr_idx: int = -1         # column index for AGR (WAGERING TAX section)
    current_section = _SECTION_OTHER

    for row in ws.iter_rows(values_only=True):
        if not any(c not in (None, "") for c in row):
            continue

        label = _header_str(row[0]) if row[0] is not None else ""
        cells = [_header_str(c) for c in row]

        # Detect section transitions by header rows
        section = _classify_header_row(cells)
        if section != _SECTION_OTHER:
            current_section = section
            if current_section == _SECTION_TOTAL_TAX:
                # Find "Wagering Tax" column (not "Sports Wagering Tax")
                col_wag_tax_idx = -1
                for i, h in enumerate(cells):
                    # Match "Wagering Tax" but not "Sports Wagering Tax"
                    if _WAG_TAX_RE.search(h) and not re.search(r"sports", h, re.I):
                        col_wag_tax_idx = i
                        # keep scanning — take the last match (rightmost pure wag tax)
                        # Actually we want the one that isn't "Sports Wagering Tax"
                # If still -1, fall back to any wagering tax column
                if col_wag_tax_idx == -1:
                    for i, h in enumerate(cells):
                        if _WAG_TAX_RE.search(h):
                            col_wag_tax_idx = i
            elif current_section == _SECTION_WAG_TAX:
                # AGR is the last meaningful column header
                col_agr_idx = -1
                for i, h in enumerate(cells):
                    if _AGR_HDR_RE.search(h):
                        col_agr_idx = i
            continue

        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue

        if current_section == _SECTION_TOTAL_TAX and col_wag_tax_idx >= 0:
            cents = _to_cents(row[col_wag_tax_idx] if col_wag_tax_idx < len(row) else None)
            if cents is not None:
                out.append(MetricFact(
                    state="IN", operator=label,
                    vertical="commercial-casino", period=period,
                    metric_name="tax_paid", value_usd_cents=cents,
                    source_url=source_url,
                ))

        elif current_section == _SECTION_WAG_TAX and col_agr_idx >= 0:
            cents = _to_cents(row[col_agr_idx] if col_agr_idx < len(row) else None)
            if cents is not None:
                out.append(MetricFact(
                    state="IN", operator=label,
                    vertical="commercial-casino", period=period,
                    metric_name="ggr", value_usd_cents=cents,
                    source_url=source_url,
                ))

    return out


# ---------------------------------------------------------------------------
# Sports sheet parser
# ---------------------------------------------------------------------------

def _parse_sports_sheet(ws, period: str, source_url: str) -> list[MetricFact]:
    """Extract Handle, AGR (ggr), and Tax for each IN sports-wagering operator.

    Guards against the "State Wide Handle by Sport" / "Monthly Sports Wagering
    Handle by Sport" sub-table that appears at the bottom of the IGC Sheet7.
    That section uses sport-category names (Football, Basketball, …) as row
    labels and has Month / YTD columns — it must never be emitted as operator
    facts.  Two defences are applied:
      1. Sub-table header stop: as soon as a row whose first cell matches
         _SPORT_SUBTABLE_RE is encountered, iteration stops.
      2. Explicit name guard: _is_sport_category() rejects any sport-category
         label that slips through (e.g. if the header row is missing or the
         column layout changes).
    """
    out: list[MetricFact] = []
    col_handle: list[int] = []
    col_agr: list[int] = []
    col_tax: list[int] = []
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not any(c not in (None, "") for c in row):
            continue

        label = _header_str(row[0]) if row[0] is not None else ""
        cells = [_header_str(c) for c in row]

        # Stop at the sport-category sub-table.  The sub-table header row
        # contains "State Wide Handle by Sport" or "Monthly Sports Wagering
        # Handle by Sport" in the first cell.
        if header_found and label and _SPORT_SUBTABLE_RE.search(label):
            break

        if not header_found:
            if any(_HANDLE_RE.search(h) for h in cells):
                for i, h in enumerate(cells):
                    if _HANDLE_RE.search(h):
                        col_handle.append(i)
                    elif _AGR_RE.search(h):
                        col_agr.append(i)
                    elif _TAX_RE.search(h):
                        col_tax.append(i)
                header_found = True
            continue

        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue

        for ci in col_handle:
            if ci < len(row):
                cents = _to_cents(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="handle", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_agr:
            if ci < len(row):
                cents = _to_cents(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="ggr", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_tax:
            if ci < len(row):
                cents = _to_cents(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="tax_paid", value_usd_cents=cents,
                        source_url=source_url,
                    ))

    return out


# ---------------------------------------------------------------------------
# PDF helpers — shared numeric parsing
# ---------------------------------------------------------------------------

_MONEY_PDF_RE = re.compile(r"[\$,\s\xa0]")
_NEG_PDF_RE = re.compile(r"^\((-?[\d.]+)\)$")

# Numeric-looking token: optional $ or (, digits/commas/dot, optional ) or %
_VALUE_TOKEN_RE = re.compile(r"^\(?-?\$?[\d,]+(?:\.\d+)?\)?$")


def _to_cents_pdf(raw: object) -> int | None:
    """Parse a cell value (string or number) into integer USD cents."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(round(float(raw) * 100))
    s = str(raw).strip()
    if s in ("", "-", "--", "N/A", "n/a"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = _MONEY_PDF_RE.sub("", s)
    if not s:
        return None
    try:
        cents = int((Decimal(s) * 100).quantize(Decimal("1")))
    except InvalidOperation:
        return None
    return -cents if neg else cents


def _header_cells(row: list) -> list[str]:
    return [re.sub(r"\s+", " ", str(c or "").replace("\n", " ")).strip() for c in row]


# ---------------------------------------------------------------------------
# Word-position clustering helpers (Indiana standard-layout tables)
#
# Indiana PDFs use a standard table layout: one row per operator/casino,
# columns are the metrics (Win, Wagering Tax, Handle, Adjusted Gross
# Revenue, etc.). pdfplumber extract_tables() fails on these because they
# have no visible ruling lines (borderless tables).  The approach mirrors
# ia_metrics.py but operates on column-oriented data rather than Iowa's
# transposed layout.
#
# Algorithm:
#   1. extract_words() with x/y tolerances
#   2. Cluster words by top y-coordinate into visual rows
#   3. Find the first row that contains known column-header keywords
#   4. Record the x-anchor (left edge) of each header word cluster
#   5. For subsequent data rows: leftmost non-numeric words = operator label;
#      numeric words to the right = values; each value assigned to the
#      column whose anchor is closest (≤ value's right edge)
# ---------------------------------------------------------------------------

_Y_TOL = 3   # pixels — words within this y-range are on the same line
_X_GAP = 8   # pixels — gap larger than this separates adjacent column anchors


def _group_words_into_rows(words: list[dict], y_tol: int = _Y_TOL) -> list[list[dict]]:
    """Cluster word dicts by ``top`` y-coordinate into visual text rows."""
    if not words:
        return []
    by_top: dict[int, list[dict]] = defaultdict(list)
    for w in words:
        bucket = round(w["top"] / y_tol) * y_tol
        by_top[bucket].append(w)
    rows = []
    for k in sorted(by_top.keys()):
        rows.append(sorted(by_top[k], key=lambda w: w["x0"]))
    return rows


def _merge_word_row(row: list[dict]) -> str:
    """Concatenate all word text on the row into a single string."""
    return " ".join(w["text"] for w in row)


def _find_header_row_idx(
    rows: list[list[dict]],
    col_detector,
) -> int:
    """Return index of first row whose merged text matches col_detector.

    col_detector(merged_text) → True if the row looks like a header.
    Returns -1 if not found.
    """
    for i, row in enumerate(rows):
        merged = _merge_word_row(row)
        if col_detector(merged):
            return i
    return -1


def _build_column_anchors(header_row: list[dict]) -> list[tuple[float, float, str]]:
    """Infer column anchors from a header row of word dicts.

    Multi-word column names (e.g. "Wagering Tax", "Adjusted Gross Revenue")
    appear as consecutive words with small gaps.  Words separated by a gap
    larger than _X_GAP are in different columns.

    Returns list of (x0, x1, label) tuples for each column cluster,
    sorted by x0 ascending.
    """
    if not header_row:
        return []
    clusters: list[list[dict]] = []
    current: list[dict] = [header_row[0]]
    for w in header_row[1:]:
        if w["x0"] - current[-1]["x1"] <= _X_GAP:
            current.append(w)
        else:
            clusters.append(current)
            current = [w]
    clusters.append(current)
    result = []
    for c in clusters:
        x0 = c[0]["x0"]
        x1 = c[-1]["x1"]
        label = " ".join(w["text"] for w in c)
        result.append((x0, x1, label))
    return result


def _assign_to_column(
    word: dict, anchors: list[tuple[float, float, str]]
) -> int:
    """Return index of the column anchor best matching this word's position.

    Strategy: pick the anchor whose x0 is the largest one that is still
    ≤ word's right edge (x1). Values are right-aligned, so their x1 sits
    at or just past the column's right boundary, while x0 of the anchor
    is always to the left of the value.
    """
    best_i = -1
    best_x0 = -1.0
    for i, (ax0, _ax1, _lbl) in enumerate(anchors):
        if ax0 <= word["x1"] and ax0 > best_x0:
            best_x0 = ax0
            best_i = i
    return best_i


# ---------------------------------------------------------------------------
# Word-cluster casino page parser — page 0 of the IGC monthly PDF
# ---------------------------------------------------------------------------
#
# Page 0 layout (three stacked sections, all operator rows):
#
#   Section A — "TOTAL TAX" section
#     Header:  TOTAL TAX | Location | Supplemental Tax | Sports Wagering Tax
#              | Wagering Tax | Total Tax
#     Data:    one row per casino; we extract the Wagering Tax column
#              (second-rightmost value) → metric_name='tax_paid'
#
#   Section B — Win / Free-Play / Taxable AGR section (IGNORED as GGR source)
#     Header:  Win | Free Play | Other * | Taxable AGR
#
#   Section C — "WAGERING TAX" breakdown section
#     Header:  WAGERING TAX | No. of Table Games | Table Win | No. of EGD/Slots
#              | EGD/Slot Win | AGR
#     Data:    one row per casino; we extract the AGR column (rightmost value)
#              → metric_name='ggr'
#
# The column assignment uses x-position: the rightmost value is always the
# "Total Tax" / "AGR" column, and the second-rightmost is "Wagering Tax" /
# "EGD/Slot Win".  We use _assign_to_column() with the anchors from the
# actual header row to avoid hard-coded pixel thresholds.
# ---------------------------------------------------------------------------


def _parse_casino_page_words(page, period: str, source_url: str) -> list[MetricFact]:
    """Word-position clustering for the casino summary page (PDF page 0).

    Uses a state-machine to walk rows top-to-bottom and switch parsing
    mode when a section header is recognised.  Reads only the two
    monthly sections — never the Win/Free-Play intermediate section:

      State A — TOTAL TAX section
        Header: "TOTAL TAX ... Wagering Tax Total Tax"
        Data:   one row per casino; emit Wagering Tax column → tax_paid
        Stop:   on next section header

      State B — WAGERING TAX breakdown section
        Header: "WAGERING TAX ... EGD/Slot Win AGR"
        Data:   one row per casino; emit AGR (rightmost) column → ggr
        Stop:   on next section header or end of rows

      State SKIP — Win/Free-Play section and all other sections
        Entered when a non-A/non-B header row is detected.

    Operator labels:
      The TOTAL TAX section has a "Location" column immediately after the
      operator name.  We strip any trailing capitalised city-name token by
      using the Location column x0 as the right boundary of the label.
      For the WAGERING TAX section no location column is present; the
      label boundary is determined by the first numeric anchor.
    """
    words = page.extract_words(
        x_tolerance=3, y_tolerance=3,
        keep_blank_chars=False, use_text_flow=True,
    )
    rows = _group_words_into_rows(words)
    if not rows:
        return []

    # ---- Section-header classifiers ----

    def _is_total_tax_hdr(text: str) -> bool:
        """'TOTAL TAX Location ... Wagering Tax Total Tax'."""
        return bool(_TOTAL_TAX_HDR_RE.search(text)) and bool(_WAG_TAX_RE.search(text))

    def _is_wagering_tax_hdr(text: str) -> bool:
        """'WAGERING TAX No. of Table Games ... EGD/Slot Win AGR'."""
        return (
            bool(_WAG_TAX_RE.search(text))
            and bool(re.search(r"\begd\b|\begd/slot", text, re.I))
            and bool(_AGR_HDR_RE.search(text))
        )

    def _is_any_section_hdr(text: str) -> bool:
        """True for any section header that should switch mode."""
        return (
            _is_total_tax_hdr(text)
            or _is_wagering_tax_hdr(text)
            or bool(re.search(r"\bwin\b.{1,30}free\s+play|free\s+play.{1,30}\bwin\b|taxable\s+agr", text, re.I))
        )

    out: list[MetricFact] = []

    # State machine variables
    _ST_SKIP = 0
    _ST_TOTAL_TAX = 1   # emitting tax_paid
    _ST_WAG_TAX = 2     # emitting ggr

    state = _ST_SKIP
    anchors: list[tuple[float, float, str]] = []
    label_x_boundary = 0.0
    col_target: list[int] = []    # column indices to read in current state
    metric_name = ""

    for row in rows:
        if not row:
            continue
        merged = _merge_word_row(row)

        # ---- Check for section header ----
        if _is_total_tax_hdr(merged):
            state = _ST_TOTAL_TAX
            anchors = _build_column_anchors(row)
            if not anchors:
                state = _ST_SKIP
                continue
            # Label boundary for the TOTAL TAX section:
            # The section has a "Location" column (e.g. "East Chicago" or just
            # "Gary") that sits between the operator name and the first numeric
            # column.  City words occasionally start a few pixels left of the
            # Location column header due to PDF rendering (e.g. "East" at 162.7
            # vs "Location" header at 168.7).  Casino names themselves never
            # extend past x1 ≈ 110 px, but the "A"/"B" Hard-Rock suffix reaches
            # x0 ≈ 105 px.  Use the Location anchor x0 minus 15 px — this keeps
            # the "A"/"B" suffix (x0 ≈ 105) while excluding city names
            # (which start ≥ 120 px).
            location_x0 = anchors[1][0] if len(anchors) > 1 else anchors[0][1]
            label_x_boundary = max(location_x0 - 15.0, anchors[0][1] + 5.0)

            # Identify the "Wagering Tax" column (not "Sports Wagering Tax").
            # When both "Supplemental Tax" and "Sports Wagering Tax" merge into
            # one anchor (x-gap < _X_GAP), the pure "Wagering Tax" anchor is the
            # one that matches _WAG_TAX_RE but does NOT contain "sports".
            col_target = []
            for i, (_x0, _x1, lbl) in enumerate(anchors):
                if _WAG_TAX_RE.search(lbl) and not re.search(r"sports", lbl, re.I):
                    col_target.append(i)
            # Fallback: rightmost Wagering Tax anchor
            if not col_target:
                for i, (_x0, _x1, lbl) in enumerate(anchors):
                    if _WAG_TAX_RE.search(lbl):
                        col_target = [i]
            metric_name = "tax_paid"
            continue

        if _is_wagering_tax_hdr(merged):
            state = _ST_WAG_TAX
            anchors = _build_column_anchors(row)
            if not anchors:
                state = _ST_SKIP
                continue
            # In WAGERING TAX, the label is purely the casino name (no Location).
            # Use the x0 of the second anchor (first non-name column).
            label_x_boundary = anchors[1][0] if len(anchors) > 1 else anchors[0][1]

            # AGR is the rightmost anchor (standalone "AGR" label).
            col_target = []
            for i, (_x0, _x1, lbl) in enumerate(anchors):
                if _AGR_HDR_RE.search(lbl):
                    col_target.append(i)
            metric_name = "ggr"
            continue

        if _is_any_section_hdr(merged):
            # Entering Win/Free-Play or any other non-target section.
            state = _ST_SKIP
            continue

        if state == _ST_SKIP or not col_target:
            continue

        # ---- Data row: extract label and target column value ----
        label_words = [w for w in row if w["x0"] < label_x_boundary]
        value_words = [w for w in row if w["x0"] >= label_x_boundary]
        if not label_words:
            continue

        label = " ".join(w["text"] for w in label_words).strip()
        # Strip trailing "**" annotation that some racino names carry.
        label = re.sub(r"\*+$", "", label).strip()
        # Normalise IGC sub-operator suffixes: Hard Rock Northern Indiana was
        # taxed in two tranches (A and B) through June 2025; the WAGERING TAX
        # section always uses the combined name.  Strip the trailing " A" / " B"
        # so that tax_paid facts join correctly to the ggr facts.
        label = re.sub(r"\s+[AB]\s*$", "", label).strip()
        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue
        # Skip rows that are all-uppercase sub-headers with no dollar values.
        if label.isupper() and not any(_VALUE_TOKEN_RE.match(w["text"]) for w in value_words):
            continue

        # Assign numeric tokens to column anchors.
        col_values: dict[int, str] = {}
        for w in value_words:
            if not _VALUE_TOKEN_RE.match(w["text"]):
                continue
            ci = _assign_to_column(w, anchors)
            if ci >= 0 and ci not in col_values:
                col_values[ci] = w["text"]

        for ci in col_target:
            raw = col_values.get(ci)
            if raw is not None:
                cents = _to_cents_pdf(raw)
                # Sanity gate: individual operator monthly value must be ≤ $100M.
                # This rejects any YTD leak or cumulative figures.
                if cents is not None and abs(cents) <= 10_000_000_000:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="commercial-casino", period=period,
                        metric_name=metric_name, value_usd_cents=cents,
                        source_url=source_url,
                    ))

    return out


# ---------------------------------------------------------------------------
# Word-cluster sports page parser
# ---------------------------------------------------------------------------

def _parse_sports_page_words(page, period: str, source_url: str) -> list[MetricFact]:
    """Word-position clustering fallback for sports wagering section pages.

    Identifies the header row (contains "Handle" or "Wager"), builds
    per-column anchors, then reads each operator row.
    Emits metric_name='handle', 'ggr' (AGR), and 'tax_paid'.

    Indiana sports PDF quirks handled here:
    1. The "Tax" column header sits ~2px above the other header words, causing
       it to land in a separate row bucket at _Y_TOL=3.  Fix: collect all words
       within ±8 px of the header row's y-centre for anchor building.
    2. Handle column values have x0 well left of the "Handle" header word's x0.
       Using anchors[1][0] as label_x_boundary incorrectly absorbs Handle values
       into the operator label.  Fix: use the midpoint between the label-anchor's
       x1 and the Handle-anchor's x0 as the boundary.
    3. Tax column values are consistently ~2 px above their operator row (same
       PDF rendering offset as the header).  At _Y_TOL=3 they become orphan rows
       that appear immediately before the operator row they belong to.  Fix:
       accumulate no-label rows as a pending orphan and prepend them to the next
       labeled row before processing.
    """
    words = page.extract_words(
        x_tolerance=3, y_tolerance=3,
        keep_blank_chars=False, use_text_flow=True,
    )
    rows = _group_words_into_rows(words)
    if not rows:
        return []

    def _is_sports_header(text: str) -> bool:
        return bool(_HANDLE_RE.search(text)) and bool(_AGR_RE.search(text) or _TAX_RE.search(text))

    hdr_idx = _find_header_row_idx(rows, _is_sports_header)
    if hdr_idx == -1:
        # Looser fallback: any row with "Handle"
        hdr_idx = _find_header_row_idx(rows, lambda t: bool(_HANDLE_RE.search(t)))
    if hdr_idx == -1:
        return []

    # Build column anchors from all words within ±8 px of the header row's
    # y-centre.  This captures the "Tax" header word which sits ~2 px above
    # the other header words and would otherwise be placed in a different row
    # bucket when using _Y_TOL=3.
    hdr_y_centre = (
        sum(w["top"] for w in rows[hdr_idx]) / max(len(rows[hdr_idx]), 1)
    )
    header_words: list[dict] = [
        w for w in words if abs(w["top"] - hdr_y_centre) <= 8
    ]

    # Data starts after the header row; skip any non-numeric continuation lines.
    data_start = hdr_idx + 1
    while data_start < len(rows):
        candidate = rows[data_start]
        if not any(_VALUE_TOKEN_RE.match(w["text"]) for w in candidate):
            data_start += 1
        else:
            break

    anchors = _build_column_anchors(sorted(header_words, key=lambda w: w["x0"]))
    if not anchors:
        return []

    col_handle: list[int] = []
    col_agr: list[int] = []
    col_tax: list[int] = []
    for i, (_x0, _x1, lbl) in enumerate(anchors):
        if i == 0:
            continue
        if _HANDLE_RE.search(lbl):
            col_handle.append(i)
        elif _AGR_RE.search(lbl):
            col_agr.append(i)
        elif _TAX_RE.search(lbl):
            col_tax.append(i)

    # label_x_boundary: midpoint between the label-column's x1 and the first
    # data-column's x0.  Using anchors[1][0] directly as the boundary causes
    # Handle values (whose x0 < header word x0) to bleed into the operator label.
    # The midpoint (~199 px for the 2026 PDF) safely separates operator names
    # (x1 < 140 px) from numeric values (x0 > 270 px).
    if len(anchors) > 1:
        label_x_boundary = (anchors[0][1] + anchors[1][0]) / 2.0
    else:
        label_x_boundary = anchors[0][1]

    # Sentinel: the "Monthly Sports Wagering Handle by Sport" sub-table at the
    # bottom of the sports summary page contains sport-category rows (Football,
    # Basketball, …) with a Month / YTD column layout — not operator rows.
    # Stop processing as soon as this sub-table header is encountered.
    # Orphan-row merge: Tax column values render ~2 px above their operator row,
    # landing in a separate row bucket and appearing *before* the operator row
    # when iterating by y-position.  We accumulate these no-label orphan rows as
    # a pending buffer and prepend them to the next labeled operator row.
    out: list[MetricFact] = []
    pending_orphan: list[dict] = []

    for row in rows[data_start:]:
        if not row:
            continue
        # Stop at the sport-category sub-table.  _SPORT_SUBTABLE_RE (module-level)
        # matches both the 2025 heading "State Wide Handle by Sport" and the
        # 2026+ heading "Monthly Sports Wagering Handle by Sport".
        row_text = " ".join(w["text"] for w in row)
        if _SPORT_SUBTABLE_RE.search(row_text):
            break

        label_words = [w for w in row if w["x0"] < label_x_boundary]
        if not label_words:
            # No operator label — this is an orphan value row.
            # Buffer it; it will be merged into the next labeled row.
            if any(_VALUE_TOKEN_RE.match(w["text"]) for w in row):
                pending_orphan.extend(row)
            continue

        # Labeled operator row: absorb any buffered orphan words first.
        effective_row = pending_orphan + row
        pending_orphan = []

        label = " ".join(w["text"] for w in label_words).strip()
        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue

        value_words = [w for w in effective_row if w["x0"] >= label_x_boundary]
        col_values: dict[int, str] = {}
        for w in value_words:
            if not _VALUE_TOKEN_RE.match(w["text"]):
                continue
            ci = _assign_to_column(w, anchors)
            if ci >= 0 and ci not in col_values:
                col_values[ci] = w["text"]

        for ci in col_handle:
            raw = col_values.get(ci)
            if raw is not None:
                cents = _to_cents_pdf(raw)
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="handle", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_agr:
            raw = col_values.get(ci)
            if raw is not None:
                cents = _to_cents_pdf(raw)
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="ggr", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_tax:
            raw = col_values.get(ci)
            if raw is not None:
                cents = _to_cents_pdf(raw)
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="tax_paid", value_usd_cents=cents,
                        source_url=source_url,
                    ))

    return out


# ---------------------------------------------------------------------------
# Original extract_tables()-based helpers (kept for older PDFs that DO have
# ruling lines and where extract_tables() succeeds).
# ---------------------------------------------------------------------------

def _parse_casino_pdf_table(table: list[list], period: str, source_url: str) -> list[MetricFact]:
    """Extract GGR (AGR) and Wagering Tax from a casino PDF table (ruled lines).

    Handles the two relevant sections of the IGC monthly table:
    - TOTAL TAX section → tax_paid (Wagering Tax column, not Sports Wagering Tax)
    - WAGERING TAX section → ggr (AGR column, rightmost)
    Skips the Win/Free-Play section and YTD data.
    """
    out: list[MetricFact] = []
    current_section = _SECTION_OTHER
    col_wag_tax_idx: int = -1
    col_agr_idx: int = -1

    for row in table:
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        label = _header_str(row[0]) if row[0] is not None else ""
        cells = _header_cells(row)

        # Detect section transition
        section = _classify_header_row(cells)
        if section != _SECTION_OTHER:
            current_section = section
            if current_section == _SECTION_TOTAL_TAX:
                col_wag_tax_idx = -1
                for i, h in enumerate(cells):
                    if _WAG_TAX_RE.search(h) and not re.search(r"sports", h, re.I):
                        col_wag_tax_idx = i
                if col_wag_tax_idx == -1:
                    for i, h in enumerate(cells):
                        if _WAG_TAX_RE.search(h):
                            col_wag_tax_idx = i
            elif current_section == _SECTION_WAG_TAX:
                col_agr_idx = -1
                for i, h in enumerate(cells):
                    if _AGR_HDR_RE.search(h):
                        col_agr_idx = i
            continue

        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue

        if current_section == _SECTION_TOTAL_TAX and col_wag_tax_idx >= 0:
            if col_wag_tax_idx < len(row):
                cents = _to_cents_pdf(row[col_wag_tax_idx])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="commercial-casino", period=period,
                        metric_name="tax_paid", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        elif current_section == _SECTION_WAG_TAX and col_agr_idx >= 0:
            if col_agr_idx < len(row):
                cents = _to_cents_pdf(row[col_agr_idx])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="commercial-casino", period=period,
                        metric_name="ggr", value_usd_cents=cents,
                        source_url=source_url,
                    ))

    return out


def _parse_sports_pdf_table(table: list[list], period: str, source_url: str) -> list[MetricFact]:
    """Extract Handle, AGR (ggr), and tax facts from a sports PDF table (ruled lines).

    Guards against the sport-category sub-table at the bottom of the sports
    summary page (same two-defence approach as _parse_sports_sheet):
      1. Stop iteration when a row's first cell matches _SPORT_SUBTABLE_RE.
      2. Reject any row label that matches _SPORT_CATEGORY_NAMES.
    """
    out: list[MetricFact] = []
    col_handle: list[int] = []
    col_agr: list[int] = []
    col_tax: list[int] = []
    header_found = False

    for row in table:
        if not row or all((c is None or str(c).strip() == "") for c in row):
            continue
        label = _header_str(row[0]) if row[0] is not None else ""
        cells = _header_cells(row)

        # Stop at the sport-category sub-table header.
        if header_found and label and _SPORT_SUBTABLE_RE.search(label):
            break

        if not header_found:
            if any(_HANDLE_RE.search(h) for h in cells):
                for i, h in enumerate(cells):
                    if _HANDLE_RE.search(h):
                        col_handle.append(i)
                    elif _AGR_RE.search(h):
                        col_agr.append(i)
                    elif _TAX_RE.search(h):
                        col_tax.append(i)
                header_found = True
            continue

        if not label or _is_skip_row(label) or _is_sport_category(label):
            continue

        for ci in col_handle:
            if ci < len(row):
                cents = _to_cents_pdf(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="handle", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_agr:
            if ci < len(row):
                cents = _to_cents_pdf(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="ggr", value_usd_cents=cents,
                        source_url=source_url,
                    ))

        for ci in col_tax:
            if ci < len(row):
                cents = _to_cents_pdf(row[ci])
                if cents is not None:
                    out.append(MetricFact(
                        state="IN", operator=label,
                        vertical="sports-wagering", period=period,
                        metric_name="tax_paid", value_usd_cents=cents,
                        source_url=source_url,
                    ))

    return out


def _parse_casino_pdf_page(page, period: str, source_url: str) -> list[MetricFact]:
    """Try extract_tables() first; fall back to word-position clustering.

    Skips YTD / cumulative pages (page 1 of the IGC monthly PDF).
    """
    # Guard: skip YTD summary pages — they contain cumulative year-to-date
    # totals that must never be recorded as a single-month value.
    txt_snippet = (page.extract_text() or "")[:300]
    if _YTD_PAGE_RE.search(txt_snippet):
        return []

    out: list[MetricFact] = []
    for tbl in page.extract_tables() or []:
        out.extend(_parse_casino_pdf_table(tbl, period, source_url))
    if not out:
        out.extend(_parse_casino_page_words(page, period, source_url))
    return out


def _parse_sports_pdf_page(page, period: str, source_url: str) -> list[MetricFact]:
    """Try extract_tables() first; fall back to word-position clustering.

    Skips the per-sportsbook "Detail of Sports Wagering Tax" page — it has a
    different multi-column layout and contains only sub-operator breakdowns, not
    the monthly operator-level summary we need.
    """
    txt_snippet = (page.extract_text() or "")[:300]
    if _SPORTS_DETAIL_PAGE_RE.search(txt_snippet):
        return []

    out: list[MetricFact] = []
    for tbl in page.extract_tables() or []:
        out.extend(_parse_sports_pdf_table(tbl, period, source_url))
    if not out:
        out.extend(_parse_sports_page_words(page, period, source_url))
    return out


# ---------------------------------------------------------------------------
# Public entry point — XLSX
# ---------------------------------------------------------------------------

def parse_in_xlsx(
    path: Path,
    vertical: str,
    source_url: str,
) -> list[MetricFact]:
    """Parse an Indiana Gaming Commission combined monthly XLSX.

    Args:
        path:       Local path to the downloaded XLSX.
        vertical:   "commercial-casino" or "sports-wagering".
        source_url: Canonical download URL (used for period extraction and
                    stored on every MetricFact).

    Returns:
        List of MetricFact records for the requested vertical.
    """
    period = _period_from_url(source_url)
    if period is None:
        raise ValueError(f"Cannot derive period from URL: {source_url!r}")

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[MetricFact] = []
    try:
        for sheet_name in wb.sheetnames:
            sn = sheet_name.strip()
            if vertical == "commercial-casino":
                # Named sheets (older files): "Casino", "Riverboat", etc.
                # IGC XLSX sheets are named "Sheet1", "Sheet2", … — no keyword in name.
                # Sheet1 always holds the combined monthly tax/AGR summary page.
                # We also accept any sheet whose name matches the casino pattern.
                if _CASINO_SHEET_RE.search(sn) or sn.lower() in ("sheet1",):
                    out.extend(_parse_casino_sheet(wb[sheet_name], period, source_url))
            elif vertical == "sports-wagering":
                # Sports data lives in a sheet named "Sheet7" or a sports-keyword sheet.
                # Fall back to scanning all sheets for sports headers.
                if _SPORTS_SHEET_RE.search(sn) or sn.lower() in ("sheet7",):
                    out.extend(_parse_sports_sheet(wb[sheet_name], period, source_url))
    finally:
        wb.close()

    return [f for f in out if f.operator]


# ---------------------------------------------------------------------------
# Public entry point — PDF
# ---------------------------------------------------------------------------

def parse_in_pdf(
    path: Path,
    vertical: str,
    source_url: str,
) -> list[MetricFact]:
    """Parse an Indiana Gaming Commission combined monthly PDF.

    Indiana switched from XLSX to PDF delivery in late 2024.  The PDF
    contains the same data as the legacy XLSX: one section for commercial
    casinos (Win = GGR) and one for sports wagering (Handle + AGR = GGR).

    Args:
        path:       Local path to the downloaded PDF.
        vertical:   "commercial-casino" or "sports-wagering".
        source_url: Canonical download URL (used for period extraction and
                    stored on every MetricFact).

    Returns:
        List of MetricFact records for the requested vertical.  Returns []
        on any parse error — never raises.
    """
    period = _period_from_url(source_url)
    if period is None:
        # Try to extract from path name as fallback (same YYYY-MM-Revenue pattern)
        period = _period_from_url(path.name)
    if period is None:
        return []

    out: list[MetricFact] = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                txt = (page.extract_text() or "").lower()
                is_casino = bool(_CASINO_SHEET_RE.search(txt))
                is_sports = bool(_SPORTS_SHEET_RE.search(txt))
                if vertical == "commercial-casino" and is_casino:
                    out.extend(_parse_casino_pdf_page(page, period, source_url))
                elif vertical == "sports-wagering" and is_sports:
                    out.extend(_parse_sports_pdf_page(page, period, source_url))
    except Exception:  # noqa: BLE001
        return []

    return [f for f in out if f.operator]


# ---------------------------------------------------------------------------
# Public dispatch entry point — routes XLSX or PDF automatically
# ---------------------------------------------------------------------------

def parse_in(
    path: Path,
    vertical: str,
    source_url: str,
) -> list[MetricFact]:
    """Dispatch to parse_in_xlsx or parse_in_pdf based on file extension.

    This is the canonical entry point for the backfill pipeline.  Indiana
    switched from XLSX to PDF delivery in late 2024; both formats remain in
    the 36-month backfill window simultaneously.
    """
    if str(path).lower().endswith(".pdf"):
        return parse_in_pdf(path, vertical, source_url)
    return parse_in_xlsx(path, vertical, source_url)
