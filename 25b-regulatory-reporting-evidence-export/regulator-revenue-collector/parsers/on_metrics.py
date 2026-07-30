# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""iGaming Ontario (iGO) monthly market performance XLSX parser.

File format
-----------
iGO publishes one combined workbook covering all three verticals.

Two sheets are parsed:

* ``(Data) Monthly Stats``
  All-market totals (no vertical breakdown). Columns:
  FiscalYearQuarter | YearMonth | CashWagers(M) | ... | NAGGR(M) | ...
  YearMonth is already in YYYY-MM format. CashWagers and NAGGR are in
  CAD millions.

* ``(Data) Product Monthly Stats``
  Per-vertical breakdown. Columns:
  FiscalYearQuarter | YearMonth | ProductCategory | CashWagers(M) | ... | NAGGR(M) | ...
  ProductCategory values: CASINO, BETTING, P2P POKER.

All monetary values are CAD millions (M) per the workbook coversheet.
FX conversion: CAD → USD ≈ 0.74 (applied before converting millions to cents).

Vertical mapping (ProductCategory → canonical vertical):
  CASINO    → igaming
  BETTING   → sports-wagering
  P2P POKER → online-poker

When ``vertical`` is passed as a non-empty string, only rows whose mapped
vertical equals the requested vertical are emitted from the product sheet.
Passing ``vertical=""`` (or any falsy value) emits all three verticals.

The all-market totals sheet has no ProductCategory column; those rows are
emitted with ``vertical`` set to the value passed in (defaulting to
``"igaming"`` if blank, to stay consistent with the collector which sets
``vertical="igaming"`` for the monthly file).

Period derivation
-----------------
1. From the ``YearMonth`` cell in each data row (already YYYY-MM).
2. Falling back to ``_period_from_url()`` when the cell is absent/unparseable.
"""
from __future__ import annotations

import re
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CAD_TO_USD: float = 0.74
# Both CashWagers and NAGGR are expressed in CAD millions in the workbook.
CAD_MILLIONS_TO_CENTS_FACTOR: float = CAD_TO_USD * 1_000_000 * 100

_PRODUCT_SHEET = "(Data) Product Monthly Stats"
_TOTALS_SHEET = "(Data) Monthly Stats"

# ProductCategory cell value → canonical vertical
_VERTICAL_MAP: dict[str, str] = {
    "CASINO": "igaming",
    "BETTING": "sports-wagering",
    "P2P POKER": "online-poker",
}

# Month name → zero-padded number
_MONTHS: dict[str, str] = {
    "January": "01", "February": "02", "March": "03", "April": "04",
    "May": "05", "June": "06", "July": "07", "August": "08",
    "September": "09", "October": "10", "November": "11", "December": "12",
}

_URL_PERIOD_RE = re.compile(
    r"(\d{4})\s+(January|February|March|April|May|June|July|"
    r"August|September|October|November|December)",
    re.IGNORECASE,
)

# Header cell substrings that identify the columns we care about
_WAGER_HEADER = "CashWagers(M)"
_GGR_HEADER = "NAGGR(M)"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _period_from_url(url: str) -> str | None:
    """Extract YYYY-MM from an iGO URL like '…2026%20February.xlsx'."""
    u = unquote(url)
    m = _URL_PERIOD_RE.search(u)
    if not m:
        return None
    return f"{m.group(1)}-{_MONTHS.get(m.group(2).title(), '01')}"


def _period_from_cell(cell_value) -> str | None:
    """Parse a YearMonth cell value that is already 'YYYY-MM'."""
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    if re.fullmatch(r"\d{4}-\d{2}", s):
        return s
    return None


def _cad_millions_to_usd_cents(value) -> int | None:
    """Convert a CAD millions float/int cell to integer USD cents."""
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    return int(round(f * CAD_MILLIONS_TO_CENTS_FACTOR))


def _find_col(header_row: tuple, substring: str) -> int | None:
    """Return the 0-based column index whose header cell contains *substring*."""
    for idx, cell in enumerate(header_row):
        if cell is not None and substring in str(cell):
            return idx
    return None


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def _parse_product_sheet(
    ws,
    vertical: str,
    period_fallback: str | None,
    source_url: str,
) -> list[MetricFact]:
    """Parse ``(Data) Product Monthly Stats`` rows.

    Emits ``handle`` (from CashWagers) and ``ggr`` (from NAGGR) for every
    product category that matches the requested *vertical*.  When *vertical*
    is falsy all three categories are emitted.
    """
    out: list[MetricFact] = []
    header_found = False
    col_yearmonth: int | None = None
    col_product: int | None = None
    col_wager: int | None = None
    col_ggr: int | None = None

    for row in ws.iter_rows(values_only=True):
        # Skip entirely empty rows.
        if not any(c is not None for c in row):
            continue

        # Detect the header row by the presence of the distinctive column names.
        if not header_found:
            if row[0] is not None and "FiscalYearQuarter" in str(row[0]):
                header_found = True
                col_yearmonth = _find_col(row, "YearMonth")
                col_product = _find_col(row, "ProductCategory")
                col_wager = _find_col(row, _WAGER_HEADER)
                col_ggr = _find_col(row, _GGR_HEADER)
            continue

        # Data row: first cell must match a FY pattern.
        if row[0] is None or not str(row[0]).startswith("FY"):
            continue

        # Determine period.
        period: str | None = None
        if col_yearmonth is not None and col_yearmonth < len(row):
            period = _period_from_cell(row[col_yearmonth])
        if period is None:
            period = period_fallback
        if period is None:
            continue

        # Determine product category → vertical.
        raw_product = (
            str(row[col_product]).strip().upper()
            if col_product is not None and col_product < len(row) and row[col_product] is not None
            else ""
        )
        mapped_vertical = _VERTICAL_MAP.get(raw_product)
        if mapped_vertical is None:
            continue  # unknown category; skip

        # Filter by requested vertical when specified.
        if vertical and mapped_vertical != vertical:
            continue

        # Emit handle and GGR.
        for col_idx, metric_name in (
            (col_wager, "handle"),
            (col_ggr, "ggr"),
        ):
            if col_idx is None or col_idx >= len(row):
                continue
            cents = _cad_millions_to_usd_cents(row[col_idx])
            if cents is None:
                continue
            out.append(MetricFact(
                state="ON",
                operator="ON Statewide",
                vertical=mapped_vertical,
                period=period,
                metric_name=metric_name,
                value_usd_cents=cents,
                source_url=source_url,
            ))

    return out


def _parse_totals_sheet(
    ws,
    vertical: str,
    period_fallback: str | None,
    source_url: str,
) -> list[MetricFact]:
    """Parse ``(Data) Monthly Stats`` all-market totals.

    These rows have no ProductCategory; they are tagged with *vertical*
    as passed in (or ``"igaming"`` if blank).
    """
    effective_vertical = vertical or "igaming"
    out: list[MetricFact] = []
    header_found = False
    col_yearmonth: int | None = None
    col_wager: int | None = None
    col_ggr: int | None = None

    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue

        if not header_found:
            if row[0] is not None and "FiscalYearQuarter" in str(row[0]):
                header_found = True
                col_yearmonth = _find_col(row, "YearMonth")
                col_wager = _find_col(row, _WAGER_HEADER)
                col_ggr = _find_col(row, _GGR_HEADER)
            continue

        if row[0] is None or not str(row[0]).startswith("FY"):
            continue

        period: str | None = None
        if col_yearmonth is not None and col_yearmonth < len(row):
            period = _period_from_cell(row[col_yearmonth])
        if period is None:
            period = period_fallback
        if period is None:
            continue

        for col_idx, metric_name in (
            (col_wager, "handle"),
            (col_ggr, "ggr"),
        ):
            if col_idx is None or col_idx >= len(row):
                continue
            cents = _cad_millions_to_usd_cents(row[col_idx])
            if cents is None:
                continue
            out.append(MetricFact(
                state="ON",
                operator="ON Statewide",
                vertical=effective_vertical,
                period=period,
                metric_name=metric_name,
                value_usd_cents=cents,
                source_url=source_url,
            ))

    return out


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_on_xlsx(path: Path, vertical: str, source_url: str) -> list[MetricFact]:
    """Parse an iGO market performance XLSX and return a list of MetricFacts.

    Parameters
    ----------
    path:
        Local path to the downloaded XLSX file.
    vertical:
        Canonical vertical string used to filter rows from the product sheet
        (``"igaming"``, ``"sports-wagering"``, or ``"online-poker"``).
        Pass ``""`` to return all verticals.
        Also used to tag rows from the all-market totals sheet (defaults to
        ``"igaming"`` when blank).
    source_url:
        Original download URL; stored verbatim on every MetricFact.
    """
    period_fallback = _period_from_url(source_url)

    wb = load_workbook(str(path), read_only=True, data_only=True)
    out: list[MetricFact] = []
    try:
        if _PRODUCT_SHEET in wb.sheetnames:
            out.extend(_parse_product_sheet(
                wb[_PRODUCT_SHEET], vertical, period_fallback, source_url,
            ))
        if _TOTALS_SHEET in wb.sheetnames:
            out.extend(_parse_totals_sheet(
                wb[_TOTALS_SHEET], vertical, period_fallback, source_url,
            ))
    finally:
        wb.close()

    return out
