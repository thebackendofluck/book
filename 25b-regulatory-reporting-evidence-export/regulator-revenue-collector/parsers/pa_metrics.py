# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""PA PGCB monthly revenue parser. One xlsx per vertical per fiscal year."""
from __future__ import annotations

import logging
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .metrics_model import MetricFact

log = logging.getLogger(__name__)

MONTH_RE = re.compile(
    r"^(January|February|March|April|May|June|July|August|"
    r"September|October|November|December)\s+(\d{4})$",
    re.IGNORECASE,
)
TOTAL_TOKENS = ("FYTD", "FY ", "TOTAL", "YTD", "GRAND TOTAL")
MONEY_RE = re.compile(r"[\$,\s]")

# Sub-section labels that appear in iGaming and Sports as structural
# headers with no data — kept as operator qualifiers (appended to parent).
_SUBSECTION_RE = re.compile(
    r"^(Interactive\s+Slots|Interactive\s+Tables|Banking\s+Tables|"
    r"Non.Banking\s+Tables|Electronic\s+Tables|Fully\s+Automated|"
    r"Hybrid\s+Tables|Total\s+Sports\s+Wagering|Retail\s+Sports\s+Wagering|"
    r"Online\s+Sports\s+Wagering|Total\s+Table\s+Games|"
    r"Non.Banking\s+Tables?\s*\d*)",
    re.IGNORECASE,
)

METRIC_MAP: list[tuple[re.Pattern[str], str]] = [
    # Most-specific patterns first to avoid early broad matches
    (re.compile(r"gross\s+(terminal\s+)?revenue", re.I),          "ggr"),
    (re.compile(r"gross\s+revenue\s*\(taxable\)", re.I),          "ggr"),
    (re.compile(r"fantasy\s+contest\s+adjusted\s+revenue", re.I), "ggr"),
    (re.compile(r"adjusted\s+gross\s+revenue", re.I),             "ggr"),
    (re.compile(r"revenue\s*\(rake", re.I),                       "ggr"),
    (re.compile(r"\brevenue\b", re.I),                            "ggr"),
    (re.compile(r"\bwin\b", re.I),                                "win"),
    (re.compile(r"state\s+tax", re.I),                            "tax_paid"),
    (re.compile(r"\btax\b", re.I),                                "tax_paid"),
    (re.compile(r"\bhandle\b", re.I),                             "handle"),
    (re.compile(r"\bwagers?\b", re.I),                            "handle"),
    (re.compile(r"\bpayouts?\b", re.I),                           "patron_winnings"),
    (re.compile(r"total\s+fees\s+collected", re.I),               "handle"),
    (re.compile(r"amount\s+won", re.I),                           "patron_winnings"),
    (re.compile(r"promotional\s+(credits?|plays?)", re.I),        "promo_credits"),
]
SKIP_METRIC_RE = re.compile(
    r"number\s+of\s+machines|%\s*change|taxable\s+w/s/d|footnote|"
    r"authorized\s+slot|report\s+notes|lsa\b|local\s+share|"
    r"cfa\s+county|operator\s+share|"
    # "Revenue" standalone (sports pre-promo total) — "Gross Revenue (Taxable)"
    # is the official taxable GGR and is captured by its own pattern above.
    r"^revenue$",
    re.I,
)

# Section-header rows that contain counts (not dollars) and whose following
# Gross Revenue would double-count sub-category rows already captured.
# When one of these is the active "sub-section", skip all metric rows.
_ROLLUP_SUBSECTION_RE = re.compile(
    r"^(Total\s+Table\s+Games|Total\s+Sports\s+Wagering)",
    re.IGNORECASE,
)


def _to_cents(raw) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return int(round(float(raw) * 100))
    s = MONEY_RE.sub("", str(raw))
    if not s or s in ("-", "--"):
        return None
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def _period_from_cell(val) -> str | None:
    """Accept string 'Month YYYY' or datetime object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return f"{val.year}-{val.month:02d}"
    m = MONTH_RE.match(str(val).strip())
    if not m:
        return None
    month = datetime.strptime(m.group(1).title(), "%B").month
    return f"{m.group(2)}-{month:02d}"


def _is_total_label(label) -> bool:
    s = str(label or "").upper()
    return any(t in s for t in TOTAL_TOKENS)


# Tokens that, by themselves (or nearly), indicate an aggregate/state total row
# that should suppress further operator attribution. More specific than TOTAL_TOKENS.
_BARE_TOTAL_RE = re.compile(
    r"^(TOTAL|GRAND\s+TOTAL|FY\s+\d|FYTD|YTD)$",
    re.IGNORECASE,
)


def _is_bare_total_label(label: str) -> bool:
    """True only for stand-alone aggregate labels like 'TOTAL', 'GRAND TOTAL'."""
    return bool(_BARE_TOTAL_RE.match(label.strip()))


def _classify_metric(label: str) -> str | None:
    if SKIP_METRIC_RE.search(label):
        return None
    for pat, name in METRIC_MAP:
        if pat.search(label):
            return name
    return None


def _is_all_caps_operator(label: str) -> bool:
    """True for ALL-CAPS casino/operator names (e.g. PENN NATIONAL)."""
    stripped = label.strip()
    if not stripped or len(stripped) < 3:
        return False
    # Allow letters, spaces, hyphens, apostrophes, commas, periods, LLC etc.
    return stripped == stripped.upper() and bool(re.search(r"[A-Z]", stripped))


def _is_operator_row(row) -> bool:
    """Row with a non-empty col-0 string and no numeric data in the rest."""
    if not row or row[0] in (None, ""):
        return False
    if not isinstance(row[0], str):
        return False
    label = row[0].strip()
    if not label or label.startswith("FOOTNOTE") or label.startswith("*"):
        return False
    if MONTH_RE.match(label):
        return False
    if _classify_metric(label):
        return False
    rest = [c for c in row[1:] if c not in (None, "", " ")]
    return len(rest) == 0


def _scan_period_cols(row) -> dict[int, str]:
    """Return {col_index: period_str} for all month-columns in a row."""
    result: dict[int, str] = {}
    for i, c in enumerate(row):
        p = _period_from_cell(c)
        if p is not None:
            result[i] = p
    return result


def _maybe_deytd(facts: list[MetricFact]) -> list[MetricFact]:
    """Detect fiscal-YTD cumulative data and convert to month-over-month deltas.

    PGCB occasionally publishes mid-year interactive/sports reports where the
    monthly columns contain fiscal-year-to-date (FYTD) cumulative totals rather
    than true monthly values.  Symptoms: per-operator values are strictly
    non-decreasing across the fiscal year (Jul → Jun), with August ≫ July,
    September ≫ August, etc.  The state GRAND TOTAL rows are excluded (they are
    always dropping them from the parser), so detection focuses on individual
    operator rows.

    Detection criterion (applied per metric_name):
      For each (operator, metric_name) group that has ≥ 3 chronologically-ordered
      positive values, count how many consecutive pairs are non-decreasing.
      If ≥ 75 % of eligible groups are monotonically non-decreasing we flag the
      entire sheet as cumulative and convert every group to deltas.

    Conversion: delta[0] = value[0]; delta[i] = value[i] − value[i−1].
    Negative deltas (amended prior-month corrections) are preserved as-is.
    """
    if not facts:
        return facts

    # Group by (operator, metric_name), keeping facts in period order
    groups: dict[tuple[str, str], list[MetricFact]] = defaultdict(list)
    for f in facts:
        groups[(f.operator, f.metric_name)].append(f)

    # Sort each group chronologically
    for key in groups:
        groups[key].sort(key=lambda f: f.period)

    # Assess monotonicity — only consider groups with ≥ 3 positive-value entries
    monotone_count = 0
    eligible_count = 0
    for (op, metric), grp in groups.items():
        # Skip rollup/total operators (they use different keys in real data)
        positive = [f for f in grp if f.value_usd_cents > 0]
        if len(positive) < 3:
            continue
        eligible_count += 1
        vals = [f.value_usd_cents for f in positive]
        non_decreasing = sum(1 for a, b in zip(vals, vals[1:]) if b >= a)
        # Strictly monotone if every pair is non-decreasing
        if non_decreasing == len(vals) - 1:
            monotone_count += 1

    if eligible_count < 3:
        # Not enough data to reliably detect cumulative encoding
        return facts

    ratio = monotone_count / eligible_count
    if ratio < 0.75:
        return facts

    log.warning(
        "pa_metrics.ytd_cumulative_detected: "
        "monotone_groups=%d eligible_groups=%d ratio=%.2f — converting to deltas",
        monotone_count, eligible_count, ratio,
    )

    # Convert all groups to deltas
    out: list[MetricFact] = []
    for grp in groups.values():
        if not grp:
            continue
        prev_cents = 0
        for f in grp:  # already sorted by period
            delta = f.value_usd_cents - prev_cents
            prev_cents = f.value_usd_cents
            out.append(MetricFact(
                state=f.state,
                operator=f.operator,
                vertical=f.vertical,
                period=f.period,
                metric_name=f.metric_name,
                value_usd_cents=delta,
                source_url=f.source_url,
            ))
    return out


def _parse_sheet(ws, vertical: str, source_url: str) -> list[MetricFact]:
    out: list[MetricFact] = []
    current_operator = ""
    current_subsection = ""
    period_cols: dict[int, str] = {}

    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        if not any(c not in (None, "", " ") for c in row):
            continue

        # --- Period header row: col-0 is None/blank, months are at col 1+ ---
        # Detect before checking operator rows so we don't skip period rows.
        if row[0] is None or (isinstance(row[0], str) and not row[0].strip()):
            month_cols = _scan_period_cols(row)
            if month_cols:
                period_cols = month_cols
                current_subsection = ""
            continue

        # --- Tables / VGT / Fantasy / Slots: col-0 = label, months at col 1+
        #     Also detects period headers where col-0 is "Establishments" (VGT)
        if isinstance(row[0], str):
            # Check if this row itself is a period header (col-0 has no data,
            # but some files put month strings starting at col 1 with a label)
            month_cols = _scan_period_cols(row[1:])
            if month_cols:
                # Re-index: add 1 because we sliced from row[1:]
                period_cols = {k + 1: v for k, v in month_cols.items()}
                current_subsection = ""
                continue

        # --- Operator / sub-section row ---
        if _is_operator_row(row):
            label = str(row[0]).strip()
            if _ROLLUP_SUBSECTION_RE.match(label):
                # Sub-section that is a rollup of what follows — skip metrics
                current_subsection = "_ROLLUP_"
            elif _is_bare_total_label(label):
                # State/grand aggregate total — stop attributing to any operator
                current_operator = ""
                current_subsection = ""
            elif _is_all_caps_operator(label):
                current_operator = label
                current_subsection = ""
                # Don't reset period_cols — sheets have one global header row
            elif _SUBSECTION_RE.match(label):
                current_subsection = label
            else:
                # Title-case section with no data: treat as operator if no
                # all-caps operator has been set yet, else as sub-section
                if not current_operator:
                    current_operator = label
                else:
                    current_subsection = label
            continue

        # --- Data row ---
        label = row[0]
        if label is None or not isinstance(label, str):
            continue
        label = label.strip()
        if not label:
            continue

        # Detect rollup-section headers (e.g. "Total Table Games") — must come
        # before _is_total_label because "TOTAL" is a substring of those labels.
        if _ROLLUP_SUBSECTION_RE.match(label):
            current_subsection = "_ROLLUP_"
            continue

        if _is_total_label(label):
            continue

        # Detect named sub-section headers that have count data in rest cols
        # (e.g. "Non-Banking Tables 3", "Banking Tables 4", "Hybrid Tables").
        # These rows EXIT rollup mode so their Gross Revenue rows get captured.
        if _SUBSECTION_RE.match(label) and _classify_metric(label) is None:
            current_subsection = label
            continue

        # Skip all metric rows while inside a rollup sub-section (avoids
        # double-counting with the detailed sub-category rows that follow).
        if current_subsection == "_ROLLUP_":
            continue

        metric_name = _classify_metric(label)
        if metric_name is None or not current_operator or not period_cols:
            continue

        # Build effective operator name incorporating sub-section
        op_name = current_operator
        if current_subsection:
            op_name = f"{current_operator} | {current_subsection}"

        for col_idx, period in period_cols.items():
            if col_idx >= len(row):
                continue
            cents = _to_cents(row[col_idx])
            if cents is None:
                continue
            out.append(MetricFact(
                state="PA", operator=op_name, vertical=vertical,
                period=period, metric_name=metric_name,
                value_usd_cents=cents, source_url=source_url,
            ))
    return out


# Sheet name patterns to skip (footnotes, map info, legacy annual summaries).
# Note: "FY xx-xx" sheets in VGT contain the fiscal-year summary and must NOT
# be skipped — only "Annual" (calendar-year) and "Weekly" sheets are legacy.
_SKIP_SHEET_RE = re.compile(
    r"^(footnote|esri_|weekly|annual\b)",
    re.IGNORECASE,
)
# VGT date-named weekly sheets: "Nov 6", "Jan22", "April 2", "June 25", etc.
# Match any full-month or abbreviated-month name followed by a space/digit.
_VGT_WEEKLY_RE = re.compile(
    r"^(January|February|March|April|May|June|July|August|September|"
    r"October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)"
    r"\s*\d",
    re.IGNORECASE,
)


def parse_pa_xlsx(path: Path, vertical: str, source_url: str) -> list[MetricFact]:
    """Walk every monthly tab, extract per-operator metrics.

    Applies FY-YTD cumulative detection after parsing each sheet: if the values
    appear monotonically non-decreasing (i.e. PGCB published mid-year cumulative
    data before replacing the file with true monthly figures), the function
    automatically converts them to month-over-month deltas so the DB receives
    correct monthly revenue rather than inflated FYTD totals.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[MetricFact] = []
    try:
        for sheet_name in wb.sheetnames:
            sn = sheet_name.strip()
            if _SKIP_SHEET_RE.match(sn):
                continue
            if _VGT_WEEKLY_RE.match(sn):
                continue
            sheet_facts = _parse_sheet(wb[sheet_name], vertical, source_url)
            out.extend(_maybe_deytd(sheet_facts))
    finally:
        wb.close()
    return out
