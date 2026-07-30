# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""UKGC quarterly Industry Statistics XLSX parser.

Source: https://www.gamblingcommission.gov.uk/statistics-and-research/publication/industry-statistics
Most-recent file inspected:
  Industry_Statistics_-_Quarterly_report_-_Financial_year_April_2025_to_March_2026
  -_Financial_Quarter_2_-_Official_Statistic.xlsx
  (July–September 2025, released 26 Feb 2026)
  URL: https://assets.ctfassets.net/j16ev64qyf6l/4UCJ50E0Ylj3E2HldP0Y04/
       70027aaadbdbd7e4e9bb2252af2fd3fe/Industry_Statistics_-_Quarterly_report_
       -_Financial_year_April_2025_to_March_2026_-_Financial_Quarter_2
       -_Official_Statistic.xlsx

Actual sheet structure (sheet names are numeric strings, not sector names):
  'Cover Note' — metadata, skip
  '0'          — contents table, skip
  '1'          — Industry Overview (aggregated, skip — per-sector sheets used instead)
  '2'          — Adult Gaming Centres (non-remote)  → vertical: arcades
  '3'          — Family Entertainment Centres       → vertical: arcades
  '4'          — Betting (non-remote)               → vertical: retail-betting
  '5'          — Bingo (non-remote)                 → vertical: bingo
  '6'          — Casino (non-remote)                → vertical: retail-casino
  '7'          — Remote Casino, Betting & Bingo     → verticals: igaming / sports-wagering / bingo
  '8'          — Lotteries (remote & non-remote)    → no GGY column, skip
  '9'          — National Lottery                   → no GGY column, skip

For sheets 2–6: header row is row index 1; GGY (£m) is column index 3.
For sheet 7: header row is row index 1; separate GGY columns for Betting (6),
  Bingo (7), Casino (8).  Column 5 is Total GGY (emitted as composite).
Data rows start at row index 2.  A trailing footnote digit on the period label
  (e.g. "Jul 2025 - Sep 20251") is stripped before parsing.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from .metrics_model import MetricFact

# Fixed exchange rate (GBP → USD); replace with fx_rates table lookup later.
_FX_GBP_USD = 1.27

# Regex to extract Financial Quarter number from URL or filename.
_FQ_URL_RE = re.compile(
    r"financial[_\- ]quarter[_\- ](\d)|[_\-/]q(\d)\b", re.IGNORECASE
)
_FY_URL_RE = re.compile(
    r"financial[_\- ]year[_\- ]april[_\- ](\d{4})", re.IGNORECASE
)

# UK financial year: Apr–Mar.  Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar.
_QUARTER_END_MONTH: dict[int, tuple[int, int]] = {
    1: (0, 6),   # same FY start year, month 06
    2: (0, 9),   # same FY start year, month 09
    3: (0, 12),  # same FY start year, month 12
    4: (1, 3),   # FY start year + 1, month 03
}

# Month abbreviation → month number (for period parsing from row labels).
_MONTH_NUM = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Sheets 2–6 configuration: sheet_id → (vertical, ggy_col_index)
# GGY (£m) is always column index 3 (0-based) for these simple sheets.
_SIMPLE_SHEETS: dict[str, tuple[str, int]] = {
    "2": ("arcades", 3),          # Adult Gaming Centres (non-remote)
    "3": ("arcades", 3),          # Family Entertainment Centres (non-remote)
    "4": ("retail-betting", 3),   # Betting (non-remote)
    "5": ("bingo", 3),            # Bingo (non-remote)
    "6": ("retail-casino", 3),    # Casino (non-remote)
}

# Sheet 7 (RCBB) has multiple GGY columns.
# col_index → vertical
# IMPORTANT: 'remote-total' (col 5) is the SUM of cols 6/7/8 — emitting it as
# a separate fact double-counts when the dashboard sums across verticals.
# We deliberately omit it; downstream consumers can re-derive the total from
# the per-vertical rows if needed.
_RCBB_GGY_COLS: dict[int, str] = {
    6: "sports-wagering",   # Betting GGY
    7: "bingo-remote",      # Bingo GGY
    8: "igaming",           # Casino GGY
}

# --- ANNUAL XLSX (different sheet layout — 0, 1, 2, …, 10, 10a, 10b, 10c, 11, 12) ---
# In the annual report, tab "7" is Bingo non-remote (NOT RCBB), so the quarterly
# logic above would ingest the wrong columns. The RCBB data is in 10/10a/10b/10c,
# and sheet "10" is only an aggregate — the sub-tabs contain Turnover + GGY with
# explicit column headers. The column positions below were pinned against the
# FY 2024-25 XLSX (published 2025-10) and verified stable back to FY 2008-09.
_ANNUAL_REMOTE_SHEETS: dict[str, tuple[str, int]] = {
    "10a": ("igaming", 10),         # Casino remote, GGY Total at col 10
    "10b": ("sports-wagering", 6),  # Betting remote, GGY Total at col 6
    "10c": ("bingo-remote", 3),     # Bingo remote, GGY at col 3
}
# Non-remote annual sheets; each has a simple "Total GGY" layout at col 1.
# Disabled by default because the global dashboard aggregates remote only —
# flip _EMIT_NON_REMOTE to True when a retail view is needed.
_EMIT_NON_REMOTE = False
_ANNUAL_NON_REMOTE_SHEETS: dict[str, tuple[str, int]] = {
    "5a": ("arcades-agc", 1),
    "5b": ("arcades-fec", 1),
    "6a": ("retail-betting-otc", 1),
    "7a": ("retail-bingo", 1),
    "8a": ("retail-casino", 1),
    "9":  ("society-lotteries", 1),
    "11": ("national-lottery", 1),
}


def _period_from_url(source_url: str) -> str | None:
    """Derive YYYY-MM (last month of quarter) from URL / filename."""
    fq_m = _FQ_URL_RE.search(source_url)
    fy_m = _FY_URL_RE.search(source_url)
    if not fq_m or not fy_m:
        return None
    quarter = int(fq_m.group(1) or fq_m.group(2))
    fy_start = int(fy_m.group(1))
    year_offset, month = _QUARTER_END_MONTH[quarter]
    return f"{fy_start + year_offset}-{month:02d}"


def _period_from_row_label(label: str) -> str | None:
    """Parse YYYY-MM from a row label like 'Jul 2025 - Sep 20251'.

    Strips trailing footnote digits from the year and returns the end month.
    """
    # Strip trailing footnote digit(s): "20251" → "2025"
    clean = re.sub(r"(\d{4})\d+", r"\1", label)
    # Match "MMM YYYY - MMM YYYY"
    m = re.search(
        r"([A-Za-z]{3})\s+(\d{4})\s*[-–]\s*([A-Za-z]{3})\s+(\d{4})",
        clean,
    )
    if not m:
        return None
    end_mon_str = m.group(3).lower()
    end_year = int(m.group(4))
    month_num = _MONTH_NUM.get(end_mon_str)
    if month_num is None:
        return None
    return f"{end_year}-{month_num:02d}"


def _is_data_row(label: object) -> bool:
    """Return True if the label looks like a quarter period string."""
    if label is None:
        return False
    s = str(label).strip()
    return bool(re.search(r"[A-Za-z]{3}\s+\d{4}", s))


# Annual rows have the form "Apr YYYY - Mar YYYY+1" with two distinct years.
# Quarterly rows have the form "Apr - Jun YYYY" with a single shared year.
# `(?<!\w)` prevents matching mid-token years; `(?!\d)` allows trailing
# footnote letters like 2025P, 2023R without losing the year capture.
_YEAR_RE = re.compile(r"(?<!\d)(\d{4})(?!\d)")


def _is_annual_period_row(label: object) -> bool:
    """True iff the row label is an annual FY range (two distinct years
    AND it actually contains a `-` or `–` between two month tokens — this
    rules out the sheet *title* row which mentions both years in prose)."""
    if label is None:
        return False
    s = str(label).strip()
    if not re.search(
        r"[A-Za-z]{3}\s+\d{4}[A-Za-z]?\s*[-–]\s*[A-Za-z]{3}\s+\d{4}", s
    ):
        return False
    years = _YEAR_RE.findall(s)
    return len(years) >= 2 and years[0] != years[1]


def _looks_like_period_label(label: object) -> bool:
    """True for any row whose label resembles a period (annual OR quarterly).

    Used to decide whether the data block has *ended*. Quarterly form is
    'MMM - MMM YYYY'; annual form is 'MMM YYYY - MMM YYYY[A-Z]?'. The
    sheet's title row ('Industry Statistics. Annual report April 2024 to
    March 2025') intentionally does NOT match because it has no MMM-MMM
    range form.
    """
    if label is None:
        return False
    s = str(label).strip()
    quarterly = re.search(
        r"[A-Za-z]{3}\s*[-–]\s*[A-Za-z]{3}\s+\d{4}[A-Za-z]?", s
    )
    annual = re.search(
        r"[A-Za-z]{3}\s+\d{4}[A-Za-z]?\s*[-–]\s*[A-Za-z]{3}\s+\d{4}", s
    )
    return bool(quarterly or annual)


def _gbp_millions_to_usd_cents(gbp_millions: float) -> int:
    """Convert GBP millions → USD cents."""
    return int(gbp_millions * 1_000_000 * _FX_GBP_USD * 100)


def _parse_uk_annual(wb, source_url: str) -> list[MetricFact]:
    """Parse the UKGC ANNUAL Industry Statistics XLSX (10/10a/10b/10c layout).

    The annual report has 17 years of historical FY rows per sheet. We emit
    one MetricFact per (sheet, FY-row) — period is derived from the row label
    ("Apr 2024 - Mar 2025" → "2025-03"). Verticals come from the sheet
    mapping (`_ANNUAL_REMOTE_SHEETS`); GGY column index is sheet-specific.
    """
    out: list[MetricFact] = []
    for sheet_id, (vertical, ggy_col) in _ANNUAL_REMOTE_SHEETS.items():
        if sheet_id not in wb.sheetnames:
            continue
        ws = wb[sheet_id]
        # Tab 10b contains TWO data blocks: rows 8-42 are Total Turnover/GGY,
        # rows 47+ are a per-sport turnover breakdown (Cricket, Dogs, Football,
        # ...). We only want the first block; reading the second would conflate
        # GGY values with Football turnover and inflate the totals ~4×.
        # Strategy:
        #   - Quarterly rows ("Apr - Jun 2024") are interleaved with annual
        #     rows ("Apr 2024 - Mar 2025") in block 1. We skip quarterly for
        #     emission (annual XLSX → annual facts only), but we do NOT break
        #     on them, otherwise the loop exits at the first quarterly row.
        #   - The block ends at a row whose label doesn't look like a period
        #     at all (blank, header repeat, footer note). At that point, after
        #     we've already collected at least one fact, we break.
        seen_data = False
        for row in ws.iter_rows():
            label = row[0].value
            if _looks_like_period_label(label):
                seen_data = True
                if not _is_annual_period_row(label):
                    continue  # quarterly — keep going but don't emit
                if len(row) <= ggy_col:
                    continue
                val = row[ggy_col].value
                if not isinstance(val, (int, float)) or val != val:
                    continue
                period = _period_from_row_label(str(label))
                if period is None:
                    continue
                out.append(MetricFact(
                    state="UK",
                    operator="UK Statewide",
                    vertical=vertical,
                    period=period,
                    metric_name="ggr",
                    value_usd_cents=_gbp_millions_to_usd_cents(float(val)),
                    source_url=source_url,
                ))
            else:
                if seen_data:
                    break
    if _EMIT_NON_REMOTE:
        for sheet_id, (vertical, ggy_col) in _ANNUAL_NON_REMOTE_SHEETS.items():
            if sheet_id not in wb.sheetnames:
                continue
            ws = wb[sheet_id]
            for row in ws.iter_rows():
                label = row[0].value
                if not _is_data_row(label):
                    continue
                if len(row) <= ggy_col:
                    continue
                val = row[ggy_col].value
                if not isinstance(val, (int, float)) or val != val:
                    continue
                period = _period_from_row_label(str(label))
                if period is None:
                    continue
                out.append(MetricFact(
                    state="UK",
                    operator="UK Statewide",
                    vertical=vertical,
                    period=period,
                    metric_name="ggr",
                    value_usd_cents=_gbp_millions_to_usd_cents(float(val)),
                    source_url=source_url,
                ))
    return out


def parse_uk_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse a UKGC quarterly Industry Statistics XLSX.

    Returns one MetricFact(metric_name="ggr") per (sheet, data-row) combination
    that contains a valid GGY figure.  GGY is reported in £ millions; converted
    to value_usd_cents via FX 1.27.

    Period is derived from each data row's label (e.g. "Jul 2025 - Sep 2025").
    The URL-derived period is used only as a fallback when row label parsing fails.
    This means the XLSX's full multi-quarter history is captured correctly.
    """
    url_period = _period_from_url(source_url)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    facts: list[MetricFact] = []

    # Detect annual XLSX: presence of sub-tabs like "10a" identifies the annual
    # publication, which has a completely different sheet layout from the
    # quarterly file. Routing to the annual code path avoids reading random
    # columns out of mis-identified sheets (root cause of the FY-history
    # inflation that pushed the global page total into the trillions).
    if "10a" in wb.sheetnames:
        try:
            return _parse_uk_annual(wb, source_url)
        finally:
            wb.close()

    try:
        # --- Process simple per-sector sheets (2–6) ---
        for sheet_id, (vertical, ggy_col) in _SIMPLE_SHEETS.items():
            if sheet_id not in wb.sheetnames:
                continue
            ws = wb[sheet_id]
            for row in ws.iter_rows():
                label = row[0].value
                if not _is_data_row(label):
                    continue
                if len(row) <= ggy_col:
                    continue
                val = row[ggy_col].value
                if not isinstance(val, (int, float)) or val != val:
                    continue
                period = _period_from_row_label(str(label)) or url_period
                if period is None:
                    continue
                facts.append(MetricFact(
                    state="UK",
                    operator="UK Statewide",
                    vertical=vertical,
                    period=period,
                    metric_name="ggr",
                    value_usd_cents=_gbp_millions_to_usd_cents(float(val)),
                    source_url=source_url,
                ))

        # --- Process sheet 7 (Remote Casino, Betting & Bingo) ---
        if "7" in wb.sheetnames:
            ws = wb["7"]
            for row in ws.iter_rows():
                label = row[0].value
                if not _is_data_row(label):
                    continue
                period = _period_from_row_label(str(label)) or url_period
                if period is None:
                    continue
                for col_idx, vertical in _RCBB_GGY_COLS.items():
                    if len(row) <= col_idx:
                        continue
                    val = row[col_idx].value
                    if not isinstance(val, (int, float)) or val != val:
                        continue
                    facts.append(MetricFact(
                        state="UK",
                        operator="UK Statewide",
                        vertical=vertical,
                        period=period,
                        metric_name="ggr",
                        value_usd_cents=_gbp_millions_to_usd_cents(float(val)),
                        source_url=source_url,
                    ))

    finally:
        wb.close()

    return facts
