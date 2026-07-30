# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Sweden Spelinspektionen quarterly market XLSX parser.

Source:
  https://www.spelinspektionen.se/globalassets/dokument/statistik/
  omsattning-spelmarknaden-per-kvartal.xlsx

Layout (single sheet "Blad1"):
  Row 4 (0-indexed): header — col 0 = label string, remaining cols are either
    a quarterly label like "2024\\njan-mars" or an annual-total label like
    "Helåret 2024".  Annual columns are skipped.
  Rows 5+: one row per vertical; col 0 is the Swedish category name; numeric
    cells are MSEK nettoomsättning (GGR = net gaming revenue).

Vertical mapping  (Swedish label → canonical vertical):
  "Kommersiellt onlinespel"  → igaming          (online casino + poker + bingo)
  "Vadhållning"              → sports-wagering
  "Statligt lotteri"         → lottery
  "Statligt kasinospel"      → land-based-casino
  "Allmännyttiga lotterier"  → lottery           (rikslotterier + hallbingo rows)
  "Landbaserat kommersiellt" → land-based-betting

Quarter → YYYY-MM end-month mapping:
  jan-mars  (Q1) → 03
  apr-juni  (Q2) → 06
  juli-sep  (Q3) → 09
  okt-dec   (Q4) → 12

FX: SEK millions → USD cents  =  value * 1_000_000 * 0.095 * 100
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Union

from openpyxl import load_workbook

from .metrics_model import MetricFact

# ── FX constant ──────────────────────────────────────────────────────────────
_SEK_TO_USD = 0.095  # 1 SEK ≈ 0.095 USD (fixed approximation)

# ── Quarter-header patterns ───────────────────────────────────────────────────
# Handles: "2024\njan-mars", "2024 jan-mars", "2024:1", "Q1 2024",
#          "2024Q1", "2024 Q1"
_QUARTER_MONTH = {
    "jan-mars": "03",
    "q1": "03",
    "1": "03",
    "apr-juni": "06",
    "q2": "06",
    "2": "06",
    "juli-sep": "09",
    "q3": "09",
    "3": "09",
    "okt-dec": "12",
    "q4": "12",
    "4": "12",
}

# Compiled patterns for flexible quarter-header parsing
_YEAR_FIRST_RE = re.compile(
    r"(\d{4})[^\d]*(jan-mars|apr-juni|juli-sep|okt-dec|q[1-4]|[1-4])\b",
    re.IGNORECASE,
)
_QUARTER_FIRST_RE = re.compile(
    r"(q[1-4]|[1-4])[^\d]*(\d{4})",
    re.IGNORECASE,
)
_COLON_RE = re.compile(r"(\d{4}):([1-4])")


def _parse_quarter_header(raw: object) -> str | None:
    """Return 'YYYY-MM' for a quarterly column header, or None to skip it."""
    if raw is None:
        return None
    cell = str(raw).strip().replace("\n", " ")
    lower = cell.lower()

    # Skip annual-total columns ("Helåret YYYY", "Helåret\n2022", etc.)
    if "helår" in lower or "helaret" in lower:
        return None

    # "YYYY:Q" notation  e.g. "2024:1"
    m = _COLON_RE.search(cell)
    if m:
        year, q = m.group(1), m.group(2)
        month = _QUARTER_MONTH.get(q)
        return f"{year}-{month}" if month else None

    # Year-first: "2024 jan-mars", "2024Q1", "2024 Q1"
    m = _YEAR_FIRST_RE.search(lower)
    if m:
        year = m.group(1)
        qkey = m.group(2).lower().replace(" ", "")
        month = _QUARTER_MONTH.get(qkey)
        return f"{year}-{month}" if month else None

    # Quarter-first: "Q1 2024", "1 2024"
    m = _QUARTER_FIRST_RE.search(lower)
    if m:
        qkey = m.group(1).lower()
        year = m.group(2)
        month = _QUARTER_MONTH.get(qkey)
        return f"{year}-{month}" if month else None

    return None


# ── Vertical mapping ──────────────────────────────────────────────────────────
# Checked via substring match (case-insensitive) on the row label.
# Order matters: more specific patterns first.
_VERTICAL_MAP: list[tuple[str, str]] = [
    ("kommersiellt onlinespel", "igaming"),
    ("vadhållning", "sports-wagering"),
    ("statligt lotteri", "lottery"),
    ("statligt kasinospel", "land-based-casino"),
    ("allmännyttiga", "lottery"),
    ("hallbingo", "lottery"),
    ("rikslotterier", "lottery"),
    ("landbaserat kommersiellt", "land-based-betting"),
]


def _map_vertical(label: str) -> str | None:
    """Return canonical vertical string or None if the row should be skipped."""
    low = label.lower()
    for fragment, vertical in _VERTICAL_MAP:
        if fragment in low:
            return vertical
    return None


# ── SEK millions → USD cents ──────────────────────────────────────────────────
def _sek_millions_to_usd_cents(value: float) -> int:
    return int(value * 1_000_000 * _SEK_TO_USD * 100)


# ── Public entry point ────────────────────────────────────────────────────────

def parse_se_xlsx(
    path: Union[Path, bytes, io.IOBase],
    source_url: str,
) -> list[MetricFact]:
    """Parse Sweden Spelinspektionen quarterly market XLSX.

    Parameters
    ----------
    path:
        ``pathlib.Path`` to the .xlsx file on disk, raw ``bytes`` of the
        downloaded file, or any file-like object accepted by openpyxl.
    source_url:
        The URL the file was fetched from; stored verbatim in every
        ``MetricFact.source_url``.

    Returns
    -------
    list[MetricFact]
        One fact per (vertical, quarter) combination found in the workbook.
        Annual-total columns ("Helåret …") are skipped.  Rows that do not
        match any known vertical are silently ignored.
    """
    # Accept path, bytes, or file-like
    if isinstance(path, (bytes, bytearray)):
        source: object = io.BytesIO(path)
    else:
        source = path

    wb = load_workbook(source, read_only=True, data_only=True)
    facts: list[MetricFact] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Collect all rows into a list so we can locate the header row.
            rows = list(ws.iter_rows(values_only=True))

            # ── Locate header row ────────────────────────────────────────────
            # The header row is the first row where col-0 is a non-empty
            # string AND at least one subsequent cell parses as a quarter.
            header_idx: int | None = None
            col_periods: list[str | None] = []

            for ridx, row in enumerate(rows):
                if not row or row[0] is None:
                    continue
                label_cell = str(row[0]).strip()
                if not label_cell:
                    continue
                # Try to parse quarter headers from this row.
                candidate_periods = [
                    _parse_quarter_header(c) for c in row[1:]
                ]
                if any(p is not None for p in candidate_periods):
                    header_idx = ridx
                    col_periods = candidate_periods
                    break

            if header_idx is None:
                # Sheet has no recognisable quarter headers — skip it.
                continue

            # ── Parse data rows ──────────────────────────────────────────────
            for row in rows[header_idx + 1 :]:
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                if not label:
                    continue

                vertical = _map_vertical(label)
                if vertical is None:
                    continue

                data_cells = row[1:]
                for cidx, period in enumerate(col_periods):
                    if period is None:
                        continue  # annual total or unrecognised header
                    if cidx >= len(data_cells):
                        break
                    cell_value = data_cells[cidx]
                    if cell_value is None:
                        continue
                    try:
                        sek_millions = float(cell_value)
                    except (TypeError, ValueError):
                        continue
                    if sek_millions == 0.0:
                        continue  # skip genuine zero cells (e.g. Covid closure)

                    # Quarterly publication → distribute across the 3 months
                    # of the quarter so the monthly chart line is smooth and
                    # comparable to monthly-cadence regulators.
                    quarterly_cents = _sek_millions_to_usd_cents(sek_millions)
                    monthly_cents = quarterly_cents // 3
                    yr, mo = int(period[:4]), int(period[5:7])
                    # period MM is end of quarter (3/6/9/12); spread back 3 mos
                    for offset in range(0, 3):
                        m = mo - offset
                        y = yr
                        if m < 1:
                            m += 12
                            y -= 1
                        facts.append(
                            MetricFact(
                                state="SE",
                                operator="SE Statewide",
                                vertical=vertical,
                                period=f"{y:04d}-{m:02d}",
                                metric_name="ggr",
                                value_usd_cents=monthly_cents,
                                source_url=source_url,
                            )
                        )
    finally:
        wb.close()

    return facts
