# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""
Lightweight Excel summarizer.

Reads the first worksheet of an .xlsx file, finds the first numeric column,
and returns the top-line totals. Good enough for the website "preview" card;
the full data is always one click away in the original file.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models import SummaryRow


def summarise_xlsx(path: Path, max_rows: int = 5) -> list[SummaryRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = []
        for raw in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in raw if c is not None]
            if not cells or len(cells) < 2:
                continue
            label = cells[0]
            value = cells[-1]
            if not label or not value:
                continue
            rows.append(SummaryRow(label=label, value=value))
            if len(rows) >= max_rows:
                break
        return rows
    finally:
        wb.close()
