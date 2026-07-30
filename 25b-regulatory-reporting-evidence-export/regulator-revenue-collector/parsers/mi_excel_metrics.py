# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Michigan Gaming Control Board (MGCB) — GovDelivery attachment Excel parser.

Each monthly MGCB GovDelivery bulletin includes two Excel attachments:
  * ``Internet Gaming - <Month> <Year>.xlsx``
  * ``Internet Sports Betting - <Month> <Year>.xlsx``

Both files follow the same layout:
  * One sheet per calendar year (``Internet Gaming 2024``, etc.).
  * Row 2  — Operator names (one operator spans 3 columns: GGR, AGR, Tax).
  * Row 6  — Column headers: ``Month``, ``Gross Receipts``, ``AGR``, ``Tax``,
             repeated per operator.  At position ``totals_col`` (the last
             operator block) the headers become ``Total Gross Receipts``,
             ``Total Adjusted Gross Receipts``, ``Total Internet Gaming State
             Tax / Payment`` (or the sports-wagering equivalents).
  * Rows 7–18 — Monthly data rows (January … December).  Rows beyond the
               latest published month are empty / None.
  * Row 19  — Annual total row (label = ``Total``).

For the sports-betting file the ``Total`` block additionally has a
``Total Handle`` column immediately before ``Total Gross Sports Betting
Receipts``.

Attachment URLs for confirmed files:
  Internet Gaming:
    2025-01-21  https://content.govdelivery.com/attachments/MIGCB/2025/01/21/file_attachments/3138414/Internet%20Gaming%20-%20December%202024.xlsx
    2026-03-17  https://content.govdelivery.com/attachments/MIGCB/2026/03/17/file_attachments/3586362/Internet%20Gaming%20-%20February%202026.xlsx
  Internet Sports Betting:
    2025-01-21  https://content.govdelivery.com/attachments/MIGCB/2025/01/21/file_attachments/3138413/Internet%20Sports%20Betting%20-%20December%202024.xlsx
    2026-03-17  https://content.govdelivery.com/attachments/MIGCB/2026/03/17/file_attachments/3586360/Internet%20Sports%20Betting%20-%20February%202026.xlsx
"""
from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import openpyxl.worksheet.worksheet

from .metrics_model import MetricFact

logger = logging.getLogger(__name__)

_MONTHS: dict[str, str] = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
}

_CENTS_PER_DOLLAR = 100


def _to_cents(value: object) -> int | None:
    """Convert a raw Excel cell value (float/int) to integer cents."""
    if value is None:
        return None
    try:
        fval: float = float(value)  # type: ignore[arg-type]
        return int(round(fval * _CENTS_PER_DOLLAR))
    except (TypeError, ValueError):
        return None


def _find_totals_col(ws: "openpyxl.worksheet.worksheet.Worksheet") -> int | None:
    """Return the 0-based column index of the *Total* operator block (row 2).

    The MGCB spreadsheet has a fixed layout: one operator block per three
    columns (GGR, AGR, Tax), and the last block is labelled ``All Internet
    Gaming Operators`` / ``All Internet Sports Betting Operators``.
    """
    for col_idx, cell in enumerate(ws[2]):  # row 2 = operators row
        val = cell.value
        if val is not None and "all" in str(val).lower():
            return col_idx
    return None


def _find_totals_col_sports(ws: "openpyxl.worksheet.worksheet.Worksheet") -> int | None:
    """Same as ``_find_totals_col`` but for sports-betting sheets.

    The sports Excel header row has ``Total Handle`` as the first column
    of the All-operators block, so we search for the ``Total Handle`` label
    in row 6 (0-based index 5) as a secondary anchor.
    """
    col = _find_totals_col(ws)
    if col is not None:
        return col
    # Fallback: look for 'Total Handle' in row 6 (header row)
    for col_idx, cell in enumerate(ws[6]):
        if cell.value and "total handle" in str(cell.value).lower():
            return col_idx
    return None


def parse_mi_igaming_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse an MGCB Internet Gaming Excel attachment.

    Emits MetricFacts with:
      vertical = 'igaming'
      metric_name in {'ggr', 'agr', 'tax_paid'}

    Each published file may contain multiple sheets (one per calendar year).
    All sheets with a recognised year-label are parsed so that a single call
    on the Feb-2026 file extracts both the full 2025 history and the
    Jan-Feb 2026 data.

    Args:
        path:       Local path to the downloaded .xlsx file.
        source_url: Canonical URL used as provenance on every emitted fact.

    Returns:
        List of MetricFact instances; empty list on any parse error.
    """
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("mi_excel: cannot open %s: %s", path, exc)
        return []

    facts: list[MetricFact] = []

    for sheet_name in wb.sheetnames:
        # Only process sheets named like "Internet Gaming YYYY"
        parts = sheet_name.split()
        if not parts:
            continue
        year_str = parts[-1]
        try:
            year = int(year_str)
        except ValueError:
            continue
        if year < 2021 or year > 2030:
            continue

        ws = wb[sheet_name]
        totals_col = _find_totals_col(ws)
        if totals_col is None:
            logger.warning("mi_excel: no All-operators col in sheet '%s' of %s", sheet_name, path)
            continue

        # Column layout at totals_col (0-based):
        #   totals_col+0: Total Gross Receipts (GGR)
        #   totals_col+1: Total Adjusted Gross Receipts (AGR)
        #   totals_col+2: Total Internet Gaming State Tax / Payment
        for row in ws.iter_rows(min_row=7, values_only=True):
            month_cell = row[0]
            if month_cell is None:
                break
            month_name = str(month_cell).strip().lower()
            if month_name == "total":
                break  # stop at annual total row
            month_num = _MONTHS.get(month_name)
            if not month_num:
                continue

            period = f"{year:04d}-{month_num}"

            ggr_cents = _to_cents(row[totals_col] if totals_col < len(row) else None)
            agr_cents = _to_cents(row[totals_col + 1] if totals_col + 1 < len(row) else None)
            tax_cents = _to_cents(row[totals_col + 2] if totals_col + 2 < len(row) else None)

            if ggr_cents is not None and ggr_cents > 0:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="igaming",
                    period=period,
                    metric_name="ggr",
                    value_usd_cents=ggr_cents,
                    source_url=source_url,
                ))
            if agr_cents is not None and agr_cents != 0:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="igaming",
                    period=period,
                    metric_name="agr",
                    value_usd_cents=agr_cents,
                    source_url=source_url,
                ))
            if tax_cents is not None and tax_cents > 0:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="igaming",
                    period=period,
                    metric_name="tax_paid",
                    value_usd_cents=tax_cents,
                    source_url=source_url,
                ))

    logger.info("mi_excel: parsed %d igaming facts from %s", len(facts), path)
    return facts


def parse_mi_sports_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse an MGCB Internet Sports Betting Excel attachment.

    Emits MetricFacts with:
      vertical = 'sports-wagering'
      metric_name in {'handle', 'ggr', 'agr', 'tax_paid'}

    Column layout at the All-operators block (0-based):
      totals_col+0: Total Handle
      totals_col+1: Total Gross Sports Betting Receipts (GGR)
      totals_col+2: Total Adjusted Gross Sports Betting Receipts (AGR)
      totals_col+3: Total Internet Sports Betting State Tax / Payment

    Args:
        path:       Local path to the downloaded .xlsx file.
        source_url: Canonical URL used as provenance on every emitted fact.

    Returns:
        List of MetricFact instances; empty list on any parse error.
    """
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("mi_excel: cannot open %s: %s", path, exc)
        return []

    facts: list[MetricFact] = []

    for sheet_name in wb.sheetnames:
        parts = sheet_name.split()
        if not parts:
            continue
        year_str = parts[-1]
        try:
            year = int(year_str)
        except ValueError:
            continue
        if year < 2021 or year > 2030:
            continue

        ws = wb[sheet_name]
        totals_col = _find_totals_col_sports(ws)
        if totals_col is None:
            logger.warning(
                "mi_excel: no All-operators col in sheet '%s' of %s", sheet_name, path,
            )
            continue

        # Confirm header layout: verify that column totals_col+1 contains 'gross'
        # (i.e., the handle column is at totals_col+0 as expected for sports).
        header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
        has_handle = (
            totals_col < len(header_row)
            and header_row[totals_col] is not None
            and "handle" in str(header_row[totals_col]).lower()
        )
        if has_handle:
            handle_off, ggr_off, agr_off, tax_off = 0, 1, 2, 3
        else:
            # Older layout (rare): no handle column, starts directly at GGR
            handle_off, ggr_off, agr_off, tax_off = None, 0, 1, 2

        for row in ws.iter_rows(min_row=7, values_only=True):
            month_cell = row[0]
            if month_cell is None:
                break
            month_name = str(month_cell).strip().lower()
            if month_name == "total":
                break
            month_num = _MONTHS.get(month_name)
            if not month_num:
                continue

            period = f"{year:04d}-{month_num}"

            def _get(offset: int | None) -> int | None:
                if offset is None:
                    return None
                col = totals_col + offset
                return _to_cents(row[col] if col < len(row) else None)

            handle_cents = _get(handle_off)
            ggr_cents = _get(ggr_off)
            agr_cents = _get(agr_off)
            tax_cents = _get(tax_off)

            if handle_cents is not None and handle_cents > 0:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="sports-wagering",
                    period=period,
                    metric_name="handle",
                    value_usd_cents=handle_cents,
                    source_url=source_url,
                ))
            if ggr_cents is not None:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="sports-wagering",
                    period=period,
                    metric_name="ggr",
                    value_usd_cents=ggr_cents,
                    source_url=source_url,
                ))
            if agr_cents is not None:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="sports-wagering",
                    period=period,
                    metric_name="agr",
                    value_usd_cents=agr_cents,
                    source_url=source_url,
                ))
            if tax_cents is not None and tax_cents > 0:
                facts.append(MetricFact(
                    state="MI",
                    operator="MI Statewide",
                    vertical="sports-wagering",
                    period=period,
                    metric_name="tax_paid",
                    value_usd_cents=tax_cents,
                    source_url=source_url,
                ))

    logger.info("mi_excel: parsed %d sports facts from %s", len(facts), path)
    return facts
