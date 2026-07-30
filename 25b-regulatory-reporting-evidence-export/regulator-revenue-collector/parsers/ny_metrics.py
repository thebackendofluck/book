# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""NY State Gaming Commission per-operator XLSX parser."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import openpyxl

from .metrics_model import MetricFact

_HEADER_SCAN_ROWS = 30
_BANNER_LOOKUP_ROWS = 4

_METRIC_MAP: dict[tuple[str, str], str] = {
    ("slot/etg's",         "ggr"):    "ggr",      # commercial casino slot/etg
    ("slot/etgs",          "ggr"):    "ggr",
    ("table games",        "ggr"):    "table_ggr",
    ("table games",        "drop"):   "handle",
    ("poker tables",       "ggr"):    "poker_ggr",
    ("sports wagering",    "ggr"):    "ggr",
    ("sports wagering",    "handle"): "handle",
    ("mobile sports wagering", "handle"): "handle",
    ("mobile sports wagering", "ggr"):    "ggr",
    ("mobile sports wagering",
        "to platform provider"):     "net_revenue_platform",
    ("mobile sports wagering",
        "to education"):             "tax_paid",
    # Prior-period fines/penalties get added into the "to education" cell
    # when the NYGC records a retroactive adjustment.  We capture this
    # column so we can subtract it out of tax_paid later.
    ("mobile sports wagering",
        "fines & penalties"):        "prior_period_adj",
}
# "Total GGR" is a derived sum column — do NOT map it to "ggr".
# Mapping it caused col 23 (Total GGR = 0 in the current-FY placeholder sheet) to
# win the dedup race before the real Slot/ETG GGR column could be stored.
_HEADER_ONLY: dict[str, str] = {}


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def _find_header_row(ws):
    for row_idx in range(1, _HEADER_SCAN_ROWS + 1):
        if _norm(ws.cell(row=row_idx, column=1).value) == "month":
            return row_idx
    return None


def _section_for_column(ws, header_row, col):
    if header_row <= 1:
        return ""
    start = header_row - 1
    stop = max(1, header_row - _BANNER_LOOKUP_ROWS)
    for r in range(start, stop - 1, -1):
        for dc in range(0, 8):
            target_col = col - dc
            if target_col < 1:
                break
            v = _norm(ws.cell(row=r, column=target_col).value)
            if not v:
                continue
            if v in {"mobile sports", "mobile sports wagering"}:
                return "mobile sports wagering"
            if v in {"slot/etg's", "slot/etgs", "table games",
                     "poker tables", "sports wagering", "sports\nwagering"}:
                return v.replace("\n", " ")
    return ""


def _build_column_map(ws, header_row):
    out = {}
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        header = _norm(ws.cell(row=header_row, column=col).value)
        if not header:
            continue
        if header in _HEADER_ONLY:
            out[col] = _HEADER_ONLY[header]
            continue
        section = _section_for_column(ws, header_row, col)
        canon = _METRIC_MAP.get((section, header))
        if canon:
            out[col] = canon
    return out


def _to_cents(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        if s in {"", "-", "N/A", "n/a"}:
            return None
        try:
            v = float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return int(round(float(v) * 100))
    return None


def _iter_data_rows(ws, header_row):
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        first = row[0] if row else None
        if isinstance(first, datetime):
            yield first, row
        elif first is None:
            continue
        else:
            return


def parse_ny_xlsx(path: Path, operator: str, vertical: str, source_url: str) -> list[MetricFact]:
    """Extract monthly $$$ history from one NY per-operator Excel file."""
    # Guard: NY occasionally serves a 404 HTML page with .xlsx in the URL.
    # openpyxl will raise "File is not a zip file" — skip cleanly.
    try:
        with open(path, "rb") as fp:
            magic = fp.read(2)
        if magic != b"PK":
            return []
    except OSError:
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    facts: list[MetricFact] = []
    seen: set[tuple[str, str]] = set()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = _find_header_row(ws)
            if header_row is None:
                continue
            col_map = _build_column_map(ws, header_row)
            if not col_map:
                continue
            for dt, row in _iter_data_rows(ws, header_row):
                period = f"{dt.year:04d}-{dt.month:02d}"
                # Collect all mapped metric values for this row first so we
                # can apply cross-column corrections before storing facts.
                row_values: dict[str, int] = {}
                for col, metric_name in col_map.items():
                    if col - 1 >= len(row):
                        continue
                    cents = _to_cents(row[col - 1])
                    if cents is None:
                        continue
                    # Keep first non-None value per metric for this row
                    # (mirrors the cross-period dedup in `seen`).
                    if metric_name not in row_values:
                        row_values[metric_name] = cents

                # Correction: when the NYGC includes prior-period fines in
                # the "Net Revenue to Education" cell, that cell holds
                #   monthly_tax + prior_period_adj
                # rather than the monthly tax alone.  Subtract the adjustment
                # to recover the correct monthly tax_paid figure.
                # Jan-2025 BetMGM: $30.66M → $7.55M after removing $23.11M adj.
                adj = row_values.get("prior_period_adj", 0)
                if adj and "tax_paid" in row_values:
                    row_values["tax_paid"] -= adj

                # Do not persist the internal helper column.
                row_values.pop("prior_period_adj", None)

                for metric_name, cents in row_values.items():
                    key = (period, metric_name)
                    if key in seen:
                        continue
                    seen.add(key)
                    facts.append(MetricFact(
                        state="NY",
                        operator=operator,
                        vertical=vertical,
                        period=period,
                        metric_name=metric_name,
                        value_usd_cents=cents,
                        source_url=source_url,
                    ))
    finally:
        wb.close()
    facts.sort(key=lambda f: (f.period, f.metric_name))
    return facts
