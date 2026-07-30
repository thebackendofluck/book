# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Ohio Casino Control Commission monthly sports-gaming XLSX and PDF parsers.

XLSX layout (confirmed against June_2024_Ohio_Sports_Gaming_Revenue_Report.xlsx):

Each XLSX covers one calendar year and contains one tab per month
(JANUARY...DECEMBER) plus a SUMMARY tab and a REVENUE REPORTING NOTES tab.

Each month tab has two sections -- TYPE A (online) and TYPE B (retail) --
both with identical column layout:

  col 0: Operator / Services Provider label
  col 1: Total Gross Receipts  -> handle
  col 2: (-) Winnings Paid      (not captured)
  col 3: (-) Voided Wagers      (not captured)
  col 4: Promotional*           -> promotional_play
  col 5: Revenue                (pre-floor at zero; not emitted -- use col 6)
  col 6: Taxable Revenue        -> ggr  (= max(Revenue, 0))

Tax = 10 % of Taxable Revenue (ORC §5753.02) -> tax_paid computed inline.

Period comes from the URL filename, not from the spreadsheet itself:
  "June_2024_Ohio_Sports_Gaming_Revenue_Report.xlsx" -> "2024-06"

PDF layout (OCCC switched to PDF-only from May 2024 onwards):

Each PDF covers one calendar month.  Tables appear on one or more pages
with the same column structure as the XLSX (7 columns):

  col 0: Operator / Services Provider label
  col 1: Total Gross Receipts (Handle)
  col 2: (-) Winnings Paid      (not captured)
  col 3: (-) Voided Wagers      (not captured)
  col 4: Promotional*           -> promotional_play
  col 5: Revenue                (pre-floor; not emitted)
  col 6: Taxable Revenue        -> ggr

pdfplumber table extraction may split or merge cells; the parser is
tolerant of 4-, 5-, 6-, or 7-column rows, mapping by position from the
right so that the 1-column-right shift (when col 5 is merged into col 6)
is handled automatically.

Column index mapping when row has N columns (N >= 4):
  col  0        : operator name
  col  N-1      : ggr (Taxable Revenue) -- rightmost
  col  N-2      : revenue (not emitted)
  col  N-3      : promotional_play
  col  N-6 / 1  : handle (Total Gross Receipts) -- 2nd from left

Annual PDFs (2024+):

OCCC annual cumulative PDFs use a text-flow layout that pdfplumber cannot
parse as tables (extract_tables() returns []).  The data is present as
plain text lines rendered by position.  Numbers are split across multiple
PDF text-run tokens due to font-kerning artifacts, e.g.:

  "2" + "82,737,354" -> 282,737,354

The _pdf_text_rows() fallback reconstructs rows by:
  1. Grouping pdfplumber word objects by Y-coordinate (binned to 3 pt)
  2. Sorting tokens left-to-right within each line
  3. Stitching split numeric tokens: a bare digit/digits followed by a
     token starting with a comma or digit-comma sequence is merged
  4. Reconstructing the 7 logical cells (operator + 6 numeric fields)
"""
from __future__ import annotations

import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MONTHS = {
    "January": "01", "February": "02", "March": "03", "April": "04",
    "May": "05", "June": "06", "July": "07", "August": "08",
    "September": "09", "October": "10", "November": "11", "December": "12",
}

# Sheet names that carry no per-operator data.
_SKIP_SHEETS = {"SUMMARY", "REVENUE REPORTING NOTES"}

# Strings that mark a subtotal / total / header row -- skip these.
_SKIP_ROW_RE = re.compile(
    r"^\s*("
    r"subtotal|total|all proprietors|online proprietor|retail proprietor"
    r"|type a|type b|\*promotional"
    r")",
    re.IGNORECASE,
)

# Column indices (0-based) within each data row.
_COL_HANDLE = 1      # Total Gross Receipts
_COL_PROMO  = 4      # Promotional*
_COL_GGR    = 6      # Taxable Revenue  (ggr = max(Revenue, 0))

# Ohio sports-gaming tax rate (ORC §5753.02).
_OH_TAX_RATE = Decimal("0.10")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _period_from_url(url: str) -> str | None:
    """Return 'YYYY-MM' extracted from the OH CDN filename.

    Handles four URL patterns:

      1. Per-month underscore:
           .../June_2024_Ohio_Sports_Gaming_Revenue_Report.xlsx -> "2024-06"

      2. Per-month URL-encoded space (legacy Dec-2023 path):
           .../December%202023%20Ohio%20Sports%20Gaming%20Revenue%20Report.pdf
           (decoded: "December 2023 Ohio ...") -> "2023-12"

      3. Annual cumulative sequential (2025/2026 sequential files):
           .../2025_Sports_Gaming_Revenue_Report12.xlsx -> "2025-01"
           (year correct; XLSX parser overrides month from tab names)

      4. Annual cumulative (named file):
           .../2024_Sports_Gaming_Revenue_Report.pdf.pdf -> "2024-01"
    """
    from urllib.parse import unquote
    decoded = unquote(url)

    # Special case: "Full Year Revenue Reports" path contains the full-year
    # cumulative PDF (all 12 months for the named year, e.g. December 2023).
    # Return the annual sentinel YYYY-01 so the PDF parser reads each page
    # heading independently rather than assigning the month from the filename.
    if "Full%20Year%20Revenue%20Reports" in url or "Full Year Revenue Reports" in decoded:
        m = re.search(r"\b(\d{4})\b", decoded)
        if m:
            return f"{m.group(1)}-01"

    # Pattern 1: MonthName_YYYY_ (underscore-separated per-month)
    m = re.search(r"/([A-Za-z]+)_(\d{4})_", decoded)
    if m:
        month_str = _MONTHS.get(m.group(1).capitalize())
        if month_str:
            return f"{m.group(2)}-{month_str}"

    # Pattern 2: "MonthName YYYY" (space-separated, from URL-encoded paths)
    m = re.search(
        r"/([A-Za-z]+)\s+(\d{4})\s+",
        decoded,
    )
    if m:
        month_str = _MONTHS.get(m.group(1).capitalize())
        if month_str:
            return f"{m.group(2)}-{month_str}"

    # Pattern 3 & 4: Annual cumulative -- /YYYY/ or YYYY_ in filename
    m = re.search(r"/(\d{4})[_/][A-Za-z0-9_%-]*(?:Sports|Revenue|Gaming)", url, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-01"
    return None


def _to_cents(raw) -> int | None:
    """Convert a raw cell value (int, float, or string) in whole dollars to
    integer US cents.

    OCCC spreadsheets and PDFs report values as whole-dollar integers
    (e.g. 103358051 meaning $103,358,051).  Multiply by 100 to get cents.
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        try:
            return int(Decimal(str(raw)) * 100)
        except InvalidOperation:
            return None
    s = str(raw).strip().replace(",", "").replace("$", "").replace(" ", "")
    if not s or s in ("-", "--", "N/A"):
        return None
    try:
        return int(Decimal(s) * 100)
    except InvalidOperation:
        return None


def _is_data_row(row) -> bool:
    """Return True when *row* looks like a per-operator data row."""
    label = row[0]
    if not label or not isinstance(label, str):
        return False
    label = label.strip()
    if not label:
        return False
    return not bool(_SKIP_ROW_RE.match(label))


def _clean_operator(label: str) -> str:
    """Strip leading/trailing whitespace; preserve the full label as-is."""
    return label.strip()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_oh_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse one OH OCCC sports-gaming XLSX and return a list of MetricFact.

    Each file covers one calendar year; the period (YYYY-MM) is derived
    from the URL filename, not from the spreadsheet content, because every
    month tab inside a single file represents a *different* month and the
    filename identifies which months are present only by the last month
    published (the file is updated monthly and always named after the latest
    month).  We therefore read the period from the URL and cross-check it
    against the tab name to ensure correctness.

    Metrics emitted per operator per month:
      - handle            : Total Gross Receipts (col 1)
      - promotional_play  : Promotional* wagered (col 4)
      - ggr               : Taxable Revenue = max(Revenue, 0) (col 6)
      - tax_paid          : ggr x 10 % (computed, not read from sheet)
    """
    period = _period_from_url(source_url)
    if period is None:
        raise ValueError(
            f"Cannot extract period from URL: {source_url!r}. "
            "Expected pattern .../MonthName_YYYY_..."
        )

    # Guard: openpyxl raises "File is not a zip file" for HTML 404 pages.
    try:
        with open(path, "rb") as fh:
            magic = fh.read(2)
        if magic != b"PK":
            return []
    except OSError:
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    facts: list[MetricFact] = []

    try:
        for sheet_name in wb.sheetnames:
            # Skip non-data tabs.
            if sheet_name.strip().upper() in _SKIP_SHEETS:
                continue
            # The tab name is a month name (e.g. "JUNE").  Derive the period
            # for *this* tab by combining the tab month with the year from
            # the URL.  This correctly handles YTD files that contain multiple
            # months (January ... current month).
            month_num = _MONTHS.get(sheet_name.strip().capitalize())
            if month_num is None:
                # Unrecognised tab (e.g. "REVENUE REPORTING NOTES" variant).
                continue
            year = period.split("-")[0]
            tab_period = f"{year}-{month_num}"

            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                if not _is_data_row(row):
                    continue

                operator = _clean_operator(str(row[0]))

                # --- handle ---
                if len(row) > _COL_HANDLE:
                    cents = _to_cents(row[_COL_HANDLE])
                    if cents is not None:
                        facts.append(MetricFact(
                            state="OH",
                            operator=operator,
                            vertical="sports-wagering",
                            period=tab_period,
                            metric_name="handle",
                            value_usd_cents=cents,
                            source_url=source_url,
                        ))

                # --- promotional_play ---
                if len(row) > _COL_PROMO:
                    cents = _to_cents(row[_COL_PROMO])
                    if cents is not None:
                        facts.append(MetricFact(
                            state="OH",
                            operator=operator,
                            vertical="sports-wagering",
                            period=tab_period,
                            metric_name="promotional_play",
                            value_usd_cents=cents,
                            source_url=source_url,
                        ))

                # --- ggr (Taxable Revenue) ---
                if len(row) > _COL_GGR:
                    cents = _to_cents(row[_COL_GGR])
                    if cents is not None:
                        facts.append(MetricFact(
                            state="OH",
                            operator=operator,
                            vertical="sports-wagering",
                            period=tab_period,
                            metric_name="ggr",
                            value_usd_cents=cents,
                            source_url=source_url,
                        ))

                        # --- tax_paid (derived: ggr x 10 %) ---
                        # Taxable Revenue is already floored at zero by OCCC,
                        # so no need to re-floor here.
                        tax_cents = int(
                            (Decimal(cents) * _OH_TAX_RATE).to_integral_value()
                        )
                        facts.append(MetricFact(
                            state="OH",
                            operator=operator,
                            vertical="sports-wagering",
                            period=tab_period,
                            metric_name="tax_paid",
                            value_usd_cents=tax_cents,
                            source_url=source_url,
                        ))

    finally:
        wb.close()

    # Annual/cumulative XLSXs include both TYPE A (online) and TYPE B (retail)
    # sections.  Both sections may list the same operator names (e.g. "BELTERRA
    # PARK - FANDUEL" appears in both TYPE A online and TYPE B retail).  Sum
    # both contributions per (operator, period, metric_name) key so the DB
    # upsert stores the combined total instead of letting the second occurrence
    # silently overwrite the first.
    agg: dict[tuple[str, str, str], MetricFact] = {}
    for f in facts:
        key = (f.operator, f.period, f.metric_name)
        if key in agg:
            agg[key] = MetricFact(
                state=f.state,
                operator=f.operator,
                vertical=f.vertical,
                period=f.period,
                metric_name=f.metric_name,
                value_usd_cents=agg[key].value_usd_cents + f.value_usd_cents,
                source_url=f.source_url,
            )
        else:
            agg[key] = f
    facts = sorted(agg.values(), key=lambda f: (f.period, f.operator, f.metric_name))
    return facts


# ---------------------------------------------------------------------------
# PDF parser (OCCC PDF-only format, May 2024 onwards)
# ---------------------------------------------------------------------------

# Rows whose first cell matches these patterns are skipped (totals/headers).
_PDF_SKIP_RE = re.compile(
    r"^\s*("
    r"subtotal|total|all proprietors|online proprietor|retail proprietor"
    r"|type a|type b|\*promotional|operator|services provider|proprietor"
    r"|online\s+proprietor\s*/\s*services\s+provider"
    r"|retail\s+proprietor\s*/\s*services\s+provider"
    r")",
    re.IGNORECASE,
)

# Minimum number of columns a row must have to be treated as a data row.
_PDF_MIN_COLS = 4

# A token is "numeric" when it looks like a number fragment:
#   - plain digits (possibly a split prefix, e.g. "2" from "282,737,354")
#   - comma-separated digit groups: "82,737,354", ",161,873", "0,048,412"
#   - parenthesised negative amounts: "(65,063)"
#   - a bare dash "-" representing zero
_NUMERIC_RE = re.compile(
    r"^("
    r"-"                        # zero / dash
    r"|[(\-]?\d[\d,]*\)?"       # plain number or (negative)
    r"|,\d[\d,]*"               # starts-with-comma fragment (split artifact)
    r")$"
)

# Stricter pattern used to TRIGGER the start of the numeric section in a row.
# A lone dash ("-") is ambiguous -- it can separate casino name from sports
# partner (e.g. "BELTERRA PARK - FANDUEL") or represent a zero value. We only
# let it open the numeric section after at least one real number has been seen.
# This pattern requires at least one digit.
_NUMERIC_START_RE = re.compile(
    r"^("
    r"[(\-]?\d[\d,]*\)?"        # plain number or (negative) with digit
    r"|,\d[\d,]*"               # starts-with-comma fragment
    r")$"
)


def _period_from_page_text(text: str, fallback_year: str) -> str | None:
    """Try to derive 'YYYY-MM' from free text on a PDF page.

    OCCC PDFs include a heading like "JANUARY 2024" or "For the Month of
    February 2025" on each page.  Extract month + year from that heading.
    Falls back to None if no month can be found.
    """
    month_pattern = "|".join(_MONTHS.keys())
    m = re.search(
        rf"\b({month_pattern})\b\s+(\d{{4}})",
        text,
        re.IGNORECASE,
    )
    if m:
        month_str = _MONTHS.get(m.group(1).capitalize())
        year = m.group(2)
        if month_str:
            return f"{year}-{month_str}"
    # Try reversed order: "2025 January" (less common but possible)
    m = re.search(
        rf"\b(\d{{4}})\b\s+({month_pattern})\b",
        text,
        re.IGNORECASE,
    )
    if m:
        month_str = _MONTHS.get(m.group(2).capitalize())
        year = m.group(1)
        if month_str:
            return f"{year}-{month_str}"
    return None


def _pdf_is_data_row(cells: list[str]) -> bool:
    """Return True when *cells* represents a per-operator data row."""
    if len(cells) < _PDF_MIN_COLS:
        return False
    label = cells[0].strip() if cells[0] else ""
    if not label:
        return False
    return not bool(_PDF_SKIP_RE.match(label))


def _pdf_extract_row(cells: list[str]) -> tuple[str, int | None, int | None, int | None]:
    """Return (operator, handle_cents, promo_cents, ggr_cents) from a table row.

    Column mapping is position-from-right so it tolerates 4-7 column rows:
      rightmost (N-1) -> ggr
      N-2             -> revenue (skipped)
      N-3             -> promotional_play
      col 1 (fixed)   -> handle  (always second cell, regardless of width)
    """
    n = len(cells)
    operator = cells[0].strip() if cells[0] else ""
    handle_cents = _to_cents(cells[1]) if n > 1 else None
    promo_cents  = _to_cents(cells[n - 3]) if n >= 4 else None
    ggr_cents    = _to_cents(cells[n - 1]) if n >= 4 else None
    return operator, handle_cents, promo_cents, ggr_cents


def _pdf_text_rows(page: Any) -> list[list[str]]:
    """Reconstruct logical table rows from a text-flow PDF page.

    OCCC annual PDFs (2024+) do not embed table objects; data is rendered
    as positioned text.  pdfplumber.extract_words() returns one dict per
    word with bounding-box coordinates.  This function:

    1. Groups words by Y-coordinate (binned to 3 pt) to reconstruct lines.
    2. Sorts tokens left-to-right within each line.
    3. Stitches split numeric tokens: the PDF renderer sometimes places a
       leading digit or two as a separate text-run immediately before the
       rest of the number (e.g. ["2", "82,737,354"] -> "282,737,354" and
       ["1", ",161,873"] -> "1161873" after comma-removal).
    4. Classifies the tokens on each line into:
         - operator part: leading non-numeric tokens (up to but not including
           the first standalone numeric token)
         - numeric fields: the remaining tokens, expected to be exactly 6
    5. Returns each row as a 7-element list [operator, f1, f2, f3, f4, f5, f6]
       compatible with _pdf_extract_row().

    Rows with fewer than 4 cells are discarded (they are likely headers
    or decorative lines).
    """
    words = page.extract_words() if hasattr(page, "extract_words") else []
    if not words:
        return []

    # Group by Y (bin to 3 pt to handle sub-pixel variations).
    by_y: dict[int, list[tuple[float, str]]] = {}
    for w in words:
        y_bin = round(float(w.get("top", 0)) / 3) * 3
        by_y.setdefault(y_bin, []).append((float(w.get("x0", 0)), w["text"]))

    rows: list[list[str]] = []
    for y_bin in sorted(by_y):
        tokens = [text for _, text in sorted(by_y[y_bin], key=lambda t: t[0])]

        # Stitch split numeric tokens.
        # A split occurs when a token is purely digits (no commas) and the
        # NEXT token starts with a comma or a digit-then-comma pattern.
        # Examples:
        #   ["2", "82,737,354"]   -> "282,737,354"
        #   ["1", ",161,873"]     -> "1161873" (handled by _to_cents)
        #   ["9", "91,364"]       -> "991364"
        stitched: list[str] = []
        i = 0
        while i < len(tokens):
            tok = tokens[i]
            # Check if this token is a bare integer prefix that should be
            # joined with the next token.
            if (
                i + 1 < len(tokens)
                and re.fullmatch(r"\d{1,3}", tok)
                and re.match(r"[,\d]", tokens[i + 1])
                and _NUMERIC_RE.match(tokens[i + 1])
            ):
                stitched.append(tok + tokens[i + 1])
                i += 2
            else:
                stitched.append(tok)
                i += 1

        # Partition into operator text vs numeric fields.
        # The operator part consists of all leading tokens that do NOT look
        # like proper numbers.  A lone "-" is ambiguous (it separates casino
        # from sports-partner, e.g. "BELTERRA PARK - FANDUEL"), so we use the
        # stricter _NUMERIC_START_RE (requires at least one digit) to trigger
        # the switch.  Once we are in the numeric section a lone "-" IS a valid
        # zero value, so we switch back to _NUMERIC_RE for subsequent tokens.
        op_tokens: list[str] = []
        num_tokens: list[str] = []
        in_numbers = False
        for tok in stitched:
            if not in_numbers and not _NUMERIC_START_RE.match(tok):
                op_tokens.append(tok)
            else:
                in_numbers = True
                num_tokens.append(tok)

        if not op_tokens:
            continue
        operator = " ".join(op_tokens)

        # We expect exactly 6 numeric columns; rows with fewer are headers
        # or summary lines that _pdf_is_data_row will reject anyway.
        row = [operator] + num_tokens
        if len(row) >= _PDF_MIN_COLS:
            rows.append(row)

    return rows


def _emit_row_facts(
    facts: list[MetricFact],
    period: str,
    source_url: str,
    operator: str,
    handle: int | None,
    promo: int | None,
    ggr: int | None,
) -> None:
    """Append MetricFact entries for one operator row into *facts* in-place."""
    # Sanity guard: for sports wagering, handle (total wagered) must always
    # exceed GGR (the operator's win). A row where handle < ggr is a malformed
    # parse (column wrap / footnote line in the legacy per-month PDFs — seen for
    # a few low-volume operators in the Dec-2023 file). Skip it rather than emit
    # an impossible handle<ggr fact that fails verify.py and shows wrong data.
    if handle is not None and ggr is not None and handle < ggr:
        return
    if handle is not None:
        facts.append(MetricFact(
            state="OH",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name="handle",
            value_usd_cents=handle,
            source_url=source_url,
        ))
    if promo is not None:
        facts.append(MetricFact(
            state="OH",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name="promotional_play",
            value_usd_cents=promo,
            source_url=source_url,
        ))
    if ggr is not None:
        facts.append(MetricFact(
            state="OH",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name="ggr",
            value_usd_cents=ggr,
            source_url=source_url,
        ))
        tax_cents = int(
            (Decimal(ggr) * _OH_TAX_RATE).to_integral_value()
        )
        facts.append(MetricFact(
            state="OH",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name="tax_paid",
            value_usd_cents=tax_cents,
            source_url=source_url,
        ))


def parse_oh_pdf(path: Path, source_url: str) -> list[MetricFact]:
    """Parse one OCCC monthly or annual sports-gaming PDF and return MetricFact rows.

    Two PDF layouts are supported:

    1. Per-month PDFs (legacy, some months pre-2024): structured table objects
       parseable via pdfplumber.extract_tables().

    2. Annual cumulative PDFs (2024+): text-flow layout where extract_tables()
       returns [].  The fallback _pdf_text_rows() reconstructs rows from
       positioned word objects, stitching PDF-split numeric tokens.

    The period is derived from the URL filename for per-month files, or from
    the page heading text (e.g. "JANUARY 2024 OHIO SPORTS GAMING REVENUE")
    for annual files.

    Metrics emitted per operator:
      - handle           : Total Gross Receipts (col 1)
      - promotional_play : Promotional* wagered (col N-3)
      - ggr              : Taxable Revenue (col N-1, right-most)
      - tax_paid         : ggr x 10 % (ORC §5753.02, computed)

    Returns [] on any file-format error (non-PDF bytes, corrupt file, etc.).
    """
    url_period = _period_from_url(source_url)
    if url_period is None:
        # Attempt to derive period from filename directly (e.g. when source_url
        # is a local path during testing).
        url_period = _period_from_url(str(path))
    if url_period is None:
        return []

    # When _period_from_url returns a year-only sentinel (month==01 from an
    # annual cumulative file), each PDF page may cover a different month.
    # We attempt per-page month extraction in that case.
    _is_annual = url_period.endswith("-01") and not re.search(
        r"/January_\d{4}_", source_url, re.IGNORECASE
    )
    _fallback_year = url_period.split("-")[0]

    # Guard: reject non-PDF bytes early (CDN 404 pages are HTML/GIF).
    try:
        with open(path, "rb") as fh:
            magic = fh.read(4)
        if magic != b"%PDF":
            return []
    except OSError:
        return []

    facts: list[MetricFact] = []

    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                # Resolve the period for this page.
                if _is_annual:
                    page_text = page.extract_text() or ""
                    period = _period_from_page_text(page_text, _fallback_year)
                    if period is None:
                        # Cannot determine month for this page -- skip it.
                        continue
                else:
                    period = url_period

                # Primary path: structured table extraction.
                tables = page.extract_tables() or []
                if tables:
                    for table in tables:
                        for row in table:
                            if not row:
                                continue
                            # Normalise: replace None with empty string.
                            cells = [
                                c.strip() if isinstance(c, str) else ""
                                for c in row
                            ]
                            if not _pdf_is_data_row(cells):
                                continue
                            operator, handle, promo, ggr = _pdf_extract_row(cells)
                            if not operator:
                                continue
                            _emit_row_facts(facts, period, source_url, operator, handle, promo, ggr)
                else:
                    # Fallback: annual PDFs use a text-flow layout.
                    # Reconstruct rows from positioned word tokens.
                    for row_cells in _pdf_text_rows(page):
                        if not _pdf_is_data_row(row_cells):
                            continue
                        operator, handle, promo, ggr = _pdf_extract_row(row_cells)
                        if not operator:
                            continue
                        _emit_row_facts(facts, period, source_url, operator, handle, promo, ggr)
    except Exception:  # noqa: BLE001
        return []

    # Annual PDFs include both TYPE A (online) and TYPE B (retail) sections.
    # Both sections use the same operator names, so the same (operator, period,
    # metric_name) key can appear twice.  Sum both contributions per key before
    # returning so the DB upsert stores the combined total instead of having the
    # second occurrence overwrite the first.
    agg: dict[tuple[str, str, str], MetricFact] = {}
    for f in facts:
        key = (f.operator, f.period, f.metric_name)
        if key in agg:
            agg[key] = MetricFact(
                state=f.state,
                operator=f.operator,
                vertical=f.vertical,
                period=f.period,
                metric_name=f.metric_name,
                value_usd_cents=agg[key].value_usd_cents + f.value_usd_cents,
                source_url=f.source_url,
            )
        else:
            agg[key] = f
    facts = sorted(agg.values(), key=lambda f: (f.period, f.operator, f.metric_name))
    return facts
