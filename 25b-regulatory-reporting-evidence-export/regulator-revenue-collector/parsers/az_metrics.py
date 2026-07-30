# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Arizona Department of Gaming — Event Wagering monthly revenue parser.

AZ publishes one PDF per month (all publicly-available monthly report files
verified via Internet Archive CDX as of 2026-06-08 are PDFs).  The filename
format changed several times since the September 2021 launch:

Era 1 (2021-2022):  EW Revenue Report for Website - {Mon} {YYYY}.pdf
                    EW Revenue Report for Website July 2022.pdf   (Jul 2022)
                    EW Report for Website - Aug 2022.pdf          (Aug 2022)
                    EW April 2022 Revenue Report.pdf              (Apr 2022)
Era 2 (2023-2024):  EW Website Revenue Report-{Mon} {YYYY}.pdf
Era 3 (2024-2025):  EW Website Report - {Mon} {YYYY}.pdf
                    EW Revenue Report for ADG website - May 2025.pdf
Era 4 (2025+):      EW Website Report-{Month} {YYYY} UNAUDITED*.pdf
                    EW Revenue June 2025 - UNAUDITED-Revised *.pdf

The report shows per-operator rows with monthly event-wagering activity.

Observed layout (verified against public ADG report structure):

  Section header: "Event Wagering Activity"  (or similar)
  Columns (left-to-right):
    Operator / Licensee name
    Amount Wagered (handle)
    Patron Winnings / Amount Paid to Patrons
    Adjusted Gross Revenue / Net Wagering Revenue   ← ggr

Some editions include a "Tax" or "Privilege Fee" column.

Period is derived from the source URL filename; examples:
  "EW Website Report - Mar 2025.pdf"               → 2025-03
  "EW Website Report-Dec 2024.pdf"                 → 2024-12
  "EW Website Revenue Report-Apr 2024.pdf"         → 2024-04
  "EW Revenue Report for Website - Jan 2023.pdf"   → 2023-01
  "EW Website Report-January 2026 UNAUDITED.pdf"   → 2026-01

XLSX support: ADG also publishes per-operator submission templates as XLSX
(e.g. "EW Mobile Revenue Template-Sept 2021.xlsx"); this parser handles them
with openpyxl.  The layout is a vertical form:

  Row ~3:  title "Monthly* Report of Event Wagering (Mobile|Retail) Revenue"
  Col D ("Totals") or col E contains the numeric values in fixed row positions:
    Line 1  = Gross Receipts (handle)
    Line 6  = Adjusted Gross Event Wagering Receipts (ggr)
    Line 10 = Total Revenue for this Month
    Line 11 = Privilege Fee (tax_paid)

  Period is read from row 5 col E ("Reporting Month / Day / Year:") or from
  the filename.  Operator is read from row 5 col C ("Operator or Designee:").

Returns [] on parse failure — never raises.
"""
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pdfplumber

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MONTHS = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    # "may" already covered by full name
    "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "sept": "09", "oct": "10", "nov": "11", "dec": "12",
}

# Match all known AZ filename variants.  The key insight is that every
# filename ends with a month token and a 4-digit year, optionally followed
# by extra words (e.g. "UNAUDITED UPDATED CB 17 APR").
#
# The prefix is one of:
#   "EW Revenue Report for Website"         (era 1, 2021-2022)
#   "EW Revenue Report for ADG website"     (era 3, 2025)
#   "EW Website Revenue Report"             (era 2, 2023-2024)
#   "EW Website Report"                     (era 3+, 2024+)
#   "EW Report for Website"                 (Aug 2022 variant)
#   "EW Revenue June 2025 ..."              (month-first variant 2025)
#   "EW {FullMonth} {YYYY} Revenue Report"  (Apr 2022 non-standard)
#   "EW Mobile Revenue Template"            (XLSX submission form)
#   "EW Retail Revenue Template"            (XLSX submission form)
#
# Between prefix and month there are various combinations of spaces, dashes,
# and percent-encoded spaces (%20).
_URL_PERIOD_RE = re.compile(
    r"EW(?:[%20 _-]+(?:Revenue[%20 _-]+(?:Report|Rpt)[%20 _-]+for[%20 _-]+(?:ADG[%20 _-]+)?[Ww]ebsite"
    r"|Website[%20 _-]+Revenue[%20 _-]+Report"
    r"|Website[%20 _-]+Report"
    r"|Report[%20 _-]+for[%20 _-]+Website"
    r"|(?:Mobile|Retail)[%20 _-]+Revenue[%20 _-]+Template"
    r"))"
    r"[%20 _-]*-?[%20 _-]*"          # separator (dash optional, spaces around it)
    r"([A-Za-z]+)"                    # month name or abbreviation
    r"[%20 _-]+"                      # separator
    r"(\d{4})",                       # 4-digit year
    re.IGNORECASE,
)

# Secondary regex for the month-first variant:
# "EW Revenue Report June 2025 - UNAUDITED..." or "EW April 2022 Revenue Report"
# Skips any intervening keyword tokens (Revenue, Report, Rpt, Website, ADG) before
# the month name, so we find the first actual month token.
_URL_PERIOD_MONTH_FIRST_RE = re.compile(
    r"EW[%20 _-]+(?:(?:Revenue|Report|Rpt|Website|ADG)[%20 _-]+)*([A-Za-z]+)[%20 _-]+(\d{4})",
    re.IGNORECASE,
)

# Column header detection — we look for these substrings (case-insensitive)
# to locate which column index maps to which metric.
_HDR_HANDLE = re.compile(r"amount\s+wager|handle|total\s+wager", re.I)
_HDR_PATRON = re.compile(r"patron\s+win|amount\s+paid|payout", re.I)
_HDR_GGR = re.compile(r"adjusted\s+gross|net\s+wager|gross\s+revenue|agr\b|ngr\b", re.I)
_HDR_TAX = re.compile(r"privilege\s+fee|tax|fee", re.I)

# Row-level totals / footer sentinel — skip these rows.
_TOTAL_RE = re.compile(r"^\s*(?:total|grand\s+total|subtotal|statewide)\s*$", re.I)

# Number token: optional leading dollar sign / open-paren, digits, commas,
# optional decimal, optional close-paren for negatives.
_NUM_RE = re.compile(r"^\(?-?\$?[\d,]+(?:\.\d+)?\)?$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _period_from_url(url: str) -> str | None:
    """Extract YYYY-MM from the PDF/XLSX filename embedded in the URL.

    Tries the standard prefix-first regex first; falls back to the
    month-first variant used by some 2022/2025 filenames.
    """
    m = _URL_PERIOD_RE.search(url)
    if m:
        mon = _MONTHS.get(m.group(1).lower())
        if mon:
            return f"{m.group(2)}-{mon}"
    # Month-first variant: "EW Revenue June 2025 - UNAUDITED..." or
    # "EW April 2022 Revenue Report..."
    m2 = _URL_PERIOD_MONTH_FIRST_RE.search(url)
    if m2:
        mon = _MONTHS.get(m2.group(1).lower())
        if mon:
            return f"{m2.group(2)}-{mon}"
    return None


def _to_cents(raw: str | None) -> int | None:
    """Convert a formatted dollar string to integer cents."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s in ("-", "--", "N/A", "n/a"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = re.sub(r"[$, ]", "", s)
    if not s:
        return None
    try:
        cents = int((Decimal(s) * 100).to_integral_value())
        return -cents if neg else cents
    except (InvalidOperation, ValueError):
        return None


def _is_number(val: str | None) -> bool:
    """True if val looks like a formatted dollar/number token."""
    if not val:
        return False
    return bool(_NUM_RE.match(str(val).strip()))


def _find_col(headers: list[str | None], pattern: re.Pattern) -> int | None:
    """Return the first column index whose header text matches pattern."""
    for i, h in enumerate(headers):
        if h and pattern.search(str(h)):
            return i
    return None


def _sum_cents(*raw: str | None) -> int | None:
    """Sum one or more raw dollar tokens, treating unparsable/'-' as 0.

    Returns None only when every token is unparsable/blank, so a row with
    no data at all still produces no fact (rather than a spurious $0 fact).
    """
    total = 0
    any_val = False
    for r in raw:
        cents = _to_cents(r)
        if cents is not None:
            total += cents
            any_val = True
    return total if any_val else None


# ---------------------------------------------------------------------------
# Real-world "Retail | Mobile" layout parser
# ---------------------------------------------------------------------------
#
# Verified against live ADG PDFs spanning Sept 2021 - Jan 2026 (all naming
# eras): every monthly report uses one consistent per-operator row layout,
# regardless of filename convention:
#
#   Operator: Retail Mobile Retail Mobile Retail Mobile Retail Mobile Retail Mobile Retail 8% Mobile 10%
#   <Operator Name> $ v1 $ v2 $ v3 $ v4 $ v5 $ v6 $ v7 $ v8 $ v9 $ v10 $ v11 $ v12
#
# The 12 values are six Retail/Mobile pairs, in this fixed order:
#   1. Gross Event Wagering Receipts (Wagers)                        -> handle
#   2. Winnings Paid to Players (Payouts)                            -> patron_winnings
#   3. Adjusted Gross Event Wagering Receipts prior to Free Bets
#      Allowable Deduction                                            (not emitted)
#   4. Free Bets / Promotional Credits Deduction Allowed and Taken     (not emitted)
#   5. Adjusted Gross Event Wagering Receipts Subject to Privilege
#      Fees                                                          -> ggr
#   6. Privilege Fees                                                 -> tax_paid
#
# Group 5 (not group 3) is used for 'ggr': it is the actual taxable revenue
# base privilege fees are assessed against, its magnitude reconciles with
# the "Total ... Combined" summary row printed at the foot of every report,
# and it matches ADG's own published monthly totals (~$35-55M/month).
#
# pdfplumber's extract_tables() badly mangles this layout — it splits the
# Retail/Mobile sub-header across cells and merges data columns unevenly —
# so we parse the text layer directly instead of relying on x-coordinates.
# Every per-operator line has exactly 12 "$"-prefixed values, which makes
# splitting on "$" a reliable, layout-independent way to isolate them. Some
# values render with a stray internal space (a pdfplumber word-spacing
# artifact, e.g. "4 3,239,581.03"); `_to_cents` already strips whitespace
# so this is handled transparently.

_OPERATOR_ROW_VALUES = 12
_OPERATOR_HEADER_RE = re.compile(r"Operator:\s*Retail\s+Mobile", re.I)


def _parse_operator_lines(
    text: str, period: str, source_url: str,
) -> list[MetricFact]:
    """Parse the real per-operator 'Retail | Mobile' text layout.

    Returns [] when the expected column header isn't present in this page's
    text, so callers can fall back to other strategies for any layout that
    differs (e.g. a future ADG redesign).
    """
    if not _OPERATOR_HEADER_RE.search(text):
        return []

    facts: list[MetricFact] = []
    for line in text.splitlines():
        parts = line.split("$")
        if len(parts) != _OPERATOR_ROW_VALUES + 1:
            continue

        operator = parts[0].strip().rstrip(":").strip()
        if not operator or not re.search(r"[A-Za-z]", operator):
            continue
        if "total" in operator.lower() or "combined" in operator.lower():
            continue

        values = parts[1:]
        handle = _sum_cents(values[0], values[1])
        patron = _sum_cents(values[2], values[3])
        ggr = _sum_cents(values[8], values[9])
        tax = _sum_cents(values[10], values[11])

        base = dict(
            state="AZ",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            source_url=source_url,
        )
        if ggr is not None:
            facts.append(MetricFact(**base, metric_name="ggr", value_usd_cents=ggr))
        if handle is not None:
            facts.append(MetricFact(**base, metric_name="handle", value_usd_cents=handle))
        if patron is not None:
            facts.append(MetricFact(
                **base, metric_name="patron_winnings", value_usd_cents=patron,
            ))
        if tax is not None:
            facts.append(MetricFact(**base, metric_name="tax_paid", value_usd_cents=tax))

    return facts


# ---------------------------------------------------------------------------
# Table parser
# ---------------------------------------------------------------------------

def _parse_table(
    table: list[list[str | None]],
    period: str,
    source_url: str,
) -> list[MetricFact]:
    """Parse one pdfplumber table into MetricFacts."""
    if not table:
        return []

    # Identify the header row — first row that contains at least two
    # recognisable column-header substrings.
    hdr_idx: int | None = None
    col_operator: int | None = None
    col_handle: int | None = None
    col_patron: int | None = None
    col_ggr: int | None = None
    col_tax: int | None = None

    for row_i, row in enumerate(table[:5]):  # header is always in first 5 rows
        h_idx = _find_col(row, _HDR_HANDLE)
        g_idx = _find_col(row, _HDR_GGR)
        if h_idx is not None or g_idx is not None:
            hdr_idx = row_i
            col_handle = h_idx
            col_ggr = g_idx
            col_patron = _find_col(row, _HDR_PATRON)
            col_tax = _find_col(row, _HDR_TAX)
            # Operator column: the leftmost non-numeric, non-empty header cell
            # that wasn't already assigned.
            assigned = {col_handle, col_ggr, col_patron, col_tax} - {None}
            for i, h in enumerate(row):
                if i not in assigned and h and not _is_number(h):
                    col_operator = i
                    break
            break

    if hdr_idx is None or col_ggr is None:
        # No recognisable header — try positional fallback:
        # col 0 = operator, col 1 = handle, col 2 = patron_winnings, col 3 = ggr
        if table and len(table[0]) >= 4:
            col_operator, col_handle, col_patron, col_ggr = 0, 1, 2, 3
            hdr_idx = 0
        else:
            return []

    facts: list[MetricFact] = []
    for row in table[hdr_idx + 1:]:
        if not row:
            continue

        def cell(idx: int | None) -> str | None:
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        op_raw = cell(col_operator)
        if not op_raw or _TOTAL_RE.match(op_raw):
            continue
        # Skip rows that look like sub-headers or blank padding.
        if not _is_number(cell(col_ggr)) and not _is_number(cell(col_handle)):
            continue

        operator = re.sub(r"\s+", " ", op_raw).strip()

        base = dict(
            state="AZ",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            source_url=source_url,
        )

        ggr = _to_cents(cell(col_ggr))
        if ggr is not None:
            facts.append(MetricFact(**base, metric_name="ggr", value_usd_cents=ggr))

        handle = _to_cents(cell(col_handle))
        if handle is not None:
            facts.append(MetricFact(**base, metric_name="handle", value_usd_cents=handle))

        patron = _to_cents(cell(col_patron))
        if patron is not None:
            facts.append(MetricFact(
                **base, metric_name="patron_winnings", value_usd_cents=patron,
            ))

        tax = _to_cents(cell(col_tax))
        if tax is not None:
            facts.append(MetricFact(**base, metric_name="tax_paid", value_usd_cents=tax))

    return facts


# ---------------------------------------------------------------------------
# Text-layer fallback (for PDFs where pdfplumber finds no tables)
# ---------------------------------------------------------------------------

# Matches a line like:
#   DraftKings   $12,345,678   $11,000,000   $1,345,678
# where we have at least one money token after the operator name.
_TEXT_ROW_RE = re.compile(
    r"^(.+?)\s{2,}"            # operator (2+ spaces before first number)
    r"(\(?-?\$?[\d,]+(?:\.\d+)?\)?)(?:\s+"  # first money value
    r"(\(?-?\$?[\d,]+(?:\.\d+)?\)?))?(?:\s+"  # optional second money value
    r"(\(?-?\$?[\d,]+(?:\.\d+)?\)?))?",  # optional third money value
    re.I,
)

# Keywords in page text that confirm this is an event wagering revenue table.
_SECTION_RE = re.compile(
    r"event\s+wager|sports\s+wager|adjusted\s+gross|amount\s+wager",
    re.I,
)


def _parse_text_fallback(
    text: str,
    period: str,
    source_url: str,
) -> list[MetricFact]:
    """Extract facts from raw text when pdfplumber finds no tables.

    This handles PDFs that use vector-drawn table lines (no rule objects),
    which pdfplumber cannot detect as formal tables.  We rely on consistent
    multi-space column alignment.
    """
    if not _SECTION_RE.search(text):
        return []

    # Determine column order from header line.
    col_order: list[str] = []  # e.g. ["handle", "patron_winnings", "ggr"]
    for line in text.splitlines():
        stripped = line.strip()
        if _HDR_HANDLE.search(stripped) or _HDR_GGR.search(stripped):
            # Build ordered list of metrics from left to right in this header line.
            tokens = re.split(r"\s{2,}", stripped)
            for tok in tokens:
                if _HDR_HANDLE.search(tok):
                    col_order.append("handle")
                elif _HDR_PATRON.search(tok):
                    col_order.append("patron_winnings")
                elif _HDR_GGR.search(tok):
                    col_order.append("ggr")
                elif _HDR_TAX.search(tok):
                    col_order.append("tax_paid")
            break

    # Default column order when header not found: handle, patron_winnings, ggr
    if not col_order:
        col_order = ["handle", "patron_winnings", "ggr"]

    facts: list[MetricFact] = []
    for line in text.splitlines():
        m = _TEXT_ROW_RE.match(line.strip())
        if not m:
            continue
        op_raw = m.group(1).strip()
        if _TOTAL_RE.match(op_raw) or not op_raw:
            continue
        values = [m.group(2), m.group(3), m.group(4)]

        base = dict(
            state="AZ",
            operator=re.sub(r"\s+", " ", op_raw),
            vertical="sports-wagering",
            period=period,
            source_url=source_url,
        )
        for metric_name, raw_val in zip(col_order, values):
            cents = _to_cents(raw_val)
            if cents is not None:
                facts.append(MetricFact(**base, metric_name=metric_name, value_usd_cents=cents))

    return facts


# ---------------------------------------------------------------------------
# XLSX parser (operator submission templates)
# ---------------------------------------------------------------------------

# Line numbers in the ADG XLSX template that carry financial data.
# These are the row labels (col B) in both the Mobile and Retail templates.
_XLSX_LINE_METRICS: dict[str | int, str] = {
    1:  "handle",       # Gross Receipts (Write/Handle)
    6:  "ggr",          # Adjusted Gross Event Wagering Receipts
    10: "ggr",          # Total Revenue for this Month (same as line 6 + adj)
    11: "tax_paid",     # Privilege Fee (10% of line 10)
}

# "Mobile" vs "Retail" in the sheet title helps identify the sub-vertical;
# both are mapped to 'sports-wagering' since AZ doesn't separate retail/mobile
# in the public aggregate reports.
_DATE_RE = re.compile(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})")


def _period_from_xlsx_date(raw) -> str | None:
    """Parse a cell value like '9/30/2021' or a datetime object into YYYY-MM."""
    if raw is None:
        return None
    # openpyxl may return a datetime object when data_only=True.
    if hasattr(raw, "year") and hasattr(raw, "month"):
        return f"{raw.year:04d}-{raw.month:02d}"
    s = str(raw).strip()
    m = _DATE_RE.search(s)
    if not m:
        return None
    month_val = int(m.group(1))
    year_raw = int(m.group(3))
    if year_raw < 100:
        year_raw += 2000
    if not (1 <= month_val <= 12 and 2020 <= year_raw <= 2035):
        return None
    return f"{year_raw:04d}-{month_val:02d}"


def _parse_az_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse an AZ ADG per-operator XLSX submission template.

    The template has a fixed vertical layout with named line numbers in
    column B and a "Totals" value in column E (index 4) or the nearest
    non-None cell to the right.  Period and operator are read from the
    header rows; both fall back to the filename/URL if missing.

    Returns [] on any parse failure.
    """
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError:
        return []

    period = _period_from_url(source_url) or _period_from_url(str(path))
    operator: str | None = None

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        return []

    # Pass 1: scan header rows for period and operator name.
    for row in rows[:12]:
        if row is None:
            continue
        # Row 5 (index 4): col C = "Operator or Designee:", col D = operator value
        #                   col E = "Reporting Month / Day / Year:", col F = date
        # Layout is 0-indexed: (None, None, label_c, value_d, label_e, value_f, ...)
        cells = list(row)
        if len(cells) >= 6:
            label_c = str(cells[2]).strip() if cells[2] else ""
            if "operator" in label_c.lower() and cells[3]:
                operator = str(cells[3]).strip() or None
            label_e = str(cells[4]).strip() if cells[4] else ""
            if "reporting month" in label_e.lower() and period is None:
                period = _period_from_xlsx_date(cells[5])

    if period is None:
        return []

    if not operator:
        operator = "AZ Statewide"

    # Pass 2: scan data rows for line numbers and associated values.
    # Col B (index 1) is the line number (int or str like "1").
    # Col E (index 4) is the "Totals" column; some cells may be None if the
    # operator left them blank, so skip None values.
    facts: list[MetricFact] = []
    # Track which line numbers we've already emitted to avoid double-counting
    # line 6 (ggr) when line 10 is also labelled "ggr".
    emitted_metrics: set[str] = set()

    for row in rows:
        if not row:
            continue
        cells = list(row)
        if len(cells) < 5:
            continue
        line_raw = cells[1]
        if line_raw is None:
            continue
        # Normalise: int or string digit.
        try:
            line_num = int(line_raw)
        except (TypeError, ValueError):
            continue

        metric = _XLSX_LINE_METRICS.get(line_num)
        if metric is None:
            continue

        # Skip line 10 for 'ggr' if we already emitted it from line 6.
        if metric == "ggr" and "ggr" in emitted_metrics:
            continue

        # Value is in col E (index 4); fall back to col H (index 7) which
        # holds formula-computed totals in some template versions.
        value_raw = cells[4]
        if value_raw is None and len(cells) > 7:
            value_raw = cells[7]
        if value_raw is None:
            continue

        cents = _to_cents(str(value_raw))
        if cents is None:
            continue

        facts.append(MetricFact(
            state="AZ",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name=metric,
            value_usd_cents=cents,
            source_url=source_url,
        ))
        emitted_metrics.add(metric)

    return facts


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_az_pdf(path: Path, source_url: str) -> list[MetricFact]:
    """Parse an AZ ADG Event Wagering monthly report into MetricFacts.

    Dispatches to the PDF parser or the XLSX parser depending on the file
    magic bytes, so callers don't need to know the format.  The backfill
    harness always calls this function regardless of the ReportFile.format
    value — the actual bytes determine the path taken.

    Returns [] on any parse failure — never raises.
    """
    # Detect format from magic bytes to be robust against mis-named files.
    try:
        with open(path, "rb") as fh:
            magic = fh.read(4)
    except OSError:
        return []

    if magic == b"%PDF":
        return _parse_az_pdf_bytes(path, source_url)

    # XLSX / ZIP-based Office format (PK\x03\x04 magic).
    if magic[:2] == b"PK":
        return _parse_az_xlsx(path, source_url)

    # Unknown format — return gracefully.
    return []


def _parse_az_pdf_bytes(path: Path, source_url: str) -> list[MetricFact]:
    """Internal: parse a verified PDF file."""
    period = _period_from_url(source_url) or _period_from_url(str(path))
    if period is None:
        return []

    facts: list[MetricFact] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""

                # Strategy 1 (primary): the real per-operator "Retail |
                # Mobile" row layout used by every ADG report observed
                # 2021-2026. Parses the text layer directly since
                # extract_tables() mangles this layout's merged sub-headers.
                page_facts = _parse_operator_lines(text, period, source_url)
                if page_facts:
                    facts.extend(page_facts)
                    continue

                # Strategy 2: pdfplumber table extraction (older/other
                # layouts this parser hasn't seen live).
                tables = page.extract_tables() or []
                for tbl in tables:
                    facts.extend(_parse_table(tbl, period, source_url))

                # Strategy 3: generic multi-space text-column fallback.
                if not tables and text:
                    facts.extend(
                        _parse_text_fallback(text, period, source_url)
                    )
    except Exception:  # noqa: BLE001 — be forgiving to malformed PDFs
        return []

    # Deduplicate: keep the last fact for each (operator, metric_name) key
    # in case the same row appears on multiple pages.
    seen: dict[tuple[str, str, str, str, str], MetricFact] = {}
    for f in facts:
        key = (f.state, f.operator, f.vertical, f.period, f.metric_name)
        seen[key] = f

    result = list(seen.values())
    result.sort(key=lambda f: (f.operator, f.metric_name))
    return result
