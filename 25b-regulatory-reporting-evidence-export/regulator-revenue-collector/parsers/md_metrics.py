# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Maryland Lottery & Gaming monthly XLSX parser.

Handles two file shapes:
  Sports:  {Month}-{YYYY}-Sports-Wagering-Data.xlsx
           Rows: one licensee per two-row block (Month row + FYTD row).
           Column layout (0-indexed, values_only):
             0  Licensee name  (str, Month row only; None on FYTD row)
             1  Period date or "FYTD" sentinel
             2  Handle
             3  Prizes Paid
             5  Promotional Play
             8  Contributions to the State (= tax_paid)

           Some months (e.g. 2024-06, 2025-03, 2025-08) are saved with
           the "Bets By Sport" (by-wager-type) tab as the active sheet.
           The parser explicitly selects the per-licensee "SW Data" tab
           instead of relying on wb.active.

  Casino:  {Month}-{YYYY}-Casino-Revenues.xlsx
           Two-row blocks per property (Month row + FYTD row).
           Column layout mirrors sports but for casino properties.
           Property name detected same way: col 0 non-None on the Month row.

Emits MetricFact with state="MD", operator=licensee/property name,
vertical=passed-in arg, period=YYYY-MM from the URL, metric_name in
{handle, ggr, promotional_play, tax_paid}, value_usd_cents as int.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONTHS = {
    "January": "01", "February": "02", "March": "03", "April": "04",
    "May": "05", "June": "06", "July": "07", "August": "08",
    "September": "09", "October": "10", "November": "11", "December": "12",
}

# Sentinel strings that mark FYTD / header / total rows to skip.
_SKIP_NAMES = re.compile(
    r"total|statewide|combined|licensee|fytd|rounding|maryland lottery"
    r"|mobile and retail",
    re.I,
)


def _period_from_url(url: str) -> str | None:
    m = re.search(r"/([A-Za-z]+)-(\d{4})-(?:Sports|Casino)", url)
    if not m:
        return None
    return f"{m.group(2)}-{_MONTHS.get(m.group(1), '01')}"


def _to_cents(raw) -> int | None:
    """Convert a raw cell value (float, int, str) to integer USD cents."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in ("-", "N/A", "n/a"):
        return 0
    # Remove currency symbols, spaces, commas; keep digits, dot, minus.
    s = re.sub(r"[^\d.\-]", "", s.replace(",", ""))
    if not s:
        return 0
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def _is_data_operator(name: str) -> bool:
    """Return True when the cell looks like a real operator/property name."""
    if not name or not isinstance(name, str):
        return False
    return not _SKIP_NAMES.search(name.strip())


# ---------------------------------------------------------------------------
# Sports-wagering XLSX  (operators as row labels)
# ---------------------------------------------------------------------------

# Column indices in the "Month" data row:
_SW_HANDLE = 2
_SW_PRIZES = 3
_SW_PROMO  = 5
_SW_TAX    = 8   # "Contributions to the State"


def _parse_sports(ws, period: str, source_url: str) -> list[MetricFact]:
    out: list[MetricFact] = []
    operator: str | None = None

    for row in ws.iter_rows(values_only=True):
        col0 = row[0] if row else None

        # Identify the Month data row: col 0 is a non-empty string operator name,
        # col 1 is a datetime (or date-like) — not the string "FYTD".
        if isinstance(col0, str) and col0.strip() and not isinstance(row[1], str):
            candidate = col0.strip()
            if _is_data_operator(candidate):
                operator = candidate
                # Emit metrics for this Month row.
                def _emit(metric: str, idx: int) -> None:
                    if operator is None or idx >= len(row):
                        return
                    cents = _to_cents(row[idx])
                    if cents is None:
                        return
                    out.append(MetricFact(
                        state="MD",
                        operator=operator,
                        vertical="sports-wagering",
                        period=period,
                        metric_name=metric,
                        value_usd_cents=cents,
                        source_url=source_url,
                    ))

                # GGR = Handle - Prizes - PromoPlay  (matches "Taxable Win" col 7
                # but we prefer to derive from canonical constituents).
                handle_cents = _to_cents(row[_SW_HANDLE]) if _SW_HANDLE < len(row) else None
                prizes_cents = _to_cents(row[_SW_PRIZES]) if _SW_PRIZES < len(row) else None
                promo_cents  = _to_cents(row[_SW_PROMO])  if _SW_PROMO  < len(row) else None

                _emit("handle",           _SW_HANDLE)
                _emit("promotional_play", _SW_PROMO)
                _emit("tax_paid",         _SW_TAX)

                if (
                    handle_cents is not None
                    and prizes_cents is not None
                    and promo_cents is not None
                ):
                    ggr = max(0, handle_cents - prizes_cents - promo_cents)
                    out.append(MetricFact(
                        state="MD",
                        operator=operator,
                        vertical="sports-wagering",
                        period=period,
                        metric_name="ggr",
                        value_usd_cents=ggr,
                        source_url=source_url,
                    ))
        # FYTD row — reset operator so we don't accidentally reuse it.
        elif isinstance(col0, str) and col0.strip() == "" or col0 is None:
            if isinstance(row[1] if len(row) > 1 else None, str):
                operator = None

    return out


# ---------------------------------------------------------------------------
# Casino XLSX  (properties as row labels, same two-row-per-entity layout)
# ---------------------------------------------------------------------------

# Casino revenue worksheets have a similar two-row-per-property structure.
# Column positions may vary slightly by year; we use the same indices as sports
# where applicable and fall back to heuristics for GGR.
_CA_HANDLE = 2   # "Total Win" or "Handle" — first numeric column
_CA_TAX    = 3   # Education Trust Fund / state tax column


def _parse_casino(ws, period: str, source_url: str) -> list[MetricFact]:
    out: list[MetricFact] = []
    # Detect header row to infer column semantics dynamically.
    col_map: dict[str, int] = {}  # metric_name -> col index

    for row in ws.iter_rows(values_only=True):
        col0 = row[0] if row else None

        # Header detection: row whose col0 contains "Casino" or "Property".
        if isinstance(col0, str) and re.search(r"casino|property|licensee", col0, re.I):
            for idx, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                c = cell.lower()
                if "win" in c or "revenue" in c or "handle" in c:
                    col_map.setdefault("ggr", idx)
                if "education" in c or "trust" in c or "state" in c and "tax" in c:
                    col_map.setdefault("tax_paid", idx)
                if "promo" in c:
                    col_map.setdefault("promotional_play", idx)
            continue

        # Data row: col 0 is a non-empty property name, col 1 is a date.
        if isinstance(col0, str) and col0.strip() and not isinstance(row[1] if len(row) > 1 else None, str):
            candidate = col0.strip()
            if not _is_data_operator(candidate):
                continue

            def _emit_ca(metric: str, idx: int) -> None:
                if idx >= len(row):
                    return
                cents = _to_cents(row[idx])
                if cents is None:
                    return
                out.append(MetricFact(
                    state="MD",
                    operator=candidate,
                    vertical="commercial-casino",
                    period=period,
                    metric_name=metric,
                    value_usd_cents=cents,
                    source_url=source_url,
                ))

            ggr_col   = col_map.get("ggr",              _CA_HANDLE)
            tax_col   = col_map.get("tax_paid",          _CA_TAX)
            promo_col = col_map.get("promotional_play",  -1)

            _emit_ca("ggr",      ggr_col)
            _emit_ca("tax_paid", tax_col)
            if promo_col >= 0:
                _emit_ca("promotional_play", promo_col)

    return out


# ---------------------------------------------------------------------------
# Sheet-selection helpers
# ---------------------------------------------------------------------------

# Pattern that matches the per-licensee sports-wagering sheet names.
# Examples: "June 2024 SW Data", "July 2024 SW Data", "Sept. 2024 SW Data",
#           "February 2205 SW Data" (typo in real file — still matches).
_SW_DATA_SHEET = re.compile(r"SW\s+Data", re.I)

# Pattern that identifies the by-wager-type breakdown sheet to skip.
_BETS_BY_SPORT_SHEET = re.compile(r"Bets\s+By\s+Sport|Bets\s+by\s+Sport", re.I)

# Pattern for casino per-property sheets.
_CASINO_DATA_SHEET = re.compile(r"Casino\s+Revenue|Revenue\s+Data", re.I)


def _select_sports_sheet(wb):
    """Return the per-licensee sports-wagering worksheet.

    MD occasionally saves the XLSX with the 'Bets By Sport' (by-wager-type)
    tab as the active sheet (confirmed in 2024-06, 2025-03, 2025-08).  Always
    prefer a sheet whose name matches the SW Data pattern; fall back to the
    active sheet only when no such sheet is found.
    """
    # First pass: find by name pattern.
    for name in wb.sheetnames:
        if _SW_DATA_SHEET.search(name) and not _BETS_BY_SPORT_SHEET.search(name):
            return wb[name]
    # Second pass: any sheet that is NOT the bets-by-sport breakdown.
    non_bets_sheets = [n for n in wb.sheetnames if not _BETS_BY_SPORT_SHEET.search(n)]
    if non_bets_sheets:
        return wb[non_bets_sheets[0]]
    # Last resort — return active (should not happen in practice).
    return wb.active


def _select_casino_sheet(wb):
    """Return the per-property casino revenue worksheet."""
    for name in wb.sheetnames:
        if _CASINO_DATA_SHEET.search(name):
            return wb[name]
    return wb.active


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_md_xlsx(path: Path, vertical: str, source_url: str) -> list[MetricFact]:
    """Parse a Maryland Lottery & Gaming monthly XLSX and return MetricFact rows.

    Args:
        path:       Local path to the downloaded .xlsx file.
        vertical:   "sports-wagering" or "commercial-casino".
        source_url: Original download URL; used to derive the period and stored
                    on every returned fact.

    Returns:
        List of MetricFact objects, one per (operator, metric_name) combination
        for the current-month data rows (FYTD rows are skipped).
    """
    period = _period_from_url(source_url)
    if period is None:
        # Fall back: use a placeholder so callers can still see the data.
        period = "0000-01"

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        if vertical == "sports-wagering":
            ws = _select_sports_sheet(wb)
            return _parse_sports(ws, period, source_url)
        else:
            ws = _select_casino_sheet(wb)
            return _parse_casino(ws, period, source_url)
    finally:
        wb.close()
