# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Denmark Spillemyndigheden monthly market XLSX parser.

Source URL pattern:
    https://www.spillemyndigheden.dk/uploads/{YYYY}-{MM}/Data%20for%20the%20monthly%20statistics.xlsx

The workbook has one sheet per vertical.  Each sheet has three columns:
    col 0 – Month  (English full name, e.g. "January"; may have trailing whitespace)
    col 1 – Year   (integer, e.g. 2024)
    col 2 – GGR (DKK)  (raw DKK, not millions; float/int)

Sheet → vertical mapping (current as of 2026-04; defensive on name variants):
    "Online casino"      → "igaming"
    "Betting"            → "sports-wagering"
    "Land-based casino"  → "land-based-casino"
    "Gaming machines"    → "slot-arcades"   (labelled "Slot arcades" in older exports)
    "Bingo"              → "bingo"

FX: DKK → USD ≈ 0.143
value_usd_cents = int(dkk_value * 0.143 * 100)
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DKK_TO_USD: float = 0.143

# Sheet-name fragments (lower-case) → canonical vertical
_SHEET_VERTICAL: dict[str, str] = {
    "online casino":     "igaming",
    "betting":           "sports-wagering",
    "land-based casino": "land-based-casino",
    "land-based":        "land-based-casino",
    "gaming machines":   "slot-arcades",
    "slot arcades":      "slot-arcades",
    "bingo":             "bingo",
}

# Full English month names → zero-padded month number
_MONTH_NAME: dict[str, str] = {
    "january":   "01",
    "february":  "02",
    "march":     "03",
    "april":     "04",
    "may":       "05",
    "june":      "06",
    "july":      "07",
    "august":    "08",
    "september": "09",
    "october":   "10",
    "november":  "11",
    "december":  "12",
}

# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------

def _period_from_label(s) -> Optional[str]:
    """Convert a variety of month-label formats to 'YYYY-MM'.

    Handles:
      - Full English name + int year  (primary format: "January", 2024)
      - "January 2024", "Jan-24", "2024-01"
      - Danish/partial: "Yan 2024" (defensive regex fallback)
    """
    if s is None:
        return None
    months = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "maj": "05", "may": "05", "jun": "06",
        "jul": "07", "aug": "08", "sep": "09",
        "okt": "10", "oct": "10", "nov": "11", "dec": "12",
    }
    text = str(s).strip().lower()
    # Try full month name first (e.g. "january" → "01")
    if text in _MONTH_NAME:
        return None  # year not embedded; caller must supply separately

    # ISO format: "2024-01"
    iso = re.fullmatch(r"(\d{4})-(\d{2})", text)
    if iso:
        return f"{iso.group(1)}-{iso.group(2)}"

    # Regex: grab 3-letter month prefix + 2-or-4 digit year
    m = re.search(r"([a-zæøå]{3})\D*(\d{2,4})", text)
    if not m:
        return None
    mon = m.group(1)[:3]
    yr = m.group(2)
    if len(yr) == 2:
        yr = "20" + yr
    return f"{yr}-{months[mon]}" if mon in months else None


def _period_from_month_year(month_cell, year_cell) -> Optional[str]:
    """Build 'YYYY-MM' from the two-column (month-name, year) layout used by
    Spillemyndigheden.  Falls back to _period_from_label for single-cell
    labels.
    """
    if year_cell is not None and month_cell is not None:
        month_str = str(month_cell).strip().lower()
        if month_str in _MONTH_NAME:
            try:
                yr = int(year_cell)
                return f"{yr:04d}-{_MONTH_NAME[month_str]}"
            except (ValueError, TypeError):
                pass

    # Single-cell fallback (e.g. "January 2024")
    if month_cell is not None:
        label = f"{month_cell} {year_cell}" if year_cell is not None else str(month_cell)
        return _period_from_label(label)

    return None


# ---------------------------------------------------------------------------
# Sheet vertical detection
# ---------------------------------------------------------------------------

def _detect_vertical(sheet_name: str) -> Optional[str]:
    """Map a sheet name to a canonical vertical string, or None if unknown."""
    key = sheet_name.strip().lower()
    # Exact match first
    if key in _SHEET_VERTICAL:
        return _SHEET_VERTICAL[key]
    # Substring match (handles minor naming variants)
    for fragment, vertical in _SHEET_VERTICAL.items():
        if fragment in key:
            return vertical
    return None


# ---------------------------------------------------------------------------
# Value conversion
# ---------------------------------------------------------------------------

def _dkk_to_usd_cents(dkk_value) -> Optional[int]:
    """Convert a raw DKK value to USD cents using the fixed FX rate.

    The Spillemyndigheden XLSX stores GGR in raw DKK (not millions).
    A typical monthly online-casino GGR is ~350,000,000 DKK (≈ 350M DKK).
    """
    if dkk_value is None:
        return None
    try:
        f = float(dkk_value)
    except (TypeError, ValueError):
        return None
    if f == 0.0:
        return 0
    return int(f * _DKK_TO_USD * 100)


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_dk_xlsx(path: Path, source_url: str) -> list[MetricFact]:
    """Parse a Spillemyndigheden monthly market XLSX and return MetricFact rows.

    One MetricFact per (vertical, period) with:
        state    = "DK"
        operator = "DK Statewide"
        vertical = mapped from sheet name
        period   = "YYYY-MM"
        metric_name = "ggr"
        value_usd_cents = int(dkk_ggr * 0.143 * 100)

    Sheets whose name cannot be mapped to a known vertical are skipped with a
    warning rather than raising, so new sheets added by the regulator do not
    crash existing runs.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    results: list[MetricFact] = []

    for sheet_name in wb.sheetnames:
        vertical = _detect_vertical(sheet_name)
        if vertical is None:
            # Unknown sheet — skip silently (e.g. metadata tabs)
            continue

        ws = wb[sheet_name]
        header_skipped = False

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            month_cell = row[0] if len(row) > 0 else None
            year_cell  = row[1] if len(row) > 1 else None
            ggr_cell   = row[2] if len(row) > 2 else None

            # Skip the header row (first non-empty row where col-0 is "Month")
            if not header_skipped:
                if isinstance(month_cell, str) and month_cell.strip().lower() == "month":
                    header_skipped = True
                    continue
                # If somehow the first row has no "Month" header, treat it as data
                header_skipped = True

            period = _period_from_month_year(month_cell, year_cell)
            if period is None:
                # Could not parse period — skip row
                continue

            cents = _dkk_to_usd_cents(ggr_cell)
            if cents is None:
                continue

            results.append(MetricFact(
                state="DK",
                operator="DK Statewide",
                vertical=vertical,
                period=period,
                metric_name="ggr",
                value_usd_cents=cents,
                source_url=source_url,
            ))

    wb.close()
    return results
