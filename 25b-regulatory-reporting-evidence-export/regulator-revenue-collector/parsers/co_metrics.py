# Companion code for "The Backend of Luck" - Chapter 25b, Regulatory Reporting and Evidence Export.
# https://thebackendofluck.com | https://github.com/thebackendofluck/book
# SPDX-License-Identifier: Apache-2.0
#
# FOR TESTING AND EVALUATION ONLY. NOT FOR PRODUCTION USE.
# Published to demonstrate the patterns explained in the book. This code is
# not certified for real-money gaming: operating a gambling platform requires
# your own licence, independent test-lab certification (GLI, eCOGRA or
# equivalent) and regulator approval.

"""Colorado Division of Gaming sports betting rolling XLS parser.

Source: https://sbg.colorado.gov/media/11906
File:   "Calendar Yr Monthly Summary WebPage (1).xls"

Layout (one sheet per calendar year, e.g. "2020", "2021", ...):

  Row A — month-label header: col 0 = operator name or blank,
           subsequent non-blank string cells are month abbreviations
           such as "May", "Jun", "Jul", or "May-20", "Jan 2021".
           Four data columns follow each month label (Handle, GGR, Adj GGR, Tax).

  Row B … — each operator occupies a block:
           • First row in block: col 0 = operator name (string, non-numeric),
             col 1..N = numeric values for Handle in month order.
           • Subsequent rows in block carry the remaining metrics
             (GGR / Adjusted GGR / Tax) in the same column order.

  OR — the more common layout seen in CO Excel publications:
           • col 0 = operator name (string)
           • col 1 = metric label ("Handle" / "Gross Gaming Revenue" /
                     "Adjusted Gross Revenue" / "Tax")
           • col 2..N = monthly values in the same order as the header row.

Both layouts are handled via the same column-map derived from the header row.

File format is legacy .xls (BIFF8); openpyxl cannot read it.  We use xlrd
as primary and fall back to openpyxl (after a .xlsx rename) only if xlrd
is unavailable (future-proofing in case the regulator ever switches format).
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any

from .metrics_model import MetricFact

# ---------------------------------------------------------------------------
# Period helper (provided in spec; reproduced verbatim with minor guard)
# ---------------------------------------------------------------------------
_MON_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _period_from_label(s: Any) -> str | None:
    if not s:
        return None
    m = re.search(r"([A-Za-z]{3,9})\s*[-/]?\s*(\d{2,4})", str(s))
    if not m:
        return None
    mon = m.group(1)[:3].lower()
    yr = m.group(2)
    if len(yr) == 2:
        yr = "20" + yr
    return f"{yr}-{_MON_MAP[mon]}" if mon in _MON_MAP else None


def _period_from_abbr(abbr: str, sheet_year: str) -> str | None:
    """Turn a bare month abbreviation like 'May' into '2020-05' using the
    sheet's calendar year as context."""
    key = abbr.strip()[:3].lower()
    if key not in _MON_MAP:
        return None
    return f"{sheet_year}-{_MON_MAP[key]}"


# ---------------------------------------------------------------------------
# Metric-label normalisation
# ---------------------------------------------------------------------------
_METRIC_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bhandle\b", re.I),                           "handle"),
    (re.compile(r"adjusted\s+gross", re.I),                     "ggr"),   # AGR ≈ net GGR after promos
    (re.compile(r"gross\s+gaming\s+revenue", re.I),             "ggr"),
    (re.compile(r"\bggr\b", re.I),                              "ggr"),
    (re.compile(r"\brevenue\b", re.I),                          "ggr"),
    (re.compile(r"\btax\b", re.I),                              "tax_paid"),
]

# Rows whose first-column label looks like these are skipped (totals / blanks)
_SKIP_RE = re.compile(
    r"^\s*(total|subtotal|grand\s+total|all\s+operator|industry|statewide"
    r"|average|ytd|fytd|footnote|\*|#)\b",
    re.I,
)


def _classify_metric(label: str) -> str | None:
    for pat, name in _METRIC_PATTERNS:
        if pat.search(label):
            return name
    return None


# ---------------------------------------------------------------------------
# Numeric conversion
# ---------------------------------------------------------------------------
_STRIP_RE = re.compile(r"[\$,\s]")


def _to_cents(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        val = float(raw)
        return int(round(val * 100))
    s = _STRIP_RE.sub("", str(raw))
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    if not s or s in ("-", "--", "N/A", "n/a"):
        return None
    try:
        return int(round(float(s) * 100)) * (-1 if neg else 1)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# XLS reader (xlrd primary, openpyxl fallback)
# ---------------------------------------------------------------------------

def _iter_sheets_xlrd(path: Path):
    """Yield (sheet_name, rows) where rows is list[list[Any]] via xlrd."""
    import xlrd  # type: ignore[import-untyped]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = xlrd.open_workbook(str(path))
    for sheet in wb.sheets():
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                # xlrd cell types: 0=empty,1=text,2=number,3=date,4=bool,5=error
                if cell.ctype == 1:
                    row.append(str(cell.value).strip())
                elif cell.ctype == 2:
                    # integer if whole number
                    v = cell.value
                    row.append(int(v) if v == int(v) else v)
                elif cell.ctype == 3:
                    # date serial — keep as float; unlikely in this report
                    row.append(cell.value)
                elif cell.ctype in (0, 5):
                    row.append(None)
                else:
                    row.append(cell.value)
            rows.append(row)
        yield sheet.name, rows


def _iter_sheets_openpyxl(path: Path):
    """Fallback: openpyxl read_only (works only for .xlsx)."""
    from openpyxl import load_workbook  # type: ignore[import-untyped]
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            yield name, rows
    finally:
        wb.close()


def _iter_sheets(path: Path):
    """Try xlrd first (required for .xls); fall back to openpyxl for .xlsx."""
    try:
        import xlrd  # noqa: F401
        yield from _iter_sheets_xlrd(path)
    except Exception:
        yield from _iter_sheets_openpyxl(path)


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def _year_from_sheet_name(name: str) -> str | None:
    """Extract 4-digit year from sheet name like '2020', 'CY 2021', etc."""
    m = re.search(r"(20\d{2})", name)
    return m.group(1) if m else None


def _is_operator_name(val: Any) -> bool:
    """Heuristic: col-0 value is an operator name if it's a non-empty string
    that doesn't look like a metric label or a month abbreviation."""
    if not isinstance(val, str):
        return False
    s = val.strip()
    if not s:
        return False
    if _SKIP_RE.match(s):
        return False
    if _classify_metric(s):
        return False
    # Month abbreviation alone → not an operator
    if s[:3].lower() in _MON_MAP and len(s) <= 5:
        return False
    return True


def _parse_sheet(
    sheet_name: str,
    rows: list[list[Any]],
    source_url: str,
) -> list[MetricFact]:
    """Parse one calendar-year sheet.

    Layout detection
    ----------------
    CO uses two closely related layouts across the years:

    Layout A (most years) — metric labels in col 1:
        col 0: operator name  |  col 1: "Handle" / "Gross Gaming Revenue" / ...
        col 2..N: monthly $ values aligned to month columns in the header row

    Layout B (occasional) — metric rows stacked under operator name row:
        col 0: operator name  |  col 1..N: Handle values
        col 0: (blank)        |  col 1: "Gross Gaming Revenue"  | col 2..N: values
        ...

    The parser handles both by:
    1. Scanning for a header row that has month abbreviations in col 1+.
    2. Walking subsequent rows: if col 0 is an operator name, update context.
       If col 1 is a metric label (or col 0 is blank and col 1 looks like
       a metric), emit facts for each period column.
    """
    sheet_year = _year_from_sheet_name(sheet_name)
    out: list[MetricFact] = []

    # Step 1: find the header row (contains month abbreviations)
    period_col_map: dict[int, str] = {}  # col_index -> "YYYY-MM"
    header_row_idx = -1

    for ri, row in enumerate(rows):
        found: dict[int, str] = {}
        for ci, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            s = cell.strip()
            # Try full "Jan 2021" style first
            p = _period_from_label(s)
            if p:
                found[ci] = p
                continue
            # Bare abbreviation requires sheet_year context
            if sheet_year and s[:3].lower() in _MON_MAP and len(s) <= 4:
                p2 = _period_from_abbr(s, sheet_year)
                if p2:
                    found[ci] = p2
        if len(found) >= 3:  # at least 3 month columns to be a real header
            period_col_map = found
            header_row_idx = ri
            break

    if not period_col_map:
        # Last-ditch: if sheet_year known, try to infer 12 data columns from
        # the first fully-numeric wide row and assign Jan..Dec of sheet_year.
        if sheet_year:
            for ri, row in enumerate(rows):
                numeric_cols = [
                    ci for ci, v in enumerate(row)
                    if isinstance(v, (int, float)) and not isinstance(v, bool)
                ]
                if len(numeric_cols) >= 6:
                    # guess that col 0 = operator, cols 1..12 = months Jan-Dec
                    for offset, mon_key in enumerate(sorted(_MON_MAP.keys())):
                        ci = 1 + offset
                        if ci < len(row):
                            period_col_map[ci] = f"{sheet_year}-{_MON_MAP[mon_key]}"
                    header_row_idx = ri - 1
                    break
        if not period_col_map:
            return out  # cannot parse this sheet

    # Step 2: walk data rows
    current_operator: str = ""

    for ri, row in enumerate(rows):
        if ri <= header_row_idx:
            continue
        if not row or all(c in (None, "", " ") for c in row):
            continue

        col0 = row[0] if len(row) > 0 else None
        col1 = row[1] if len(row) > 1 else None

        # Detect operator name in col 0
        if _is_operator_name(col0):
            # Could be operator-name-only row (Layout B) OR
            # operator + metric label in col 1 (Layout A).
            # If col 1 is a metric label, this is a Layout A data row.
            if isinstance(col1, str) and _classify_metric(col1):
                # Layout A: operator name + metric label on same row
                current_operator = str(col0).strip()
                metric_name = _classify_metric(str(col1).strip())
                _emit(row, period_col_map, current_operator, metric_name,
                      source_url, out)
                continue
            # Otherwise col 0 sets the operator for subsequent rows
            if not _classify_metric(str(col0)):
                current_operator = str(col0).strip()
            # The row itself might also carry Handle values (Layout B row 1)
            # Check if there are numeric values in period columns
            has_data = any(
                isinstance(row[ci], (int, float)) and not isinstance(row[ci], bool)
                for ci in period_col_map
                if ci < len(row)
            )
            if has_data and current_operator:
                # Treat this operator-name row as implicit Handle row (Layout B)
                _emit(row, period_col_map, current_operator, "handle",
                      source_url, out)
            continue

        # col 0 blank / None — look for metric label in col 1 (Layout B continuation)
        if col0 in (None, "") and isinstance(col1, str):
            metric_name = _classify_metric(col1.strip())
            if metric_name and current_operator:
                _emit(row, period_col_map, current_operator, metric_name,
                      source_url, out)
            continue

        # col 0 is a metric label directly (some sheets have no col-1 label)
        if isinstance(col0, str) and _classify_metric(col0) and current_operator:
            metric_name = _classify_metric(col0.strip())
            _emit(row, period_col_map, current_operator, metric_name,
                  source_url, out)
            continue

    return out


def _emit(
    row: list[Any],
    period_col_map: dict[int, str],
    operator: str,
    metric_name: str | None,
    source_url: str,
    out: list[MetricFact],
) -> None:
    if not metric_name or not operator:
        return
    for ci, period in period_col_map.items():
        if ci >= len(row):
            continue
        cents = _to_cents(row[ci])
        if cents is None:
            continue
        out.append(MetricFact(
            state="CO",
            operator=operator,
            vertical="sports-wagering",
            period=period,
            metric_name=metric_name,
            value_usd_cents=cents,
            source_url=source_url,
        ))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_co_xls(path: Path, source_url: str) -> list[MetricFact]:
    """Parse the Colorado Division of Gaming rolling sports-betting XLS.

    Parameters
    ----------
    path:
        Local path to the downloaded .xls file.
    source_url:
        Canonical URL of the file (stored verbatim in every MetricFact).

    Returns
    -------
    list[MetricFact]
        One fact per (operator, period, metric_name) triple found across all
        calendar-year sheets.  Metrics emitted: ``handle``, ``ggr``,
        ``tax_paid``.
    """
    out: list[MetricFact] = []
    for sheet_name, rows in _iter_sheets(path):
        # Only process year sheets; skip legend / footnote tabs
        if not _year_from_sheet_name(sheet_name):
            continue
        sheet_facts = _parse_sheet(sheet_name, rows, source_url)
        out.extend(sheet_facts)
    return out
